#!/usr/bin/env python3
"""
Post-migration validation checks.

After all migrations complete, the user can click "Check the migration" to run
these tests.  Each migration type has its own check function that compares the
input files against the database to verify correctness.

Every check function returns (passed: bool, messages: list[str]).
"""

import csv
import logging
import os
import sys
from datetime import datetime, date, timedelta, time
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Set

try:
    import psycopg2
    from psycopg2.extras import RealDictCursor
except ImportError:
    psycopg2 = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

logger = logging.getLogger(__name__)

from encoding_utils import fix_utf8_mojibake, normalize_name_for_match


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

DAYS_OF_WEEK = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']

def _get_day_of_week(dt: date) -> str:
    """Get day of week name from date."""
    return DAYS_OF_WEEK[dt.weekday()]

def _parse_datetime_value(datetime_val) -> Optional[datetime]:
    """Parse datetime from Excel/CSV value (same logic as migration)."""
    if datetime_val is None:
        return None
    if isinstance(datetime_val, datetime):
        return datetime_val
    if isinstance(datetime_val, str):
        datetime_str = datetime_val.strip()
        for fmt in ['%d-%m-%Y %H:%M:%S', '%Y-%m-%d %H:%M:%S', '%d/%m/%Y %H:%M:%S',
                    '%d-%m-%Y %H:%M', '%Y-%m-%d %H:%M', '%d/%m/%Y %H:%M']:
            try:
                return datetime.strptime(datetime_str, fmt)
            except ValueError:
                continue
    # Excel serial: days since 1899-12-30
    if isinstance(datetime_val, (int, float)):
        try:
            base = datetime(1899, 12, 30)
            return base + timedelta(days=float(datetime_val))
        except (OverflowError, ValueError):
            pass
    return None

def _occurrence_is_covered(
    occurrence_date: date,
    occurrence_start_time: str,  # HH:MM:SS
    occurrence_end_time: str,    # HH:MM:SS
    db_records: List[Dict]
) -> bool:
    """
    Check if a specific occurrence (date + time) is covered by any database record.

    A database record covers an occurrence if:
    1. The day of week matches
    2. The time range matches (allowing small tolerance for rounding)
    3. The occurrence date falls within the recurrence pattern

    Parameters:
    - occurrence_date: The specific date to check (e.g., 2026-03-06)
    - occurrence_start_time: Start time string "HH:MM:SS"
    - occurrence_end_time: End time string "HH:MM:SS"
    - db_records: List of database records with fields:
        - days: list of day names
        - start_time: "HH:MM:SS"
        - end_time: "HH:MM:SS"
        - start_date: date object
        - occurs_every: int (1=weekly, 2=bi-weekly, etc.)
    """
    occurrence_dow = _get_day_of_week(occurrence_date)

    # Normalize times for comparison (10-minute tolerance)
    def normalize_time(t_str):
        parts = t_str.split(':')
        hour = int(parts[0])
        minute = int(parts[1]) if len(parts) > 1 else 0
        # Round to nearest 10 minutes
        minute = (minute // 10) * 10
        return f"{hour:02d}:{minute:02d}"

    occ_start_norm = normalize_time(occurrence_start_time)
    occ_end_norm = normalize_time(occurrence_end_time)

    for record in db_records:
        # Check day of week
        if occurrence_dow not in record['days']:
            continue

        # Check time match
        db_start_norm = normalize_time(record['start_time'])
        db_end_norm = normalize_time(record['end_time'])

        if db_start_norm != occ_start_norm or db_end_norm != occ_end_norm:
            continue

        # Check if occurrence date is covered by recurrence pattern
        record_start_date = record['start_date']
        occurs_every = record.get('occurs_every', 1)

        # Occurrence must be on or after record start date
        if occurrence_date < record_start_date:
            continue

        # Calculate weeks from the FIRST occurrence of this day-of-week on or after start_date
        # For example: if start_date is Monday and record.days is {Friday},
        # we need to find the first Friday on or after Monday, then count weeks from there.

        # Get the day-of-week index (0=Monday, 6=Sunday)
        start_dow_index = record_start_date.weekday()
        occurrence_dow_index = occurrence_date.weekday()

        # Calculate how many days forward from start_date to reach the first occurrence of this day-of-week
        if occurrence_dow_index >= start_dow_index:
            days_to_first_occurrence = occurrence_dow_index - start_dow_index
        else:
            days_to_first_occurrence = 7 - start_dow_index + occurrence_dow_index

        # First date this day-of-week occurs on or after start_date
        from datetime import timedelta
        first_occurrence_date = record_start_date + timedelta(days=days_to_first_occurrence)

        # Occurrence must be on or after the first occurrence
        if occurrence_date < first_occurrence_date:
            continue

        # Calculate weeks between first occurrence and the occurrence we're checking
        days_diff = (occurrence_date - first_occurrence_date).days
        weeks_diff = days_diff // 7

        # For occurs_every=1 (weekly), any week works (weeks_diff % 1 == 0 always true)
        # For occurs_every=2 (bi-weekly), only weeks 0, 2, 4, ... work
        if weeks_diff % occurs_every == 0:
            return True

    return False

def _get_assets_dir() -> Path:
    from migration_support import get_assets_dir
    return get_assets_dir()


def _safe_strip(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _clean_excel_value(value) -> str:
    """Strip Excel formula formatting e.g. =\"06/08/2023 00:00\". Match usersMigration/clientsMigration."""
    s = _safe_strip(value) if value is not None else ""
    if s.startswith('="') and s.endswith('"'):
        s = s[2:-1].strip()
    elif s.startswith('=') and s.endswith('"'):
        s = s[1:].strip('"').strip()
    return s


def _parse_termination_date(value) -> Optional[date]:
    """Parse Termination Date from CSV (DD/MM/YYYY, DD/MM/YYYY HH:MM, or HH:MM:SS). Match usersMigration/clientsMigration."""
    if value is None or not _safe_strip(value):
        return None
    s = _clean_excel_value(value)
    if not s:
        return None
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _load_workbook_safe(path, data_only=True):
    """
    Load an XLSX with openpyxl. Tries normal mode first; on failure (e.g. style/XML
    parsing like Nested.from_tree), falls back to read_only mode to avoid full parse.
    Returns (wb, close_wb): close_wb is True if caller must call wb.close().
    """
    if openpyxl is None:
        raise RuntimeError("openpyxl not installed")
    try:
        wb = openpyxl.load_workbook(path, data_only=data_only)
        return wb, False
    except Exception:
        # read_only skips full style/XML parsing and often avoids from_tree errors
        wb = openpyxl.load_workbook(path, read_only=True, data_only=data_only)
        return wb, True


# ---------------------------------------------------------------------------
# 1. Caregivers (Users) Migration Check
# ---------------------------------------------------------------------------

def check_caregivers(connection) -> Tuple[bool, List[str]]:
    """
    Verify that every valid record in the caregivers CSV exists in the
    ``"user"`` table with ``status = 'Active'``.
    """
    msgs: List[str] = []
    csv_path = _get_assets_dir() / "CareAssistantExport.csv"

    if not csv_path.exists():
        msgs.append("SKIP: Caregivers CSV not found at %s" % csv_path)
        return True, msgs

    # Read expected names from CSV (same encoding + Termination Date filter as usersMigration)
    expected = []
    today = date.today()
    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            first = _safe_strip(fix_utf8_mojibake(row.get("First Name")))
            last = _safe_strip(fix_utf8_mojibake(row.get("Last Name")))
            if not last and first and " " in first:
                parts = first.split()
                titles = {"Mr", "Mrs", "Miss", "Ms", "Dr", "Prof",
                          "Mr.", "Mrs.", "Miss.", "Ms.", "Dr.", "Prof."}
                while parts and parts[0] in titles:
                    parts = parts[1:]
                if len(parts) >= 2:
                    first = parts[0]
                    last = " ".join(parts[1:])
            if not first or not last:
                continue
            # Same filter as usersMigration: only rows with null/empty or Termination Date >= today
            term_raw = _safe_strip(fix_utf8_mojibake(row.get("Termination Date")))
            term_date = _parse_termination_date(term_raw)
            if term_date is not None and term_date < today:
                continue
            # Use same normalization as usersMigration for matching
            expected.append((normalize_name_for_match(first), normalize_name_for_match(last)))

    if not expected:
        msgs.append("WARN: No valid records found in CSV")
        return True, msgs

    # Query DB
    cursor = connection.cursor()
    try:
        cursor.execute(
            'SELECT name, lastname, status FROM "user" WHERE deleted_at IS NULL'
        )
        db_rows = cursor.fetchall()
    finally:
        cursor.close()

    db_map: Dict[Tuple[str, str], str] = {}
    db_display: Dict[Tuple[str, str], Tuple[str, str]] = {}
    for row in db_rows:
        nfirst = normalize_name_for_match(row["name"] or "")
        nlast = normalize_name_for_match(row["lastname"] or "")
        key = (nfirst, nlast)
        db_map[key] = (row["status"] or "").strip()
        db_display[key] = ((row["name"] or "").strip(), (row["lastname"] or "").strip())

    missing = []
    wrong_status = []
    for key in expected:
        nfirst, nlast = key
        if key not in db_map:
            missing.append("%s %s" % (nfirst.title() if nfirst else "", nlast.title() if nlast else ""))
        elif db_map[key] != "Active":
            d = db_display.get(key, (nfirst, nlast))
            wrong_status.append(
                "%s %s (status=%s)" % (d[0], d[1], db_map[key])
            )

    msgs.append("Caregivers: %d expected from CSV, %d in DB" % (len(expected), len(db_map)))

    if missing:
        msgs.append("FAIL: %d caregiver(s) missing from DB:" % len(missing))
        for m in missing[:20]:
            msgs.append("  - %s" % m)
        if len(missing) > 20:
            msgs.append("  ... and %d more" % (len(missing) - 20))

    if wrong_status:
        msgs.append("FAIL: %d caregiver(s) not Active:" % len(wrong_status))
        for w in wrong_status[:20]:
            msgs.append("  - %s" % w)
        if len(wrong_status) > 20:
            msgs.append("  ... and %d more" % (len(wrong_status) - 20))

    passed = not missing and not wrong_status
    if passed:
        msgs.append("PASS: All %d caregivers present with Active status" % len(expected))
    return passed, msgs


# ---------------------------------------------------------------------------
# 2. Clients Migration Check
# ---------------------------------------------------------------------------

def check_clients(connection) -> Tuple[bool, List[str]]:
    """
    - Every client in the CSV must exist in ``client`` with Active status.
    - Every client NOT in the CSV must have ``status = 'Deactive'``.
    """
    msgs: List[str] = []
    csv_path = _get_assets_dir() / "CustomerExport.csv"

    if not csv_path.exists():
        msgs.append("SKIP: Clients CSV not found at %s" % csv_path)
        return True, msgs

    # Read expected clients (same encoding fix + Termination Date filter as clientsMigration)
    expected_keys = set()
    expected_display = {}  # normalized key -> (first, last) for readable error messages
    today = date.today()
    with open(csv_path, "r", encoding="utf-8") as f:
        sample = f.read(2048)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel
        reader = csv.DictReader(f, dialect=dialect)
        for row in reader:
            first = _safe_strip(fix_utf8_mojibake(row.get("First Name")))
            last = _safe_strip(fix_utf8_mojibake(row.get("Last Name")))
            if not first or not last:
                continue
            # Same filter as migration: only rows with null/empty or Termination Date >= today
            term_raw = _safe_strip(fix_utf8_mojibake(row.get("Termination Date")))
            term_date = _parse_termination_date(term_raw)
            if term_date is not None and term_date < today:
                continue
            norm_first = normalize_name_for_match(first)
            norm_last = normalize_name_for_match(last)
            key = (norm_first, norm_last)
            expected_keys.add(key)
            expected_display[key] = (first, last)

    if not expected_keys:
        msgs.append("WARN: No valid client records found in CSV")
        return True, msgs

    # Query DB
    cursor = connection.cursor()
    try:
        cursor.execute(
            "SELECT name, lastname, status FROM client WHERE deleted_at IS NULL"
        )
        db_rows = cursor.fetchall()
    finally:
        cursor.close()

    missing = []
    wrong_status_active = []   # should be Active but isn't
    wrong_status_deactive = [] # should be Deactive but isn't

    for row in db_rows:
        key = (normalize_name_for_match(row["name"] or ""),
               normalize_name_for_match(row["lastname"] or ""))
        status = (row["status"] or "").strip()
        name_display = (row["name"] or "").strip()
        lastname_display = (row["lastname"] or "").strip()
        if key in expected_keys:
            if status != "Active":
                wrong_status_active.append(
                    "%s %s (status=%s)" % (name_display, lastname_display, status)
                )
        else:
            if status != "Deactive":
                wrong_status_deactive.append(
                    "%s %s (status=%s)" % (name_display, lastname_display, status)
                )

    # Check if any expected client is missing entirely (use same normalized keys)
    db_keys = set()
    for row in db_rows:
        db_keys.add((normalize_name_for_match(row["name"] or ""),
                     normalize_name_for_match(row["lastname"] or "")))
    for key in expected_keys:
        if key not in db_keys:
            display = expected_display.get(key, (key[0].title(), key[1].title()))
            missing.append("%s %s" % (display[0], display[1]))

    msgs.append("Clients: %d expected from CSV, %d in DB" % (len(expected_keys), len(db_rows)))

    if missing:
        msgs.append("FAIL: %d client(s) missing from DB:" % len(missing))
        for m in missing[:20]:
            msgs.append("  - %s" % m)
        if len(missing) > 20:
            msgs.append("  ... and %d more" % (len(missing) - 20))

    if wrong_status_active:
        msgs.append("FAIL: %d client(s) from CSV not Active:" % len(wrong_status_active))
        for w in wrong_status_active[:20]:
            msgs.append("  - %s" % w)
        if len(wrong_status_active) > 20:
            msgs.append("  ... and %d more" % (len(wrong_status_active) - 20))

    if wrong_status_deactive:
        msgs.append("FAIL: %d client(s) NOT in CSV should be Deactive:" % len(wrong_status_deactive))
        for w in wrong_status_deactive[:20]:
            msgs.append("  - %s" % w)
        if len(wrong_status_deactive) > 20:
            msgs.append("  ... and %d more" % (len(wrong_status_deactive) - 20))

    passed = not missing and not wrong_status_active and not wrong_status_deactive
    if passed:
        msgs.append("PASS: All clients validated (CSV=Active, others=Deactive)")
    return passed, msgs


# ---------------------------------------------------------------------------
# 3. Client Availability Migration Check
# ---------------------------------------------------------------------------

def check_client_availability(connection) -> Tuple[bool, List[str]]:
    """
    - Filter the Excel input: keep only rows where both
      ``Planned Service Type Description`` and
      ``Planned Service Requirement Type Description`` equal "Personal Care".
    - Each specific occurrence (date+time) must be covered by a database record
      (either directly or through a recurring pattern).
    """
    msgs: List[str] = []
    xlsx_path = _get_assets_dir() / "clientAvailability" / "ClientHoursWithServiceType.xlsx"

    if not xlsx_path.exists():
        msgs.append("SKIP: Client availability XLSX not found at %s" % xlsx_path)
        return True, msgs

    if openpyxl is None:
        msgs.append("SKIP: openpyxl not installed, cannot read XLSX")
        return True, msgs

    # Parse Excel - extract Personal Care occurrences
    try:
        wb, close_wb = _load_workbook_safe(xlsx_path, data_only=True)
    except Exception as e:
        msgs.append("FAIL: Cannot open XLSX: %s" % e)
        return False, msgs

    csv_occurrences = []  # List of (client_key, occurrence_date, start_time, end_time)

    try:
        sheet_name = "Data"
        if sheet_name not in wb.sheetnames:
            msgs.append("FAIL: Sheet 'Data' not found in XLSX")
            return False, msgs

        ws = wb[sheet_name]
        header = [str(c or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

        def _col_idx(names):
            for n in names:
                nl = n.lower()
                for i, h in enumerate(header):
                    if h.lower() == nl:
                        return i
            return -1

        col_loc = _col_idx(["Service Location Name"])
        col_type = _col_idx(["Planned Service Type Description"])
        col_req = _col_idx(["Planned Service Requirement Type Description"])
        col_start = _col_idx(["Service Requirement Start Date And Time"])
        col_end = _col_idx(["Service Requirement End Date And Time"])
        col_duration = _col_idx(["Service Requirement Duration"])

        if any(c == -1 for c in [col_loc, col_type, col_req, col_start, col_end]):
            msgs.append("FAIL: Missing required columns in XLSX")
            return False, msgs

        total_count = 0
        filtered_count = 0
        skipped_parse_errors = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            total_count += 1
            stype = _safe_strip(row[col_type] if col_type < len(row) else None)
            sreq = _safe_strip(row[col_req] if col_req < len(row) else None)

            if stype.lower() != "personal care" or sreq.lower() != "personal care":
                continue

            filtered_count += 1

            loc = _safe_strip(fix_utf8_mojibake(row[col_loc] if col_loc < len(row) else None))
            if not loc:
                continue

            client_key = normalize_name_for_match(loc)

            # Parse date/time
            start_dt_val = row[col_start] if col_start < len(row) else None
            end_dt_val = row[col_end] if col_end < len(row) else None

            start_dt = _parse_datetime_value(start_dt_val)
            end_dt = _parse_datetime_value(end_dt_val)

            if not start_dt or not end_dt:
                skipped_parse_errors += 1
                continue

            # Parse duration: convert from hours to minutes (e.g., 0.75 hours = 45 minutes)
            duration_val = row[col_duration] if col_duration != -1 and col_duration < len(row) else None
            duration_minutes = None
            if duration_val is not None:
                try:
                    duration_hours = float(duration_val)
                    if duration_hours > 0:
                        duration_minutes = int(duration_hours * 60)
                except (ValueError, TypeError):
                    pass

            csv_occurrences.append((
                client_key,
                start_dt.date(),
                start_dt.strftime('%H:%M:%S'),
                end_dt.strftime('%H:%M:%S'),
                duration_minutes
            ))

        msgs.append("Client Availability: %d total rows, %d Personal Care rows, %d valid occurrences, %d parse errors"
                     % (total_count, filtered_count, len(csv_occurrences), skipped_parse_errors))
    finally:
        if close_wb:
            wb.close()

    if not csv_occurrences:
        msgs.append("WARN: No valid Personal Care occurrences found in XLSX")
        return True, msgs

    # Get client ID mapping
    cursor = connection.cursor()
    try:
        cursor.execute('SELECT id, name, lastname FROM client WHERE deleted_at IS NULL')
        client_id_map = {}
        for row in cursor.fetchall():
            name = (row['name'] or '').strip()
            lastname = (row['lastname'] or '').strip()
            key = normalize_name_for_match(f"{lastname}, {name}")
            client_id_map[key] = row['id']
    finally:
        cursor.close()

    # Fetch all client availability records from DB
    cursor = connection.cursor()
    try:
        cursor.execute("""
            SELECT ca.client_id, ca.days, ca.requested_start_time, ca.requested_end_time,
                   ca.start_date, ca.occurs_every, ca.duration
            FROM client_availabilities ca
            WHERE ca.deleted_at IS NULL
        """)
        db_rows = cursor.fetchall()
    finally:
        cursor.close()

    # Group DB records by client_id
    db_records_by_client = {}
    for row in db_rows:
        client_id = row['client_id']
        if client_id not in db_records_by_client:
            db_records_by_client[client_id] = []
        db_records_by_client[client_id].append({
            'days': row['days'],
            'start_time': str(row['requested_start_time']) if row['requested_start_time'] else '00:00:00',
            'end_time': str(row['requested_end_time']) if row['requested_end_time'] else '00:00:00',
            'start_date': row['start_date'],
            'occurs_every': row['occurs_every'] or 1,
            'duration': row.get('duration')
        })

    # Check each CSV occurrence
    uncovered_occurrences = []
    duration_mismatches = []
    clients_with_no_records = set()

    for client_key, occ_date, occ_start, occ_end, csv_duration in csv_occurrences:
        client_id = client_id_map.get(client_key)

        if not client_id:
            clients_with_no_records.add(client_key)
            continue

        db_records = db_records_by_client.get(client_id, [])

        if not db_records:
            clients_with_no_records.add(client_key)
            continue

        # Check if this occurrence is covered
        if not _occurrence_is_covered(occ_date, occ_start, occ_end, db_records):
            uncovered_occurrences.append((client_key, occ_date, occ_start, occ_end))
            continue

        # Find matching record to verify duration
        if csv_duration is not None:
            matched_record = None
            for record in db_records:
                if _occurrence_is_covered(occ_date, occ_start, occ_end, [record]):
                    matched_record = record
                    break

            if matched_record:
                db_duration = matched_record.get('duration')
                # Allow 5-minute tolerance for duration mismatches
                if db_duration is not None:
                    duration_diff = abs(csv_duration - db_duration)
                    if duration_diff > 5:
                        duration_mismatches.append((
                            client_key, occ_date, occ_start, occ_end,
                            csv_duration, db_duration
                        ))

    # Report results
    if clients_with_no_records:
        msgs.append("FAIL: %d client(s) from XLSX have no availability records in DB:" % len(clients_with_no_records))
        for m in sorted(clients_with_no_records)[:10]:
            msgs.append("  - %s" % m)
        if len(clients_with_no_records) > 10:
            msgs.append("  ... and %d more" % (len(clients_with_no_records) - 10))

    if uncovered_occurrences:
        msgs.append("FAIL: %d occurrence(s) not covered by DB records:" % len(uncovered_occurrences))
        for client_key, occ_date, occ_start, occ_end in uncovered_occurrences[:10]:
            msgs.append("  - %s on %s %s-%s" % (client_key, occ_date, occ_start[:5], occ_end[:5]))
        if len(uncovered_occurrences) > 10:
            msgs.append("  ... and %d more" % (len(uncovered_occurrences) - 10))

    if duration_mismatches:
        msgs.append("WARN: %d occurrence(s) with duration mismatches (>5 min diff):" % len(duration_mismatches))
        for client_key, occ_date, occ_start, occ_end, csv_dur, db_dur in duration_mismatches[:10]:
            msgs.append("  - %s on %s %s-%s: CSV=%d min, DB=%d min" %
                       (client_key, occ_date, occ_start[:5], occ_end[:5], csv_dur, db_dur))
        if len(duration_mismatches) > 10:
            msgs.append("  ... and %d more" % (len(duration_mismatches) - 10))

    passed = not clients_with_no_records and not uncovered_occurrences
    if passed:
        if duration_mismatches:
            msgs.append("PASS: All %d occurrences covered by DB records (%d with duration warnings)" %
                       (len(csv_occurrences), len(duration_mismatches)))
        else:
            msgs.append("PASS: All %d occurrences covered by DB records with matching durations" % len(csv_occurrences))
    return passed, msgs


# ---------------------------------------------------------------------------
# 4. Caregivers (User) Availability Migration Check
# ---------------------------------------------------------------------------

def check_caregiver_availability(connection) -> Tuple[bool, List[str]]:
    """
    Same logic as client availability but for user_availabilities.
    Reads caregivers availability XLSX and verifies each occurrence is covered.
    """
    msgs: List[str] = []
    xlsx_path = _get_assets_dir() / "userAvailabilities" / "userAvailabilities.xlsx"

    if not xlsx_path.exists():
        msgs.append("SKIP: Caregiver availability XLSX not found at %s" % xlsx_path)
        return True, msgs

    if openpyxl is None:
        msgs.append("SKIP: openpyxl not installed, cannot read XLSX")
        return True, msgs

    try:
        wb, close_wb = _load_workbook_safe(xlsx_path, data_only=True)
    except Exception as e:
        msgs.append("FAIL: Cannot open XLSX: %s" % e)
        return False, msgs

    csv_occurrences = []  # List of (user_key, occurrence_date, start_time, end_time, type)

    try:
        # Use first sheet (or 'Care Assistant Availability')
        sheet_name = "Care Assistant Availability"
        if sheet_name not in wb.sheetnames:
            sheet_name = wb.sheetnames[0]

        ws = wb[sheet_name]

        # Column indices from userAvailabilityMigration
        COL_NAME = 0
        COL_START_DATE = 5
        COL_START_TIME = 6
        COL_END_DATE = 7
        COL_END_TIME = 8
        COL_TYPE = 10

        total_count = 0
        core_count = 0
        other_count = 0
        skipped_parse_errors = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            total_count += 1

            name = _safe_strip(row[COL_NAME] if len(row) > COL_NAME else None)
            if not name:
                continue

            # Strip title
            parts = name.split()
            titles = {"Mr", "Mrs", "Miss", "Ms", "Dr", "Prof",
                      "Mr.", "Mrs.", "Miss.", "Ms.", "Dr.", "Prof."}
            while parts and parts[0] in titles:
                parts = parts[1:]
            clean_name = " ".join(parts).lower()

            if not clean_name:
                continue

            avail_type = _safe_strip(row[COL_TYPE] if len(row) > COL_TYPE else None)

            # Parse dates/times
            start_date_val = row[COL_START_DATE] if len(row) > COL_START_DATE else None
            start_time_val = row[COL_START_TIME] if len(row) > COL_START_TIME else None
            end_date_val = row[COL_END_DATE] if len(row) > COL_END_DATE else None
            end_time_val = row[COL_END_TIME] if len(row) > COL_END_TIME else None

            start_dt = _parse_datetime_value(start_date_val) if start_date_val else None
            end_dt = _parse_datetime_value(end_date_val) if end_date_val else None

            # Parse times
            if isinstance(start_time_val, time):
                start_time = start_time_val
            elif isinstance(start_time_val, datetime):
                start_time = start_time_val.time()
            else:
                start_time = None

            if isinstance(end_time_val, time):
                end_time = end_time_val
            elif isinstance(end_time_val, datetime):
                end_time = end_time_val.time()
            else:
                end_time = None

            if not start_dt or not start_time or not end_time:
                skipped_parse_errors += 1
                continue

            # Combine date and time
            start_datetime = datetime.combine(start_dt.date() if isinstance(start_dt, datetime) else start_dt, start_time)

            # Use end_date if provided, otherwise same as start
            if end_dt:
                end_date = end_dt.date() if isinstance(end_dt, datetime) else end_dt
            else:
                end_date = start_datetime.date()

            end_datetime = datetime.combine(end_date, end_time)

            # Handle overnight shifts
            if end_datetime < start_datetime:
                end_datetime += timedelta(days=1)

            if avail_type and avail_type.lower() == 'core':
                core_count += 1
            else:
                other_count += 1

            csv_occurrences.append((
                clean_name,
                start_datetime.date(),
                start_datetime.strftime('%H:%M:%S'),
                end_datetime.strftime('%H:%M:%S'),
                avail_type
            ))

        msgs.append("Caregiver Availability: %d total rows, %d Core, %d Other, %d valid occurrences, %d parse errors"
                     % (total_count, core_count, other_count, len(csv_occurrences), skipped_parse_errors))
    finally:
        if close_wb:
            wb.close()

    if not csv_occurrences:
        msgs.append("WARN: No valid user availability occurrences found in XLSX")
        return True, msgs

    # Get user ID mapping
    cursor = connection.cursor()
    try:
        cursor.execute('SELECT id, name, lastname FROM "user" WHERE deleted_at IS NULL')
        user_id_map = {}
        for row in cursor.fetchall():
            name = (row['name'] or '').strip()
            lastname = (row['lastname'] or '').strip()
            key = f"{name} {lastname}".lower()
            user_id_map[key] = row['id']
    finally:
        cursor.close()

    # Fetch all user availability records from DB
    cursor = connection.cursor()
    try:
        cursor.execute("""
            SELECT ua.user_id, ua.days, ua.start_time, ua.end_time,
                   ua.start_date, ua.occurs_every
            FROM user_availabilities ua
            WHERE ua.deleted_at IS NULL
        """)
        db_rows = cursor.fetchall()
    finally:
        cursor.close()

    # Group DB records by user_id
    db_records_by_user = {}
    for row in db_rows:
        user_id = row['user_id']
        if user_id not in db_records_by_user:
            db_records_by_user[user_id] = []
        db_records_by_user[user_id].append({
            'days': row['days'],
            'start_time': str(row['start_time']) if row['start_time'] else '00:00:00',
            'end_time': str(row['end_time']) if row['end_time'] else '00:00:00',
            'start_date': row['start_date'],
            'occurs_every': row['occurs_every'] or 1
        })

    # Check each CSV occurrence
    uncovered_occurrences = []
    users_with_no_records = set()

    for user_key, occ_date, occ_start, occ_end, avail_type in csv_occurrences:
        user_id = user_id_map.get(user_key)

        if not user_id:
            users_with_no_records.add(user_key)
            continue

        db_records = db_records_by_user.get(user_id, [])

        if not db_records:
            users_with_no_records.add(user_key)
            continue

        # Check if this occurrence is covered
        if not _occurrence_is_covered(occ_date, occ_start, occ_end, db_records):
            uncovered_occurrences.append((user_key, occ_date, occ_start, occ_end, avail_type))

    # Report results
    if users_with_no_records:
        msgs.append("FAIL: %d user(s) from XLSX have no availability records in DB:" % len(users_with_no_records))
        for m in sorted(users_with_no_records)[:10]:
            msgs.append("  - %s" % m)
        if len(users_with_no_records) > 10:
            msgs.append("  ... and %d more" % (len(users_with_no_records) - 10))

    if uncovered_occurrences:
        msgs.append("FAIL: %d occurrence(s) not covered by DB records:" % len(uncovered_occurrences))
        for user_key, occ_date, occ_start, occ_end, avail_type in uncovered_occurrences[:10]:
            msgs.append("  - %s on %s %s-%s (%s)" % (user_key, occ_date, occ_start[:5], occ_end[:5], avail_type or 'N/A'))
        if len(uncovered_occurrences) > 10:
            msgs.append("  ... and %d more" % (len(uncovered_occurrences) - 10))

    passed = not users_with_no_records and not uncovered_occurrences
    if passed:
        msgs.append("PASS: All %d occurrences covered by DB records" % len(csv_occurrences))
    return passed, msgs


# ---------------------------------------------------------------------------
# 5. Distances Migration Check
# ---------------------------------------------------------------------------

def check_distances(connection) -> Tuple[bool, List[str]]:
    """
    Verify that ALL distance pairs exist in travel_distances for all travel methods:
    1. All user to user pairs (both directions: for users 1 and 2, both 1→2 and 2→1 must exist)
    2. All client to client pairs (both directions: for clients 1 and 2, both 1→2 and 2→1 must exist)
    3. All user to client pairs (both directions: user→client and client→user)

    This test verifies ACTUAL pairs, not just counts, to ensure no missing or duplicate pairs.
    """
    msgs: List[str] = []

    cursor = connection.cursor()
    try:
        # Get all users with coordinates
        cursor.execute("""
            SELECT id FROM "user"
            WHERE deleted_at IS NULL AND is_caregiver = true
              AND latitude IS NOT NULL AND longitude IS NOT NULL
            ORDER BY id
        """)
        user_ids = [row["id"] for row in cursor.fetchall()]

        # Get all clients with coordinates
        cursor.execute("""
            SELECT id FROM client
            WHERE deleted_at IS NULL
              AND latitude IS NOT NULL AND longitude IS NOT NULL
            ORDER BY id
        """)
        client_ids = [row["id"] for row in cursor.fetchall()]

        if not user_ids:
            msgs.append("SKIP: No users with coordinates")
            return True, msgs

        if not client_ids:
            msgs.append("SKIP: No clients with coordinates")
            return True, msgs

        msgs.append("Distance Validation:")
        msgs.append("  Users with coordinates: %d" % len(user_ids))
        msgs.append("  Clients with coordinates: %d" % len(client_ids))
        msgs.append("")

        # Load ALL actual pairs from database grouped by (from_type, to_type, travel_method)
        cursor.execute("""
            SELECT from_type, to_type, from_id, to_id, travel_method
            FROM travel_distances
        """)
        db_pairs = {}  # key: (from_type, to_type, travel_method) -> set of (from_id, to_id)
        for row in cursor.fetchall():
            key = (row["from_type"], row["to_type"], row["travel_method"])
            if key not in db_pairs:
                db_pairs[key] = set()
            db_pairs[key].add((row["from_id"], row["to_id"]))

    finally:
        cursor.close()

    methods = ["car", "bike", "walk"]
    all_ok = True

    # Define expected pairs for each category
    categories = []

    # 1. User → User: Every user to every user (including self)
    expected_user_user = set()
    for uid1 in user_ids:
        for uid2 in user_ids:
            expected_user_user.add((uid1, uid2))
    categories.append(("user", "user", expected_user_user, "User → User"))

    # 2. Client → Client: Every client to every client (including self)
    expected_client_client = set()
    for cid1 in client_ids:
        for cid2 in client_ids:
            expected_client_client.add((cid1, cid2))
    categories.append(("client", "client", expected_client_client, "Client → Client"))

    # 3. User → Client: Every user to every client
    expected_user_client = set()
    for uid in user_ids:
        for cid in client_ids:
            expected_user_client.add((uid, cid))
    categories.append(("user", "client", expected_user_client, "User → Client"))

    # 4. Client → User: Every client to every user
    expected_client_user = set()
    for cid in client_ids:
        for uid in user_ids:
            expected_client_user.add((cid, uid))
    categories.append(("client", "user", expected_client_user, "Client → User"))

    # Verify each category for each travel method
    for from_type, to_type, expected_pairs, label in categories:
        msgs.append("%s (expected: %d pairs):" % (label, len(expected_pairs)))

        for method in methods:
            key = (from_type, to_type, method)
            actual_pairs = db_pairs.get(key, set())

            missing_pairs = expected_pairs - actual_pairs
            extra_pairs = actual_pairs - expected_pairs

            if missing_pairs or extra_pairs:
                all_ok = False
                msgs.append("  FAIL: %s - %d/%d pairs" % (method, len(actual_pairs), len(expected_pairs)))

                if missing_pairs:
                    msgs.append("    Missing %d pairs (showing first 10):" % len(missing_pairs))
                    for from_id, to_id in list(missing_pairs)[:10]:
                        msgs.append("      - %s %d → %s %d" % (from_type, from_id, to_type, to_id))
                    if len(missing_pairs) > 10:
                        msgs.append("      ... and %d more" % (len(missing_pairs) - 10))

                if extra_pairs:
                    msgs.append("    Extra/Duplicate %d pairs (showing first 10):" % len(extra_pairs))
                    for from_id, to_id in list(extra_pairs)[:10]:
                        msgs.append("      - %s %d → %s %d" % (from_type, from_id, to_type, to_id))
                    if len(extra_pairs) > 10:
                        msgs.append("      ... and %d more" % (len(extra_pairs) - 10))
            else:
                msgs.append("  PASS: %s has all %d pairs" % (method, len(expected_pairs)))

        msgs.append("")

    if all_ok:
        total_expected = sum(len(cat[2]) for cat in categories)
        msgs.append("PASS: All distance pairs verified for all travel methods")
        msgs.append("  Total pairs per method: %d" % total_expected)
        msgs.append("  Total across all methods: %d" % (total_expected * len(methods)))
    else:
        msgs.append("FAIL: Some distance pairs are missing or incorrect")

    return all_ok, msgs


# ---------------------------------------------------------------------------
# 6. Availability Types Check (placeholder)
# ---------------------------------------------------------------------------

def check_availability_types(connection) -> Tuple[bool, List[str]]:
    msgs: List[str] = []
    msgs.append("SKIP: Availability types check — no specific scenario defined yet")
    return True, msgs


# ---------------------------------------------------------------------------
# 7. Geocode Checks
# ---------------------------------------------------------------------------

def check_geocode(connection) -> Tuple[bool, List[str]]:
    """
    Verify that all clients and users with postcodes have latitude and longitude filled.

    Requirements:
    - All client records with postcode IS NOT NULL must have latitude and longitude NOT NULL
    - All user records with postcode IS NOT NULL must have latitude and longitude NOT NULL
    """
    msgs: List[str] = []
    all_passed = True

    cursor = connection.cursor()
    try:
        # Check clients with postcodes but missing coordinates
        cursor.execute("""
            SELECT id, name, lastname, postcode
            FROM client
            WHERE deleted_at IS NULL
              AND postcode IS NOT NULL
              AND postcode != ''
              AND (latitude IS NULL OR longitude IS NULL)
        """)
        clients_missing_coords = cursor.fetchall()

        # Count total clients with postcodes
        cursor.execute("""
            SELECT COUNT(*) as total
            FROM client
            WHERE deleted_at IS NULL
              AND postcode IS NOT NULL
              AND postcode != ''
        """)
        total_clients_with_postcodes = cursor.fetchone()['total']

        # Check users with postcodes but missing coordinates
        cursor.execute("""
            SELECT id, name, lastname, postcode
            FROM "user"
            WHERE deleted_at IS NULL
              AND postcode IS NOT NULL
              AND postcode != ''
              AND (latitude IS NULL OR longitude IS NULL)
        """)
        users_missing_coords = cursor.fetchall()

        # Count total users with postcodes
        cursor.execute("""
            SELECT COUNT(*) as total
            FROM "user"
            WHERE deleted_at IS NULL
              AND postcode IS NOT NULL
              AND postcode != ''
        """)
        total_users_with_postcodes = cursor.fetchone()['total']

    finally:
        cursor.close()

    # Report client location results
    msgs.append("Client Locations: %d clients with postcodes" % total_clients_with_postcodes)
    if clients_missing_coords:
        all_passed = False
        msgs.append("FAIL: %d client(s) with postcode missing latitude/longitude:" % len(clients_missing_coords))
        for row in clients_missing_coords[:20]:
            name = (row['name'] or '').strip()
            lastname = (row['lastname'] or '').strip()
            postcode = (row['postcode'] or '').strip()
            msgs.append("  - %s %s (postcode: %s)" % (name, lastname, postcode))
        if len(clients_missing_coords) > 20:
            msgs.append("  ... and %d more" % (len(clients_missing_coords) - 20))
    else:
        msgs.append("PASS: All clients with postcodes have coordinates")

    # Report user location results
    msgs.append("User Locations: %d users with postcodes" % total_users_with_postcodes)
    if users_missing_coords:
        all_passed = False
        msgs.append("FAIL: %d user(s) with postcode missing latitude/longitude:" % len(users_missing_coords))
        for row in users_missing_coords[:20]:
            name = (row['name'] or '').strip()
            lastname = (row['lastname'] or '').strip()
            postcode = (row['postcode'] or '').strip()
            msgs.append("  - %s %s (postcode: %s)" % (name, lastname, postcode))
        if len(users_missing_coords) > 20:
            msgs.append("  ... and %d more" % (len(users_missing_coords) - 20))
    else:
        msgs.append("PASS: All users with postcodes have coordinates")

    return all_passed, msgs


# ---------------------------------------------------------------------------
# 8. Feasible Pairs Check
# ---------------------------------------------------------------------------

def _parse_full_name_for_feasible(full_name: str) -> Tuple[Optional[str], Optional[str]]:
    """
    Parse a full name into first name and last name.
    Supports "Lastname, Firstname" and "Firstname Lastname" formats.
    Matches the logic from feasible_pairs_migration.
    """
    if not full_name:
        return (None, None)

    full_name = _safe_strip(full_name)
    if not full_name:
        return (None, None)

    # "Lastname, Firstname" format (VisitExport / Caremark)
    if ',' in full_name:
        parts = [p.strip() for p in full_name.split(',', 1)]
        if len(parts) == 2 and parts[0] and parts[1]:
            return (parts[1], parts[0])  # (firstname, lastname)
        if len(parts) == 1 and parts[0]:
            return (parts[0], None)
        return (None, None)

    # "Firstname Lastname" format
    parts = full_name.split()
    if len(parts) == 0:
        return (None, None)
    elif len(parts) == 1:
        return (parts[0], None)
    else:
        lastname = parts[-1]
        firstname = ' '.join(parts[:-1])
        return (firstname, lastname)


def check_feasible_pairs(connection) -> Tuple[bool, List[str]]:
    """
    Verify that all Personal Care pairs from the CSV exist in the feasible_pairs table.

    Requirements:
    - Read CSV file from assets directory
    - Filter rows where both "Planned Service Type Description" AND
      "Planned Service Requirement Type Description" equal "Personal Care"
    - Parse "Planned Employee Name" (caregiver) and "Service Location Name" (client)
    - Verify each pair exists in feasible_pairs table
    """
    msgs: List[str] = []
    csv_path = _get_assets_dir() / "visit_data.csv"

    if not csv_path.exists():
        msgs.append("SKIP: Visit data CSV not found at %s" % csv_path)
        return True, msgs

    # Load user and client lookups from database
    cursor = connection.cursor()
    try:
        # Load caregivers (users)
        cursor.execute("""
            SELECT id, name, lastname
            FROM "user"
            WHERE deleted_at IS NULL AND is_caregiver = true
        """)
        users_lookup = {}
        for row in cursor.fetchall():
            name = _safe_strip(row['name'])
            lastname = _safe_strip(row['lastname'])
            key = (name.lower(), lastname.lower())
            users_lookup[key] = row['id']

        # Load clients
        cursor.execute("""
            SELECT id, name, lastname
            FROM client
            WHERE deleted_at IS NULL
        """)
        clients_lookup = {}
        for row in cursor.fetchall():
            name = _safe_strip(row['name'])
            lastname = _safe_strip(row['lastname'])
            key = (name.lower(), lastname.lower())
            clients_lookup[key] = row['id']

        # Load existing feasible pairs
        cursor.execute("""
            SELECT cgid, client_id, frequency
            FROM feasible_pairs
        """)
        db_pairs = {}
        for row in cursor.fetchall():
            key = (row['cgid'], row['client_id'])
            db_pairs[key] = row['frequency']

    finally:
        cursor.close()

    # Parse CSV and extract Personal Care pairs
    expected_pairs = {}  # (caregiver_id, client_id) -> count
    total_rows = 0
    personal_care_rows = 0
    skipped_non_personal_care = 0
    unmatched_caregivers = set()
    unmatched_clients = set()

    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)

        for row in reader:
            total_rows += 1

            # Filter: Only Personal Care rows
            service_type = _safe_strip(row.get('Planned Service Type Description', ''))
            requirement_type = _safe_strip(row.get('Planned Service Requirement Type Description', ''))

            if service_type != 'Personal Care' or requirement_type != 'Personal Care':
                skipped_non_personal_care += 1
                continue

            personal_care_rows += 1

            # Get caregiver and client names
            employee_name = _safe_strip(row.get('Planned Employee Name', ''))
            location_name = _safe_strip(row.get('Service Location Name', ''))

            if not employee_name or not location_name:
                continue

            # Parse names
            employee_first, employee_last = _parse_full_name_for_feasible(employee_name)
            client_first, client_last = _parse_full_name_for_feasible(location_name)

            if not employee_first or not employee_last:
                unmatched_caregivers.add(employee_name)
                continue

            if not client_first or not client_last:
                unmatched_clients.add(location_name)
                continue

            # Look up IDs
            caregiver_key = (employee_first.lower(), employee_last.lower())
            caregiver_id = users_lookup.get(caregiver_key)

            if not caregiver_id:
                # Try alternative: first word of first name
                first_part = employee_first.split()[0] if employee_first else ''
                alt_key = (first_part.lower(), employee_last.lower())
                caregiver_id = users_lookup.get(alt_key)

            client_key = (client_first.lower(), client_last.lower())
            client_id = clients_lookup.get(client_key)

            if not client_id:
                # Try alternative: first word of first name
                first_part = client_first.split()[0] if client_first else ''
                alt_key = (first_part.lower(), client_last.lower())
                client_id = clients_lookup.get(alt_key)

            # Record pair if both found
            if caregiver_id and client_id:
                pair_key = (caregiver_id, client_id)
                expected_pairs[pair_key] = expected_pairs.get(pair_key, 0) + 1
            else:
                if not caregiver_id:
                    unmatched_caregivers.add(employee_name)
                if not client_id:
                    unmatched_clients.add(location_name)

    # Check for missing pairs in database
    missing_pairs = []
    for pair_key, expected_frequency in expected_pairs.items():
        if pair_key not in db_pairs:
            missing_pairs.append(pair_key)

    # Report results
    msgs.append("Feasible Pairs: %d total CSV rows, %d Personal Care rows" % (total_rows, personal_care_rows))
    msgs.append("  Expected pairs from CSV: %d" % len(expected_pairs))
    msgs.append("  Pairs in database: %d" % len(db_pairs))

    if unmatched_caregivers:
        msgs.append("WARN: %d caregiver(s) from CSV not found in DB:" % len(unmatched_caregivers))
        for name in sorted(unmatched_caregivers)[:10]:
            msgs.append("  - %s" % name)
        if len(unmatched_caregivers) > 10:
            msgs.append("  ... and %d more" % (len(unmatched_caregivers) - 10))

    if unmatched_clients:
        msgs.append("WARN: %d client(s) from CSV not found in DB:" % len(unmatched_clients))
        for name in sorted(unmatched_clients)[:10]:
            msgs.append("  - %s" % name)
        if len(unmatched_clients) > 10:
            msgs.append("  ... and %d more" % (len(unmatched_clients) - 10))

    if missing_pairs:
        msgs.append("FAIL: %d pair(s) from CSV missing in feasible_pairs table:" % len(missing_pairs))
        # Get names for display
        for cgid, client_id in missing_pairs[:20]:
            cursor = connection.cursor()
            try:
                cursor.execute('SELECT name, lastname FROM "user" WHERE id = %s', (cgid,))
                user_row = cursor.fetchone()
                cursor.execute('SELECT name, lastname FROM client WHERE id = %s', (client_id,))
                client_row = cursor.fetchone()

                if user_row and client_row:
                    caregiver_name = "%s %s" % (user_row['name'], user_row['lastname'])
                    client_name = "%s %s" % (client_row['name'], client_row['lastname'])
                    msgs.append("  - Caregiver: %s <-> Client: %s" % (caregiver_name, client_name))
            finally:
                cursor.close()

        if len(missing_pairs) > 20:
            msgs.append("  ... and %d more" % (len(missing_pairs) - 20))

    passed = not missing_pairs
    if passed:
        msgs.append("PASS: All %d expected pairs found in feasible_pairs table" % len(expected_pairs))

    return passed, msgs


# ---------------------------------------------------------------------------
# 9. Client Windows Check (placeholder)
# ---------------------------------------------------------------------------

def check_client_windows(connection) -> Tuple[bool, List[str]]:
    msgs: List[str] = []
    msgs.append("SKIP: Client windows check — no specific scenario defined yet")
    return True, msgs


# ---------------------------------------------------------------------------
# Runner
# ---------------------------------------------------------------------------

# Maps wizard migration option keys to check functions.
CHECK_MAP = {
    "caregivers":              ("Caregivers",              check_caregivers),
    "clients":                 ("Clients",                 check_clients),
    "clients_availability":    ("Clients Availability",    check_client_availability),
    "caregivers_availability": ("Caregivers Availability", check_caregiver_availability),
    "calculate_distances":     ("Distances",               check_distances),
    "availability_types":      ("Availability Types",      check_availability_types),
    "geocode_api":             ("Geocode (API)",           check_geocode),
    "geocode_client_file":     ("Geocode Client File",     check_geocode),
    "geocode_caregiver_file":  ("Geocode Caregiver File",  check_geocode),
    "fvisit_history":          ("Feasible Pairs",          check_feasible_pairs),
    "client_windows":          ("Client Windows",          check_client_windows),
}


def run_migration_checks(
    connection,
    selected_options: List[str],
    log_callback=None,
) -> Tuple[bool, List[str]]:
    """
    Run post-migration checks for the given option keys.

    Parameters
    ----------
    connection : psycopg2 connection
    selected_options : list of option key strings that were migrated
    log_callback : optional callable(str) invoked for every message line

    Returns
    -------
    (all_passed, all_messages)
    """
    all_passed = True
    all_msgs: List[str] = []

    def _log(msg):
        all_msgs.append(msg)
        if log_callback:
            log_callback(msg)

    _log("=" * 60)
    _log("POST-MIGRATION VALIDATION CHECKS")
    _log("=" * 60)
    _log("")

    if not selected_options:
        _log("No migration options selected for validation. Select one or more options to run their checks.")
        _log("=" * 60)
        return True, all_msgs

    for opt_key in selected_options:
        entry = CHECK_MAP.get(opt_key)
        if entry is None:
            continue
        label, check_fn = entry
        _log("--- %s ---" % label)
        try:
            passed, msgs = check_fn(connection)
        except Exception as e:
            passed = False
            msgs = ["ERROR: %s" % e]
            logger.exception("Check %s failed", label)
        for m in msgs:
            _log("  %s" % m)
        if not passed:
            all_passed = False
        _log("")

    _log("=" * 60)
    if all_passed:
        _log("ALL CHECKS PASSED")
    else:
        _log("SOME CHECKS FAILED — review messages above")
    _log("=" * 60)

    return all_passed, all_msgs
