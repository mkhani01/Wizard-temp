#!/usr/bin/env python3
"""
Validate today's roster – API validation against Client Hours + Caregivers Availability Excel files.

Compares expected data derived from the files with the live app API for a given date.
No database reads; API is the source of actual state.
"""

from __future__ import annotations

import logging
import re
from collections import Counter, defaultdict
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Optional, Sequence, Set, Tuple
from urllib.parse import urljoin

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import requests
except ImportError:
    requests = None

from encoding_utils import fix_utf8_mojibake, normalize_name_for_match

logger = logging.getLogger(__name__)

PERSONAL_CARE = "Personal Care"
DAYS_OF_WEEK = [
    "Monday",
    "Tuesday",
    "Wednesday",
    "Thursday",
    "Friday",
    "Saturday",
    "Sunday",
]
TITLES = [
    "Mr",
    "Mrs",
    "Miss",
    "Ms",
    "Dr",
    "Prof",
    "Mr.",
    "Mrs.",
    "Miss.",
    "Ms.",
    "Dr.",
    "Prof.",
]
LEAVE_COLOR = "#e7000b"
MINUTES_IN_DAY = 24 * 60

# Care Assistant Availability columns (0-based) — same as userAvailabilityMigration
COL_CARE_ASSISTANT_NAME = 0
COL_START_DATE = 5
COL_START_TIME = 6
COL_END_DATE = 7
COL_END_TIME = 8
COL_TYPE = 10

LogFn = Callable[[str], None]


class TestTodayError(Exception):
    pass


# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------


def _log(msgs: List[str], log_callback: Optional[LogFn], msg: str) -> None:
    msgs.append(msg)
    logger.info(msg)
    if log_callback:
        log_callback(msg)


def minutes_to_hhmm(minutes: int) -> str:
    m = max(0, min(MINUTES_IN_DAY, int(minutes)))
    return f"{m // 60:02d}:{m % 60:02d}"


def hhmm_to_minutes(hhmm: str) -> int:
    parts = str(hhmm).strip().split(":")
    h = int(parts[0])
    mi = int(parts[1]) if len(parts) > 1 else 0
    return h * 60 + mi


def datetime_to_minutes(dt: datetime) -> int:
    """Convert datetime to minutes-from-midnight, rounding to nearest minute.

    Excel serial floats often land a few microseconds short of an exact minute
    (e.g. 12:29:59.999997 instead of 12:30:00). Truncating would then report
    12:29; rounding matches what Excel displays and what the API stores.
    """
    total_seconds = (
        dt.hour * 3600
        + dt.minute * 60
        + dt.second
        + dt.microsecond / 1_000_000
    )
    return int(round(total_seconds / 60))


def strip_title(full_name: str) -> str:
    if not full_name:
        return ""
    parts = full_name.strip().split()
    while parts and parts[0] in TITLES:
        parts = parts[1:]
    return " ".join(parts)


def normalize_person_key(value: Optional[str]) -> str:
    """Normalize a person display name for matching (handles 'lastname, name')."""
    raw = normalize_name_for_match(fix_utf8_mojibake(value or ""))
    if not raw:
        return ""
    if "," in raw:
        last, first = raw.split(",", 1)
        return normalize_name_for_match(f"{first} {last}")
    return raw


def strip_name_annotations(value: Optional[str]) -> str:
    """
    Strip Excel location annotations that the API usually does not store in name/lastname.

    Examples:
      '*Ryan (C), Martin'  -> 'Ryan, Martin'
      'Keane (DS), Patrick' -> 'Keane, Patrick'
      'Bagnall(DU), George' -> 'Bagnall, George'
    """
    s = fix_utf8_mojibake(value or "").strip()
    if not s:
        return ""
    s = re.sub(r"^\*+\s*", "", s)
    s = re.sub(r"\([^)]*\)", " ", s)
    s = re.sub(r"\s*,\s*", ", ", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip(" ,")


def _name_forms(value: str) -> Set[str]:
    """Build space/comma alternate forms for one display string (no annotation stripping)."""
    keys: Set[str] = set()
    fixed = fix_utf8_mojibake(value or "")
    base = normalize_name_for_match(fixed)
    if base:
        keys.add(base)
        keys.add(normalize_person_key(fixed))
    if "," in (fixed or ""):
        last, first = fixed.split(",", 1)
        keys.add(normalize_name_for_match(f"{first.strip()} {last.strip()}"))
        keys.add(normalize_name_for_match(f"{last.strip()}, {first.strip()}"))
    else:
        parts = base.split()
        if len(parts) >= 2:
            keys.add(normalize_name_for_match(f"{parts[-1]}, {' '.join(parts[:-1])}"))
    return {k for k in keys if k}


def name_match_keys(value: Optional[str]) -> Set[str]:
    """
    Return alternate keys for a person name (space and comma forms).

    Also includes annotation-stripped forms so Excel names like 'Keane (DS), Patrick'
    match API 'Patrick Keane'.
    """
    fixed = fix_utf8_mojibake(value or "")
    keys = _name_forms(fixed)
    scrubbed = strip_name_annotations(fixed)
    if scrubbed and normalize_name_for_match(scrubbed) != normalize_name_for_match(fixed):
        keys |= _name_forms(scrubbed)
    return keys


MINUTE_MATCH_TOLERANCE = 1


def minutes_close(a: int, b: int, tolerance: int = MINUTE_MATCH_TOLERANCE) -> bool:
    """True when two minute-of-day values are within tolerance (API often truncates end by 1)."""
    return abs(int(a) - int(b)) <= int(tolerance)


def visit_matches_slot(visit: Dict[str, Any], row: Dict[str, Any]) -> bool:
    """True when API visit and Excel slot refer to the same client + window (±1 min)."""
    vkeys = name_match_keys(visit_client_name(visit))
    if not (vkeys & row["name_keys"]):
        return False
    if not minutes_close(int(visit.get("startMinute") or 0), int(row["start_minute"])):
        return False
    if not minutes_close(int(visit.get("endMinute") or 0), int(row["end_minute"])):
        return False
    return True


def parse_datetime_value(datetime_val) -> Optional[datetime]:
    if datetime_val is None:
        return None
    if isinstance(datetime_val, datetime):
        return datetime_val
    if isinstance(datetime_val, date) and not isinstance(datetime_val, datetime):
        return datetime.combine(datetime_val, time.min)
    if isinstance(datetime_val, str):
        datetime_str = datetime_val.strip()
        if not datetime_str:
            return None
        for fmt in (
            "%d-%m-%Y %H:%M:%S",
            "%Y-%m-%d %H:%M:%S",
            "%d/%m/%Y %H:%M:%S",
            "%d-%m-%Y %H:%M",
            "%Y-%m-%d %H:%M",
            "%d/%m/%Y %H:%M",
            "%Y-%m-%d",
            "%d-%m-%Y",
            "%d/%m/%Y",
        ):
            try:
                return datetime.strptime(datetime_str, fmt)
            except ValueError:
                continue
    if isinstance(datetime_val, (int, float)):
        try:
            return datetime(1899, 12, 30) + timedelta(days=float(datetime_val))
        except (OverflowError, ValueError):
            pass
    return None


def parse_date_value(date_val) -> Optional[date]:
    if date_val is None:
        return None
    if isinstance(date_val, datetime):
        return date_val.date()
    if isinstance(date_val, date):
        return date_val
    if isinstance(date_val, str):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(date_val.strip(), fmt).date()
            except ValueError:
                continue
    return None


def parse_time_value(time_val) -> Optional[time]:
    if time_val is None:
        return None
    if isinstance(time_val, time):
        return time_val
    if isinstance(time_val, datetime):
        return time_val.time()
    if isinstance(time_val, str):
        try:
            parts = time_val.strip().split(":")
            if len(parts) >= 2:
                return time(int(parts[0]), int(parts[1]), int(parts[2]) if len(parts) > 2 else 0)
        except (ValueError, IndexError):
            pass
    return None


def parse_target_date(value: Any) -> date:
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    raise TestTodayError(f"Invalid target date: {value!r} (expected YYYY-MM-DD)")


def resolve_start_end(req_start_val, req_end_val, act_start_val, act_end_val):
    req_start = parse_datetime_value(req_start_val)
    req_end = parse_datetime_value(req_end_val)
    act_start = parse_datetime_value(act_start_val)
    act_end = parse_datetime_value(act_end_val)
    start = req_start or act_start
    end = req_end or act_end
    if req_start and req_end:
        source = "requirement"
    elif act_start and act_end and not (req_start or req_end):
        source = "actual"
    else:
        source = "mixed"
    return start, end, source


def _col_idx(headers: Sequence[Any], names: Sequence[str]) -> int:
    lower_headers = [str(h or "").strip().lower() for h in headers]
    for name in names:
        nl = name.lower()
        for i, h in enumerate(lower_headers):
            if h == nl:
                return i
    return -1


def _cell_str(row, col: int) -> str:
    if col == -1 or col >= len(row) or row[col] is None:
        return ""
    return str(row[col]).strip()


def _cell_int(row, col: int, default: int = 1) -> int:
    if col == -1 or col >= len(row) or row[col] is None:
        return default
    try:
        return max(1, int(row[col]))
    except (TypeError, ValueError):
        return default


# ---------------------------------------------------------------------------
# Time-interval math (ported from server time-intervals.ts)
# ---------------------------------------------------------------------------

TimeRange = Tuple[str, str]  # (startTime HH:MM, endTime HH:MM)


def merge_day_intervals(ranges: Iterable[TimeRange]) -> List[TimeRange]:
    normalized = []
    for start_s, end_s in ranges:
        start = hhmm_to_minutes(start_s)
        end = hhmm_to_minutes(end_s)
        if end > start:
            normalized.append((start, end))
    normalized.sort(key=lambda x: x[0])
    if not normalized:
        return []
    merged = [list(normalized[0])]
    for start, end in normalized[1:]:
        last = merged[-1]
        if start <= last[1]:
            last[1] = max(last[1], end)
        else:
            merged.append([start, end])
    return [(minutes_to_hhmm(a), minutes_to_hhmm(b)) for a, b in merged]


def subtract_day_intervals(base: Iterable[TimeRange], subtract: Iterable[TimeRange]) -> List[TimeRange]:
    merged_base = [(hhmm_to_minutes(a), hhmm_to_minutes(b)) for a, b in merge_day_intervals(base)]
    merged_sub = [(hhmm_to_minutes(a), hhmm_to_minutes(b)) for a, b in merge_day_intervals(subtract)]
    result: List[Tuple[int, int]] = []
    for start, end in merged_base:
        segments = [(start, end)]
        for b_start, b_end in merged_sub:
            next_segs: List[Tuple[int, int]] = []
            for seg_start, seg_end in segments:
                if b_end <= seg_start or b_start >= seg_end:
                    next_segs.append((seg_start, seg_end))
                    continue
                if b_start > seg_start:
                    next_segs.append((seg_start, b_start))
                if b_end < seg_end:
                    next_segs.append((b_end, seg_end))
            segments = [(s, e) for s, e in next_segs if e > s]
        result.extend(segments)
    return [(minutes_to_hhmm(a), minutes_to_hhmm(b)) for a, b in result]


def union_day_intervals(*sets: Iterable[TimeRange]) -> List[TimeRange]:
    flat: List[TimeRange] = []
    for s in sets:
        flat.extend(list(s))
    return merge_day_intervals(flat)


def complement_day_intervals(ranges: Iterable[TimeRange]) -> List[TimeRange]:
    day_start, day_end = "00:00", "24:00"
    merged = merge_day_intervals(ranges)
    if not merged:
        return [(day_start, day_end)]
    result: List[TimeRange] = []
    cursor = day_start
    for start, end in merged:
        if cursor < start:
            result.append((cursor, start))
        if end > cursor:
            cursor = end
    if cursor < day_end:
        result.append((cursor, day_end))
    return result


# ---------------------------------------------------------------------------
# Multi-day window clip (ported from end-time-date.helpers.ts)
# ---------------------------------------------------------------------------


def _utc_midnight(d: date) -> datetime:
    return datetime(d.year, d.month, d.day)


def build_absolute_window(
    occurrence: date,
    start_t: time,
    end_t: time,
    offset_days: int,
) -> Tuple[datetime, datetime]:
    start_ms = _utc_midnight(occurrence) + timedelta(
        minutes=start_t.hour * 60 + start_t.minute
    )
    end_date = occurrence + timedelta(days=max(0, offset_days))
    end_ms = _utc_midnight(end_date) + timedelta(minutes=end_t.hour * 60 + end_t.minute)
    return start_ms, end_ms


def clip_absolute_window_to_date(
    absolute: Tuple[datetime, datetime],
    target: date,
) -> Optional[Tuple[int, int]]:
    day_start = _utc_midnight(target)
    day_end = day_start + timedelta(minutes=MINUTES_IN_DAY)
    start_ms = max(absolute[0], day_start)
    end_ms = min(absolute[1], day_end)
    if start_ms >= end_ms:
        return None
    start_minute = int((start_ms - day_start).total_seconds() // 60)
    end_minute = int((end_ms - day_start).total_seconds() // 60)
    return start_minute, end_minute


def cap_absolute_window_at_date(
    absolute: Tuple[datetime, datetime],
    cap: Optional[date],
) -> Tuple[datetime, datetime]:
    if cap is None:
        return absolute
    cap_end = _utc_midnight(cap) + timedelta(minutes=MINUTES_IN_DAY)
    return absolute[0], min(absolute[1], cap_end)


def is_recurring_week_active(start_date: date, candidate: date, occurs_every: int) -> bool:
    days_since = (candidate - start_date).days
    week_index = days_since // 7
    interval = occurs_every or 1
    return week_index % interval == 0


def resolve_slot_window_for_date(slot: Dict[str, Any], target: date) -> Optional[Tuple[int, int]]:
    """
    Resolve day-clipped minutes for a caregiver availability slot on target date.
    Mirrors resolveAvailabilityWindowForDate / findOccurrenceCoveringDate.
    """
    offset = max(0, int(slot.get("end_time_date_offset_days") or 0))
    start_t: time = slot["start_time"]
    end_t: time = slot["end_time"]
    slot_day = slot.get("day")

    candidates = [target - timedelta(days=back) for back in range(0, offset + 1)]
    for candidate in candidates:
        cand_day = DAYS_OF_WEEK[candidate.weekday()]
        if slot_day and cand_day != slot_day:
            continue

        if slot.get("is_temp"):
            from_d = slot.get("effective_date_from")
            to_d = slot.get("effective_date_to")
            if from_d is None:
                continue
            if candidate < from_d:
                continue
            if to_d is not None and candidate > to_d:
                continue
        else:
            start_date = slot.get("start_date")
            if start_date is None:
                continue
            if candidate < start_date:
                continue
            end_date = slot.get("end_date")
            if end_date is not None and candidate > end_date:
                continue
            if not is_recurring_week_active(
                start_date, candidate, slot.get("occurs_every") or 1
            ):
                continue

        occurrence = candidate
        absolute = build_absolute_window(occurrence, start_t, end_t, offset)
        if slot.get("is_temp"):
            absolute = cap_absolute_window_at_date(absolute, slot.get("effective_date_to"))
        clipped = clip_absolute_window_to_date(absolute, target)
        if clipped:
            return clipped
    return None


# ---------------------------------------------------------------------------
# API client
# ---------------------------------------------------------------------------


class ApiClient:
    def __init__(self, base_url: str, token: str, timeout: int = 60):
        if requests is None:
            raise TestTodayError("Missing package: requests (pip install requests)")
        self.base_url = base_url.rstrip("/") + "/"
        tok = token.strip()
        if tok.lower().startswith("bearer "):
            auth = tok
        else:
            auth = "Bearer " + tok
        self.session = requests.Session()
        self.session.headers.update(
            {
                "Authorization": auth,
                "Accept": "application/json",
            }
        )
        self.timeout = timeout

    def get(self, path: str, params: Optional[Dict[str, Any]] = None) -> Any:
        url = urljoin(self.base_url, path.lstrip("/"))
        try:
            resp = self.session.get(url, params=params or {}, timeout=self.timeout)
        except requests.RequestException as e:
            raise TestTodayError(f"API request failed: {path} ({e})") from e
        if resp.status_code >= 400:
            body = resp.text[:500]
            raise TestTodayError(
                f"API {resp.status_code} for {path}: {body}"
            )
        if not resp.content:
            return None
        try:
            payload = resp.json()
        except ValueError as e:
            raise TestTodayError(f"Invalid JSON from {path}") from e
        # NestJS ResponseInterceptor wraps as SuccessResponse: { status, data, ... }
        if (
            isinstance(payload, dict)
            and payload.get("status") == "success"
            and "data" in payload
        ):
            return payload["data"]
        return payload

    def get_paginated(
        self,
        path: str,
        params: Optional[Dict[str, Any]] = None,
        limit: int = 100,
        data_key: str = "data",
    ) -> List[Any]:
        params = dict(params or {})
        page = 1
        all_rows: List[Any] = []
        total: Optional[int] = None
        while True:
            params.update({"page": page, "limit": limit})
            payload = self.get(path, params=params)
            if payload is None:
                break
            if isinstance(payload, list):
                return payload
            rows = payload.get(data_key) if isinstance(payload, dict) else None
            if rows is None:
                # Some endpoints wrap differently
                rows = payload.get("items") if isinstance(payload, dict) else None
            # Nested paginated body: { data: [...], meta: {...} } already unwrapped
            # from SuccessResponse by get(); if data_key pointed at another object
            # with an inner list, unwrap one more level.
            if isinstance(rows, dict) and isinstance(rows.get(data_key), list):
                payload = rows
                rows = rows[data_key]
            if not isinstance(rows, list):
                raise TestTodayError(f"Unexpected paginated response from {path}")
            all_rows.extend(rows)
            meta_total = None
            if isinstance(payload, dict):
                meta_total = payload.get("total")
                if meta_total is None and isinstance(payload.get("meta"), dict):
                    meta_total = payload["meta"].get("total")
            if meta_total is not None:
                total = int(meta_total)
            if not rows:
                break
            if total is not None and len(all_rows) >= total:
                break
            if len(rows) < limit:
                break
            page += 1
            if page > 500:
                raise TestTodayError(f"Pagination safety stop for {path}")
        return all_rows


# ---------------------------------------------------------------------------
# Excel parsers
# ---------------------------------------------------------------------------


def load_client_hours_for_date(
    filepath: Path,
    target_date: date,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, int]]:
    """
    Parse Client Hours XLSX for target_date.
    Returns (unallocated_slots, cancelled_slots, stats).
    Slots are expanded by Count. Keys use normalized client name.
    """
    if openpyxl is None:
        raise TestTodayError("Missing package: openpyxl")
    if not filepath.exists():
        raise TestTodayError(f"Client Hours file not found: {filepath}")

    stats = {
        "total_rows": 0,
        "personal_care": 0,
        "on_target_date": 0,
        "unallocated": 0,
        "cancelled": 0,
        "skipped_not_personal_care": 0,
        "skipped_missing_datetime": 0,
        "skipped_wrong_date": 0,
        "skipped_empty_location": 0,
    }
    unallocated: List[Dict[str, Any]] = []
    cancelled: List[Dict[str, Any]] = []

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        if "Data" not in wb.sheetnames:
            raise TestTodayError("Sheet 'Data' not found in Client Hours workbook")
        ws = wb["Data"]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            raise TestTodayError("Client Hours workbook has no header row")

        col_loc = _col_idx(header, ["Service Location Name"])
        col_req_start = _col_idx(header, ["Service Requirement Start Date And Time"])
        col_req_end = _col_idx(header, ["Service Requirement End Date And Time"])
        col_act_start = _col_idx(header, ["Actual Start Date And Time"])
        col_act_end = _col_idx(header, ["Actual End Date And Time"])
        col_cancel = _col_idx(header, ["Cancellation Description"])
        col_count = _col_idx(header, ["Count"])
        col_svc_type = _col_idx(header, ["Planned Service Type Description"])
        col_svc_req = _col_idx(header, ["Planned Service Requirement Type Description"])

        if col_loc == -1:
            raise TestTodayError("Missing column: Service Location Name")
        if col_svc_type == -1 or col_svc_req == -1:
            raise TestTodayError(
                "Missing Planned Service Type / Requirement Type Description columns"
            )

        for row_num, row in enumerate(rows_iter, start=2):
            stats["total_rows"] += 1
            if not row:
                continue
            svc_type = _cell_str(row, col_svc_type)
            svc_req = _cell_str(row, col_svc_req)
            if svc_type != PERSONAL_CARE or svc_req != PERSONAL_CARE:
                stats["skipped_not_personal_care"] += 1
                continue
            stats["personal_care"] += 1

            raw_loc = row[col_loc] if col_loc < len(row) else None
            loc = fix_utf8_mojibake(raw_loc) if raw_loc is not None else None
            loc_str = str(loc).strip() if loc is not None else ""
            if not loc_str:
                stats["skipped_empty_location"] += 1
                continue

            req_start = row[col_req_start] if col_req_start != -1 and col_req_start < len(row) else None
            req_end = row[col_req_end] if col_req_end != -1 and col_req_end < len(row) else None
            act_start = row[col_act_start] if col_act_start != -1 and col_act_start < len(row) else None
            act_end = row[col_act_end] if col_act_end != -1 and col_act_end < len(row) else None
            start_dt, end_dt, source = resolve_start_end(req_start, req_end, act_start, act_end)
            if not start_dt or not end_dt:
                stats["skipped_missing_datetime"] += 1
                continue
            if start_dt.date() != target_date:
                stats["skipped_wrong_date"] += 1
                continue
            stats["on_target_date"] += 1

            cancel = _cell_str(row, col_cancel)
            count = _cell_int(row, col_count, default=1)
            start_m = datetime_to_minutes(start_dt)
            end_m = datetime_to_minutes(end_dt)
            # Overnight same-day: treat end next day minutes for display; visit windows are same-day minutes
            offset_days = (end_dt.date() - start_dt.date()).days
            if offset_days == 0 and end_m <= start_m:
                offset_days = 1
            if offset_days > 0:
                # Clip end to end of day for same-day roster comparison (API day-clips)
                end_m_day = MINUTES_IN_DAY if offset_days > 0 and end_dt.date() > target_date else end_m
                if end_dt.date() > target_date:
                    end_m = MINUTES_IN_DAY

            name_key = normalize_person_key(loc_str)
            item_base = {
                "row_num": row_num,
                "client_name": loc_str,
                "name_key": name_key,
                "name_keys": name_match_keys(loc_str),
                "start_minute": start_m,
                "end_minute": end_m,
                "time_source": source,
                "cancellation": cancel,
                "count": count,
            }
            is_cancel = bool(cancel) and cancel.lower() != "none"
            for slot in range(count):
                item = dict(item_base)
                item["slot"] = slot
                if is_cancel:
                    cancelled.append(item)
                    stats["cancelled"] += 1
                else:
                    unallocated.append(item)
                    stats["unallocated"] += 1
        return unallocated, cancelled, stats
    finally:
        wb.close()


def load_caregiver_availability_slots(
    filepath: Path,
    types_map: Dict[str, Dict[str, Any]],
) -> Tuple[List[Dict[str, Any]], List[str], Dict[str, int]]:
    """
    Parse caregivers availability XLSX into slot dicts (same rules as migration).
    types_map: name_lower -> {is_unavailability: bool, ...}
    """
    if openpyxl is None:
        raise TestTodayError("Missing package: openpyxl")
    if not filepath.exists():
        raise TestTodayError(f"Caregivers availability file not found: {filepath}")

    stats = {
        "total_rows": 0,
        "valid": 0,
        "skipped_empty_name": 0,
        "skipped_unknown_type": 0,
        "skipped_missing_datetime": 0,
        "skipped_bad_dates": 0,
    }
    slots: List[Dict[str, Any]] = []
    unmatched_types: List[str] = []

    # read_only avoids openpyxl crashes on Client Hours workbooks with pivot caches
    # (TypeError: Nested.from_tree() missing ...), and is enough for value iteration.
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        raise TestTodayError(
            f"Failed to open caregivers availability file {filepath.name!r}: {e}. "
            "Select an Availability Export XLSX (sheet 'Care Assistant Availability'), "
            "not the Client Hours file."
        ) from e

    try:
        preferred = "Care Assistant Availability"
        if preferred in wb.sheetnames:
            sheet_name = preferred
        elif wb.sheetnames:
            sheet_name = wb.sheetnames[0]
        else:
            raise TestTodayError(
                f"Caregivers availability file {filepath.name!r} has no sheets. "
                "Select an Availability Export XLSX."
            )
        ws = wb[sheet_name]

        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            raise TestTodayError(
                f"Caregivers availability file {filepath.name!r} has no header row."
            )
        header0 = str(header[0] or "").strip().lower() if len(header) > 0 else ""
        if sheet_name != preferred and header0 != "care assistant name":
            raise TestTodayError(
                f"File {filepath.name!r} does not look like a Caregivers Availability export "
                f"(sheets={list(wb.sheetnames)!r}). "
                "Select an Availability Export XLSX with sheet "
                "'Care Assistant Availability', not Client Hours."
            )

        for row_num, row in enumerate(rows_iter, start=2):
            stats["total_rows"] += 1
            care_name = row[COL_CARE_ASSISTANT_NAME] if len(row) > COL_CARE_ASSISTANT_NAME else None
            start_date_val = row[COL_START_DATE] if len(row) > COL_START_DATE else None
            end_date_val = row[COL_END_DATE] if len(row) > COL_END_DATE else None
            start_time_val = row[COL_START_TIME] if len(row) > COL_START_TIME else None
            end_time_val = row[COL_END_TIME] if len(row) > COL_END_TIME else None
            type_val = row[COL_TYPE] if len(row) > COL_TYPE else None

            if not care_name:
                stats["skipped_empty_name"] += 1
                continue

            type_str = str(type_val).strip() if type_val else ""
            type_key = type_str.lower()
            if type_key not in types_map:
                stats["skipped_unknown_type"] += 1
                unmatched_types.append(type_str)
                continue

            start_date = parse_date_value(start_date_val)
            end_date = parse_date_value(end_date_val)
            start_time = parse_time_value(start_time_val)
            end_time = parse_time_value(end_time_val)
            if not start_date or not start_time or not end_time:
                stats["skipped_missing_datetime"] += 1
                continue
            if not end_date:
                end_date = start_date
            if end_date < start_date:
                stats["skipped_bad_dates"] += 1
                continue

            offset_days = (end_date - start_date).days
            if offset_days == 0 and end_time <= start_time:
                offset_days = 1

            type_info = types_map[type_key]
            is_core = type_key == "core"
            cleaned = strip_title(str(care_name))
            name_key = normalize_person_key(cleaned)

            if is_core:
                target_start = start_date - timedelta(days=28)
                slots.append(
                    {
                        "user_name": str(care_name).strip(),
                        "name_key": name_key,
                        "name_keys": name_match_keys(cleaned) | name_match_keys(str(care_name)),
                        "day": DAYS_OF_WEEK[target_start.weekday()],
                        "start_time": start_time,
                        "end_time": end_time,
                        "end_time_date_offset_days": offset_days,
                        "is_temp": False,
                        "is_unavailability": bool(type_info.get("is_unavailability")),
                        "type_name": type_str,
                        "start_date": target_start,
                        "end_date": None,
                        "occurs_every": 4,
                        "effective_date_from": None,
                        "effective_date_to": None,
                        "source_row": row_num,
                    }
                )
            else:
                slots.append(
                    {
                        "user_name": str(care_name).strip(),
                        "name_key": name_key,
                        "name_keys": name_match_keys(cleaned) | name_match_keys(str(care_name)),
                        "day": DAYS_OF_WEEK[start_date.weekday()],
                        "start_time": start_time,
                        "end_time": end_time,
                        "end_time_date_offset_days": offset_days,
                        "is_temp": True,
                        "is_unavailability": bool(type_info.get("is_unavailability")),
                        "type_name": type_str,
                        "start_date": None,
                        "end_date": None,
                        "occurs_every": None,
                        "effective_date_from": start_date,
                        "effective_date_to": end_date,
                        "source_row": row_num,
                    }
                )
            stats["valid"] += 1
        return slots, unmatched_types, stats
    finally:
        wb.close()


def expected_caregiver_windows_for_date(
    slots: List[Dict[str, Any]],
    target: date,
) -> Dict[str, Dict[str, List[TimeRange]]]:
    """
    Group by name_key → availability / explicit unavailability ranges for the date.
    """
    by_user: Dict[str, Dict[str, List[TimeRange]]] = defaultdict(
        lambda: {"availability": [], "unavailability": [], "display_name": ""}
    )
    name_meta: Dict[str, Dict[str, Any]] = {}

    for slot in slots:
        clipped = resolve_slot_window_for_date(slot, target)
        if not clipped:
            continue
        start_m, end_m = clipped
        rng = (minutes_to_hhmm(start_m), minutes_to_hhmm(end_m))
        key = slot["name_key"]
        bucket = by_user[key]
        if not bucket.get("display_name"):
            bucket["display_name"] = slot["user_name"]
        if slot["is_unavailability"]:
            bucket["unavailability"].append(rng)
        else:
            bucket["availability"].append(rng)
        name_meta[key] = slot

    result: Dict[str, Dict[str, List[TimeRange]]] = {}
    for key, buckets in by_user.items():
        effective = subtract_day_intervals(buckets["availability"], buckets["unavailability"])
        if not effective and not buckets["unavailability"]:
            continue
        file_unavail = union_day_intervals(
            complement_day_intervals(effective),
            buckets["unavailability"],
        )
        result[key] = {
            "availability": effective,
            "unavailability_from_file": file_unavail,
            "explicit_unavailability": merge_day_intervals(buckets["unavailability"]),
            "display_name": buckets.get("display_name") or key,
            "name_keys": name_meta.get(key, {}).get("name_keys") or {key},
        }
    return result


# ---------------------------------------------------------------------------
# API data loaders
# ---------------------------------------------------------------------------


def fetch_availability_types(client: ApiClient) -> Dict[str, Dict[str, Any]]:
    rows = client.get_paginated(
        "panel/v1/basic-info/availability-types",
        params={},
        limit=100,
    )
    types_map: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        name = (row.get("name") or "").strip()
        if not name:
            continue
        type_str = (row.get("type") or "").strip().lower()
        is_unavail = type_str == "unavailability"
        # Same special-case as migration
        if name.lower() == "swap day":
            is_unavail = False
        types_map[name.lower()] = {
            "id": row.get("id"),
            "name": name,
            "is_unavailability": is_unavail,
            "type": type_str,
        }
    return types_map


def fetch_unallocated(client: ApiClient, target: date) -> List[Dict[str, Any]]:
    rows = client.get(f"panel/v1/roster/{target.isoformat()}/unallocated")
    if not isinstance(rows, list):
        raise TestTodayError("Unexpected unallocated response (expected list)")
    return rows


def fetch_cancellations(client: ApiClient, target: date) -> List[Dict[str, Any]]:
    rows = client.get(f"panel/v1/roster/{target.isoformat()}/cancellations")
    if not isinstance(rows, list):
        raise TestTodayError("Unexpected cancellations response (expected list)")
    return rows


def fetch_available_clients(client: ApiClient, target: date) -> List[Dict[str, Any]]:
    return client.get_paginated(
        "panel/v1/roster/available-clients",
        params={"date": target.isoformat()},
        limit=100,
    )


def fetch_available_caregivers(
    client: ApiClient,
    target: date,
    search: Optional[str] = None,
) -> List[Dict[str, Any]]:
    params: Dict[str, Any] = {"date": target.isoformat()}
    if search:
        params["search"] = search
    return client.get_paginated(
        "panel/v1/roster/available-caregivers",
        params=params,
        limit=100,
    )


def fetch_users(client: ApiClient, search: Optional[str] = None) -> List[Dict[str, Any]]:
    params: Dict[str, Any] = {}
    if search:
        params["search"] = search
    try:
        return client.get_paginated("panel/v1/users", params=params, limit=100)
    except TestTodayError:
        return []


def fetch_user_availability(
    client: ApiClient,
    target: date,
    user_id: int,
) -> Dict[str, Any]:
    data = client.get(f"panel/v1/roster/{target.isoformat()}/users/{user_id}/availability")
    if not isinstance(data, dict):
        raise TestTodayError(f"Unexpected user availability response for user {user_id}")
    return data


def visit_client_name(visit: Dict[str, Any]) -> str:
    receiver = visit.get("receiver") or visit.get("client") or {}
    if isinstance(receiver, dict):
        name = (receiver.get("name") or "").strip()
        last = (receiver.get("lastname") or "").strip()
        if name or last:
            return f"{name} {last}".strip()
    return ""


def visit_to_slot_key(visit: Dict[str, Any]) -> Tuple[str, int, int]:
    name = visit_client_name(visit)
    return (
        normalize_person_key(name),
        int(visit.get("startMinute") or 0),
        int(visit.get("endMinute") or 0),
    )


def cancel_type_name(visit: Dict[str, Any]) -> str:
    ct = visit.get("cancellationType")
    if isinstance(ct, dict):
        return (ct.get("name") or "").strip()
    return ""


# ---------------------------------------------------------------------------
# Comparisons
# ---------------------------------------------------------------------------


def _multiset_diff(
    expected: Counter,
    actual: Counter,
) -> Tuple[List[Any], List[Any]]:
    only_exp = []
    only_act = []
    for key in sorted(set(expected) | set(actual), key=lambda x: str(x)):
        e = expected[key]
        a = actual[key]
        if e > a:
            only_exp.extend([key] * (e - a))
        elif a > e:
            only_act.extend([key] * (a - e))
    return only_exp, only_act


def resolve_caregiver_user_id(
    client: ApiClient,
    target: date,
    display_name: str,
    name_keys: Set[str],
) -> Optional[int]:
    """Resolve Excel caregiver name → API user id via available-caregivers, then users."""
    search_term = strip_title(display_name).strip() or display_name
    # Try a few search fragments
    fragments = [search_term]
    parts = search_term.split()
    if len(parts) >= 2:
        fragments.append(parts[-1])  # lastname
        fragments.append(parts[0])

    candidates: List[Dict[str, Any]] = []
    seen_ids: Set[int] = set()

    for frag in fragments:
        for row in fetch_available_caregivers(client, target, search=frag):
            uid = row.get("id")
            if uid is None or int(uid) in seen_ids:
                continue
            seen_ids.add(int(uid))
            candidates.append(row)

    def _match_row(row: Dict[str, Any]) -> bool:
        full = (row.get("name") or "").strip()
        # available-caregivers returns combined name
        keys = name_match_keys(full)
        if "lastname" in row or "name" in row:
            n = (row.get("name") or "").strip()
            ln = (row.get("lastname") or "").strip()
            if ln and " " not in full:
                keys |= name_match_keys(f"{n} {ln}")
                keys |= name_match_keys(f"{ln}, {n}")
        return bool(keys & name_keys)

    matched = [c for c in candidates if _match_row(c)]
    if len(matched) == 1:
        return int(matched[0]["id"])
    if len(matched) > 1:
        # Prefer exact normalize match
        exact = [
            c
            for c in matched
            if normalize_person_key(c.get("name") or "") in name_keys
        ]
        if len(exact) == 1:
            return int(exact[0]["id"])
        return int(matched[0]["id"])

    # Fallback: users list (covers leave / unavailable caregivers)
    for frag in fragments:
        for row in fetch_users(client, search=frag):
            uid = row.get("id")
            if uid is None:
                continue
            n = (row.get("name") or "").strip()
            ln = (row.get("lastname") or "").strip()
            keys = name_match_keys(f"{n} {ln}") | name_match_keys(f"{ln}, {n}")
            if keys & name_keys:
                return int(uid)
    return None


# ---------------------------------------------------------------------------
# Main runner
# ---------------------------------------------------------------------------


def run_test_today(
    client_hours_path: str | Path,
    caregivers_availability_path: str | Path,
    target_date: str | date,
    base_url: str,
    token: str,
    log_callback: Optional[LogFn] = None,
) -> Tuple[bool, List[str]]:
    """
    Run all Validate today's roster API checks.
    Returns (all_passed, messages).
    """
    msgs: List[str] = []
    passed = True
    target = parse_target_date(target_date)
    client_hours = Path(client_hours_path)
    caregivers_xlsx = Path(caregivers_availability_path)

    _log(msgs, log_callback, "=" * 70)
    _log(msgs, log_callback, "VALIDATE TODAY'S ROSTER – API VALIDATION")
    _log(msgs, log_callback, f"Date     : {target.isoformat()}")
    _log(msgs, log_callback, f"Base URL : {base_url}")
    _log(msgs, log_callback, f"Client Hours : {client_hours}")
    _log(msgs, log_callback, f"Caregivers   : {caregivers_xlsx}")
    _log(msgs, log_callback, "=" * 70)

    client = ApiClient(base_url, token)

    # --- Parse client hours ---
    _log(msgs, log_callback, "")
    _log(msgs, log_callback, "--- Parsing Client Hours XLSX ---")
    expected_unalloc, expected_cancel, ch_stats = load_client_hours_for_date(
        client_hours, target
    )
    for k, v in ch_stats.items():
        _log(msgs, log_callback, f"  {k}: {v}")

    # --- Availability types + caregiver file ---
    _log(msgs, log_callback, "")
    _log(msgs, log_callback, "--- Fetching availability types from API ---")
    types_map = fetch_availability_types(client)
    _log(msgs, log_callback, f"  Loaded {len(types_map)} availability type(s)")
    if "core" not in types_map:
        _log(msgs, log_callback, "FAIL: 'Core' availability type not found via API")
        passed = False

    _log(msgs, log_callback, "")
    _log(msgs, log_callback, "--- Parsing Caregivers Availability XLSX ---")
    slots, unmatched_types, cg_stats = load_caregiver_availability_slots(
        caregivers_xlsx, types_map
    )
    for k, v in cg_stats.items():
        _log(msgs, log_callback, f"  {k}: {v}")
    if unmatched_types:
        uniq = sorted(set(unmatched_types))
        _log(
            msgs,
            log_callback,
            f"  WARN: skipped unknown types: {uniq[:20]}"
            + (" ..." if len(uniq) > 20 else ""),
        )

    expected_windows = expected_caregiver_windows_for_date(slots, target)
    _log(
        msgs,
        log_callback,
        f"  Caregivers with windows on {target.isoformat()}: {len(expected_windows)}",
    )

    # --- Unallocated ---
    _log(msgs, log_callback, "")
    _log(msgs, log_callback, "--- Unallocated visits ---")
    api_unalloc = fetch_unallocated(client, target)
    # Filter API to Personal Care if serviceType names/ids available — visits have serviceTypeIds
    # Count all unallocated client visits; file is Personal Care only.
    # Prefer matching by name+minutes; report count mismatch too.
    exp_counter: Counter = Counter()
    for row in expected_unalloc:
        exp_counter[(row["name_key"], row["start_minute"], row["end_minute"])] += 1

    act_counter: Counter = Counter()
    act_details: Dict[Tuple[str, int, int], List[Dict[str, Any]]] = defaultdict(list)
    for visit in api_unalloc:
        key = visit_to_slot_key(visit)
        act_counter[key] += 1
        act_details[key].append(visit)

    _log(
        msgs,
        log_callback,
        f"  Expected (file, Personal Care, no cancel): {sum(exp_counter.values())}",
    )
    _log(msgs, log_callback, f"  Actual (API unallocated): {sum(act_counter.values())}")

    # Fuzzy match: file name_key may not equal API name_key — rebuild using name_keys.
    # Track matched API rows by index so extras are detected even when id is blank.
    used_api_indices: Set[int] = set()
    matched = 0
    only_file: List[Dict[str, Any]] = []
    for row in expected_unalloc:
        found_idx: Optional[int] = None
        for idx, visit in enumerate(api_unalloc):
            if idx in used_api_indices:
                continue
            if visit_matches_slot(visit, row):
                found_idx = idx
                break
        if found_idx is not None:
            matched += 1
            used_api_indices.add(found_idx)
        else:
            only_file.append(row)

    only_api = [v for idx, v in enumerate(api_unalloc) if idx not in used_api_indices]

    # Reclassify "extras": many are Excel cancellations that the API still shows as UNALLOCATED.
    # Those are status mismatches, not "not in file".
    used_cancel_for_status: Set[int] = set()
    status_mismatch_unalloc: List[Tuple[Dict[str, Any], Dict[str, Any]]] = []
    true_only_api: List[Dict[str, Any]] = []
    for visit in only_api:
        found_cancel_idx: Optional[int] = None
        for cidx, crow in enumerate(expected_cancel):
            if cidx in used_cancel_for_status:
                continue
            if visit_matches_slot(visit, crow):
                found_cancel_idx = cidx
                break
        if found_cancel_idx is not None:
            used_cancel_for_status.add(found_cancel_idx)
            status_mismatch_unalloc.append((visit, expected_cancel[found_cancel_idx]))
        else:
            true_only_api.append(visit)

    _log(msgs, log_callback, f"  Matched: {matched}")
    if only_file:
        passed = False
        _log(msgs, log_callback, f"FAIL: {len(only_file)} expected unallocated slot(s) missing from API:")
        for r in only_file[:40]:
            near = []
            for visit in api_unalloc:
                if name_match_keys(visit_client_name(visit)) & r["name_keys"]:
                    near.append(
                        f"{minutes_to_hhmm(int(visit.get('startMinute') or 0))}-"
                        f"{minutes_to_hhmm(int(visit.get('endMinute') or 0))}"
                    )
            near_txt = f" (API same client windows: {', '.join(near[:5])})" if near else ""
            _log(
                msgs,
                log_callback,
                f"    XLSX row={r['row_num']} {r['client_name']!r} "
                f"{minutes_to_hhmm(r['start_minute'])}-{minutes_to_hhmm(r['end_minute'])}"
                f"{near_txt}",
            )
        if len(only_file) > 40:
            _log(msgs, log_callback, f"    ... and {len(only_file) - 40} more")
    if status_mismatch_unalloc:
        passed = False
        _log(
            msgs,
            log_callback,
            f"FAIL: {len(status_mismatch_unalloc)} visit(s) cancelled in file but still UNALLOCATED in API:",
        )
        for visit, crow in status_mismatch_unalloc[:40]:
            _log(
                msgs,
                log_callback,
                f"    API id={visit.get('id')} {visit_client_name(visit)!r} "
                f"{minutes_to_hhmm(int(visit.get('startMinute') or 0))}-"
                f"{minutes_to_hhmm(int(visit.get('endMinute') or 0))} "
                f"(file row={crow['row_num']} cancel={crow.get('cancellation')!r})",
            )
        if len(status_mismatch_unalloc) > 40:
            _log(msgs, log_callback, f"    ... and {len(status_mismatch_unalloc) - 40} more")
    if true_only_api:
        passed = False
        _log(
            msgs,
            log_callback,
            f"FAIL: {len(true_only_api)} API unallocated visit(s) missing from file "
            f"(treat as DELETED):",
        )
        for v in true_only_api[:40]:
            _log(
                msgs,
                log_callback,
                f"    API id={v.get('id')} {visit_client_name(v)!r} "
                f"{minutes_to_hhmm(int(v.get('startMinute') or 0))}-"
                f"{minutes_to_hhmm(int(v.get('endMinute') or 0))} "
                f"cancel='DELETED'",
            )
        if len(true_only_api) > 40:
            _log(msgs, log_callback, f"    ... and {len(true_only_api) - 40} more")
    if not only_file and not status_mismatch_unalloc and not true_only_api:
        _log(msgs, log_callback, "PASS: Unallocated visits match")

    # --- Cancellations ---
    _log(msgs, log_callback, "")
    _log(msgs, log_callback, "--- Cancellations ---")
    api_cancel = fetch_cancellations(client, target)
    expected_cancel_total = len(expected_cancel) + len(true_only_api)
    _log(
        msgs,
        log_callback,
        f"  Expected (file): {len(expected_cancel)} + DELETED orphans: {len(true_only_api)} "
        f"= {expected_cancel_total}",
    )
    _log(msgs, log_callback, f"  Actual (API): {len(api_cancel)}")

    used_cancel_indices: Set[int] = set()
    matched_c = 0
    only_file_c: List[Dict[str, Any]] = []
    type_mismatches: List[str] = []
    for row in expected_cancel:
        found = None
        found_idx: Optional[int] = None
        for idx, visit in enumerate(api_cancel):
            if idx in used_cancel_indices:
                continue
            if visit_matches_slot(visit, row):
                found = visit
                found_idx = idx
                break
        if found is not None and found_idx is not None:
            matched_c += 1
            used_cancel_indices.add(found_idx)
            api_type = cancel_type_name(found)
            exp_type = (row.get("cancellation") or "").strip()
            if api_type and exp_type and normalize_name_for_match(api_type) != normalize_name_for_match(exp_type):
                type_mismatches.append(
                    f"client={row['client_name']!r} "
                    f"{minutes_to_hhmm(row['start_minute'])}-{minutes_to_hhmm(row['end_minute'])} "
                    f"file={exp_type!r} api={api_type!r}"
                )
        else:
            only_file_c.append(row)

    # DELETED orphans expected on cancellations endpoint with type DELETED
    deleted_missing_cancel: List[Dict[str, Any]] = []
    deleted_type_mismatches: List[str] = []
    for visit in true_only_api:
        found_idx: Optional[int] = None
        found = None
        for idx, cvisit in enumerate(api_cancel):
            if idx in used_cancel_indices:
                continue
            # Match by id when possible, else name+time
            vid = str(visit.get("id") or "")
            cid = str(cvisit.get("id") or "")
            if vid and cid and vid == cid:
                found = cvisit
                found_idx = idx
                break
            if visit_matches_slot(cvisit, {
                "name_keys": name_match_keys(visit_client_name(visit)),
                "start_minute": int(visit.get("startMinute") or 0),
                "end_minute": int(visit.get("endMinute") or 0),
            }):
                found = cvisit
                found_idx = idx
                break
        if found is not None and found_idx is not None:
            matched_c += 1
            used_cancel_indices.add(found_idx)
            api_type = cancel_type_name(found)
            if normalize_name_for_match(api_type) != normalize_name_for_match("DELETED"):
                deleted_type_mismatches.append(
                    f"client={visit_client_name(visit)!r} "
                    f"{minutes_to_hhmm(int(visit.get('startMinute') or 0))}-"
                    f"{minutes_to_hhmm(int(visit.get('endMinute') or 0))} "
                    f"expected='DELETED' api={api_type!r}"
                )
        else:
            deleted_missing_cancel.append(visit)

    only_api_c = [v for idx, v in enumerate(api_cancel) if idx not in used_cancel_indices]

    # File cancellations missing from cancellations API, but present as API unallocated
    cancel_as_unalloc: List[Dict[str, Any]] = []
    truly_missing_cancel: List[Dict[str, Any]] = []
    status_mismatch_cancel_rows = {id(crow) for _v, crow in status_mismatch_unalloc}
    for row in only_file_c:
        if id(row) in status_mismatch_cancel_rows or any(
            visit_matches_slot(v, row) for v in api_unalloc
        ):
            cancel_as_unalloc.append(row)
        else:
            truly_missing_cancel.append(row)

    _log(msgs, log_callback, f"  Matched: {matched_c}")
    if cancel_as_unalloc:
        passed = False
        _log(
            msgs,
            log_callback,
            f"FAIL: {len(cancel_as_unalloc)} expected cancellation(s) still UNALLOCATED in API "
            f"(not on cancellations endpoint):",
        )
        for r in cancel_as_unalloc[:40]:
            _log(
                msgs,
                log_callback,
                f"    XLSX row={r['row_num']} {r['client_name']!r} "
                f"{minutes_to_hhmm(r['start_minute'])}-{minutes_to_hhmm(r['end_minute'])} "
                f"cancel={r['cancellation']!r}",
            )
        if len(cancel_as_unalloc) > 40:
            _log(msgs, log_callback, f"    ... and {len(cancel_as_unalloc) - 40} more")
    if truly_missing_cancel:
        passed = False
        _log(
            msgs,
            log_callback,
            f"FAIL: {len(truly_missing_cancel)} expected cancellation(s) missing from API:",
        )
        for r in truly_missing_cancel[:40]:
            _log(
                msgs,
                log_callback,
                f"    XLSX row={r['row_num']} {r['client_name']!r} "
                f"{minutes_to_hhmm(r['start_minute'])}-{minutes_to_hhmm(r['end_minute'])} "
                f"cancel={r['cancellation']!r}",
            )
        if len(truly_missing_cancel) > 40:
            _log(msgs, log_callback, f"    ... and {len(truly_missing_cancel) - 40} more")
    if deleted_missing_cancel:
        passed = False
        _log(
            msgs,
            log_callback,
            f"FAIL: {len(deleted_missing_cancel)} DELETED orphan(s) still UNALLOCATED "
            f"(expected cancellation type DELETED):",
        )
        for v in deleted_missing_cancel[:40]:
            _log(
                msgs,
                log_callback,
                f"    API id={v.get('id')} {visit_client_name(v)!r} "
                f"{minutes_to_hhmm(int(v.get('startMinute') or 0))}-"
                f"{minutes_to_hhmm(int(v.get('endMinute') or 0))} "
                f"cancel='DELETED'",
            )
        if len(deleted_missing_cancel) > 40:
            _log(msgs, log_callback, f"    ... and {len(deleted_missing_cancel) - 40} more")
    if deleted_type_mismatches:
        passed = False
        _log(msgs, log_callback, f"FAIL: {len(deleted_type_mismatches)} DELETED type mismatch(es):")
        for line in deleted_type_mismatches[:40]:
            _log(msgs, log_callback, f"    {line}")
    if only_api_c:
        passed = False
        _log(msgs, log_callback, f"FAIL: {len(only_api_c)} API cancellation(s) not in file:")
        for v in only_api_c[:40]:
            _log(
                msgs,
                log_callback,
                f"    API {visit_client_name(v)!r} "
                f"{minutes_to_hhmm(int(v.get('startMinute') or 0))}-"
                f"{minutes_to_hhmm(int(v.get('endMinute') or 0))} "
                f"type={cancel_type_name(v)!r}",
            )
    if type_mismatches:
        passed = False
        _log(msgs, log_callback, f"FAIL: {len(type_mismatches)} cancellation type name mismatch(es):")
        for line in type_mismatches[:40]:
            _log(msgs, log_callback, f"    {line}")
    if (
        not cancel_as_unalloc
        and not truly_missing_cancel
        and not deleted_missing_cancel
        and not deleted_type_mismatches
        and not only_api_c
        and not type_mismatches
    ):
        _log(msgs, log_callback, "PASS: Cancellations match")

    # --- Available clients: cancellation windows ---
    _log(msgs, log_callback, "")
    _log(msgs, log_callback, "--- Available clients (cancellation windows) ---")
    avail_clients = fetch_available_clients(client, target)
    _log(msgs, log_callback, f"  API available-clients rows: {len(avail_clients)}")

    # Build expected cancel windows per client name_key from file
    exp_cancel_windows: Dict[str, Counter] = defaultdict(Counter)
    name_key_to_display: Dict[str, str] = {}
    for row in expected_cancel:
        key = row["name_key"]
        name_key_to_display[key] = row["client_name"]
        win = (minutes_to_hhmm(row["start_minute"]), minutes_to_hhmm(row["end_minute"]))
        exp_cancel_windows[key][win] += 1

    # Index API clients by name keys
    api_client_by_keys: List[Tuple[Set[str], Dict[str, Any]]] = []
    for c in avail_clients:
        full = (c.get("name") or "").strip()
        api_client_by_keys.append((name_match_keys(full), c))

    client_window_fail = 0
    for name_key, exp_wins in exp_cancel_windows.items():
        display = name_key_to_display.get(name_key, name_key)
        keys = name_match_keys(display)
        matched_client = None
        for ckeys, crow in api_client_by_keys:
            if ckeys & keys:
                matched_client = crow
                break
        if matched_client is None:
            # Client may still appear only via cancellations list; warn
            _log(
                msgs,
                log_callback,
                f"  WARN: client {display!r} has file cancellations but not in available-clients",
            )
            # Not necessarily a hard fail if cancellations endpoint already matched
            continue
        api_wins = Counter()
        for w in matched_client.get("cancellations") or []:
            api_wins[((w.get("startTime") or "")[:5], (w.get("endTime") or "")[:5])] += 1
        only_e, only_a = _multiset_diff(exp_wins, api_wins)
        if only_e or only_a:
            client_window_fail += 1
            passed = False
            _log(
                msgs,
                log_callback,
                f"FAIL: cancellation windows mismatch for client {display!r}",
            )
            for w in only_e:
                _log(msgs, log_callback, f"    only in file: {w[0]}-{w[1]}")
            for w in only_a:
                _log(msgs, log_callback, f"    only in API:  {w[0]}-{w[1]}")
    if client_window_fail == 0:
        _log(msgs, log_callback, "PASS: Client cancellation windows OK (or none to check)")

    # --- Caregiver windows ---
    _log(msgs, log_callback, "")
    _log(msgs, log_callback, "--- Caregiver availability / unavailability windows ---")
    unresolved = 0
    window_fail = 0
    for name_key, expected in expected_windows.items():
        display = expected["display_name"]
        keys = set(expected.get("name_keys") or {name_key})
        uid = resolve_caregiver_user_id(client, target, display, keys)
        if uid is None:
            unresolved += 1
            passed = False
            _log(
                msgs,
                log_callback,
                f"FAIL: could not resolve caregiver {display!r} to a user id",
            )
            continue

        api_av = fetch_user_availability(client, target, uid)
        api_unavail_raw = api_av.get("unavailability") or []
        api_avail_raw = api_av.get("availability") or []

        leave_windows: List[TimeRange] = []
        non_leave: List[TimeRange] = []
        for w in api_unavail_raw:
            start = (w.get("startTime") or "")[:5]
            end = (w.get("endTime") or "")[:5]
            color = (w.get("color") or "").lower()
            if color == LEAVE_COLOR:
                leave_windows.append((start, end))
            else:
                non_leave.append((start, end))

        if leave_windows:
            _log(
                msgs,
                log_callback,
                f"  Leave for {display!r} (userId={uid}): "
                + ", ".join(f"{a}-{b}" for a, b in leave_windows),
            )

        # Expected = file-derived unavailability ∪ leave (exact match vs full API unavailability)
        expected_unavail = union_day_intervals(
            expected["unavailability_from_file"],
            leave_windows,
        )
        actual_unavail = merge_day_intervals(
            [
                ((w.get("startTime") or "")[:5], (w.get("endTime") or "")[:5])
                for w in api_unavail_raw
            ]
        )

        exp_c = Counter(expected_unavail)
        act_c = Counter(actual_unavail)
        only_e, only_a = _multiset_diff(exp_c, act_c)
        if only_e or only_a:
            window_fail += 1
            passed = False
            _log(
                msgs,
                log_callback,
                f"FAIL: unavailability windows mismatch for {display!r} (userId={uid})",
            )
            _log(
                msgs,
                log_callback,
                f"    file-derived: {expected['unavailability_from_file']}",
            )
            _log(msgs, log_callback, f"    leave: {leave_windows}")
            _log(msgs, log_callback, f"    expected (file∪leave): {expected_unavail}")
            _log(msgs, log_callback, f"    actual API: {actual_unavail}")
            for w in only_e:
                _log(msgs, log_callback, f"    only expected: {w[0]}-{w[1]}")
            for w in only_a:
                _log(msgs, log_callback, f"    only API:      {w[0]}-{w[1]}")
        else:
            _log(
                msgs,
                log_callback,
                f"  PASS: {display!r} (userId={uid}) unavailability windows match "
                f"({len(actual_unavail)} window(s)"
                + (f", incl. {len(leave_windows)} leave)" if leave_windows else ")"),
            )

        # Also log availability for diagnostics (not a hard fail if unavail matched)
        api_avail = merge_day_intervals(
            [
                ((w.get("startTime") or "")[:5], (w.get("endTime") or "")[:5])
                for w in api_avail_raw
            ]
        )
        if Counter(expected["availability"]) != Counter(api_avail):
            _log(
                msgs,
                log_callback,
                f"  WARN: availability ranges differ for {display!r}: "
                f"file={expected['availability']} api={api_avail}",
            )

    if unresolved:
        _log(msgs, log_callback, f"FAIL: {unresolved} caregiver(s) could not be resolved")
    if window_fail:
        _log(msgs, log_callback, f"FAIL: {window_fail} caregiver(s) had window mismatches")
    if unresolved == 0 and window_fail == 0 and expected_windows:
        _log(msgs, log_callback, "PASS: All caregiver unavailability windows match")
    elif not expected_windows:
        _log(msgs, log_callback, "WARN: No caregiver windows expected from file for this date")

    _log(msgs, log_callback, "")
    _log(msgs, log_callback, "=" * 70)
    if passed:
        _log(msgs, log_callback, "RESULT: ALL CHECKS PASSED")
    else:
        _log(msgs, log_callback, "RESULT: SOME CHECKS FAILED – see details above")
    _log(msgs, log_callback, "=" * 70)
    return passed, msgs
