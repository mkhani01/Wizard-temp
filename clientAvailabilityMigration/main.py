"""
Client Availability Migration Script
=====================================
Migrates client availability/schedule data from Excel file to database.

Configuration:
- WEEK_ROTATION: Set to 1 for weekly recurrence, 2 for bi-weekly (2-week rotation)
- AUTO_DETECT_WEEK_ROTATION: If True, compares both weeks to determine occurs_every automatically

Rules:
- Sheet: "Data"
- Filter: "Planned Service Type Description" = "Personal Care" AND
          "Planned Service Requirement Type Description" = "Personal Care"
- Client matching: "Service Location Name" column (format: "lastname, name")
- Time: "Service Requirement Start Date And Time" and "Service Requirement End Date And Time"
- Start Date: First service date - 14 days (2 weeks before)
- Recurrence: Based on WEEK_ROTATION or auto-detected from comparing both weeks
  - If both weeks have SAME schedule → occurs_every=1 (weekly)
  - If weeks are DIFFERENT → occurs_every=2 (bi-weekly)

Record rules (Personal Care only):
- Each distinct (start_time, end_time) becomes one availability record.
  E.g. 08:00-08:30 and 08:00-09:00 → two separate records.
- When X source rows have the exact same (start_time, end_time) on a date,
  one record is emitted with number_of_care_givers = X (max per date across dates).
"""

import os
import sys
import logging
from pathlib import Path
from datetime import datetime, timedelta, date, time
from collections import defaultdict
from typing import Optional, Dict, List, Tuple, Any

try:
    import psycopg2
    from psycopg2.extras import RealDictCursor, execute_values
    import openpyxl
    import pandas as pd
except ImportError as e:
    print(f"Missing required package: {e}")
    print("Please install: pip install psycopg2-binary openpyxl pandas")
    sys.exit(1)


# ============================================================================
# CONFIGURATION - Modify these values as needed
# ============================================================================

WEEK_ROTATION = 2
AUTO_DETECT_WEEK_ROTATION = True

DEFAULT_NUMBER_OF_CARE_GIVERS = 1
DEFAULT_FLEX_START = 0
DEFAULT_FLEX_END = 0
DEFAULT_FIX_WINDOW = False
DEFAULT_MIN_DURATION = None
DEFAULT_DURATION = None

AVAILABILITY_TYPE_NAME = 'Core'

# ============================================================================
# END CONFIGURATION
# ============================================================================


logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - [%(filename)s:%(lineno)d] - %(message)s',
    handlers=[
        logging.FileHandler('migration_client_availability.log', mode='w'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

DAYS_OF_WEEK = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']


class MigrationError(Exception):
    pass


def get_db_config() -> Dict[str, Any]:
    config = {
        'host': os.getenv('DB_HOST', 'localhost'),
        'port': int(os.getenv('DB_PORT', '5432')),
        'database': os.getenv('DB_NAME'),
        'user': os.getenv('DB_USER'),
        'password': os.getenv('DB_PASSWORD')
    }
    missing = [k for k, v in config.items() if not v]
    if missing:
        raise MigrationError(f"Missing database configuration: {missing}")
    return config


def connect_to_database(config: Dict[str, Any]):
    try:
        logger.info(f"Connecting to PostgreSQL at {config['host']}:{config['port']}/{config['database']}...")
        connection = psycopg2.connect(
            host=config['host'],
            port=config['port'],
            database=config['database'],
            user=config['user'],
            password=config['password'],
            cursor_factory=RealDictCursor,
            connect_timeout=10,
        )
        connection.autocommit = False
        logger.info("✓ Database connection established successfully")
        return connection
    except Exception as e:
        logger.error(f"✗ Failed to connect to database: {e}")
        raise MigrationError(f"Database connection failed: {e}")


def get_all_clients(connection) -> Dict[str, int]:
    cursor = connection.cursor()
    try:
        cursor.execute('SELECT id, name, lastname FROM client WHERE deleted_at IS NULL')
        clients = {}
        for row in cursor.fetchall():
            name = (row['name'] or '').strip()
            lastname = (row['lastname'] or '').strip()
            key = f"{lastname}, {name}".strip().lower()
            if key:
                clients[key] = row['id']
        logger.info(f"✓ Loaded {len(clients)} clients from database")
        return clients
    finally:
        cursor.close()


def get_availability_type(connection, type_name: str) -> Tuple[int, bool]:
    cursor = connection.cursor()
    try:
        cursor.execute(
            "SELECT id, name, type, category FROM availability_types "
            "WHERE LOWER(name) = %s AND deleted_at IS NULL",
            (type_name.lower(),)
        )
        row = cursor.fetchone()
        
        if not row:
            cursor.execute("SELECT id, name, type, category FROM availability_types WHERE deleted_at IS NULL")
            all_types = cursor.fetchall()
            logger.error("Available availability types in database:")
            for t in all_types:
                logger.error(f"  - ID: {t['id']}, Name: '{t['name']}', Type: '{t['type']}', Category: '{t['category']}'")
            raise MigrationError(f"CRITICAL: '{type_name}' availability type not found in database.")
        
        type_id = row['id']
        is_unavailability = (row['type'] or '').strip().lower() == 'unavailability'
        category = (row['category'] or '').strip()
        
        if category not in ['CLIENT', 'BOTH']:
            raise MigrationError(f"Availability type '{type_name}' has category '{category}'. Must be CLIENT or BOTH.")
        
        logger.info(f"✓ Found '{type_name}': ID={type_id}, is_unavailability={is_unavailability}, category={category}")
        return type_id, is_unavailability
    finally:
        cursor.close()


def parse_datetime_value(datetime_val) -> Optional[datetime]:
    if datetime_val is None:
        return None
    if isinstance(datetime_val, datetime):
        return datetime_val
    if isinstance(datetime_val, str):
        datetime_str = datetime_val.strip()
        for fmt in ['%d-%m-%Y %H:%M:%S', '%Y-%m-%d %H:%M:%S', '%d/%m/%Y %H:%M:%S',
                    '%d-%m-%Y %H:%M', '%Y-%m-%d %H:%M', '%d/%m/%Y %H:%M']:
            try:
                return datetime.strptime(datetime_str, fmt)
            except ValueError:
                continue
    # Excel serial: days since 1899-12-30 (with fractional day = time)
    if isinstance(datetime_val, (int, float)):
        try:
            base = datetime(1899, 12, 30)
            return base + timedelta(days=float(datetime_val))
        except (OverflowError, ValueError):
            pass
    return None


def get_day_of_week(date_obj: date) -> str:
    return DAYS_OF_WEEK[date_obj.weekday()]


def format_time_str(t: time) -> str:
    return t.strftime('%H:%M:%S')


def normalize_time_for_slot(t: time) -> time:
    """Normalize time so same logical slot groups together.
    Floors to 10-minute boundary so e.g. 20:00 (string) and 20:05 (Excel serial) both become 20:00."""
    minute_floor = (t.minute // 10) * 10

    return time(t.hour, minute_floor, 0)


def format_date_str(d: date) -> str:
    return d.strftime('%Y-%m-%d')


def get_week_number(date_obj: date, reference_date: date) -> int:
    """Get week number (1, 2, 3, etc.) relative to reference date"""
    days_diff = (date_obj - reference_date).days
    if days_diff < 0:
        return 0
    return (days_diff // 7) + 1


def find_excel_column_index(header_row: tuple, column_names: List[str]) -> int:
    for col_name in column_names:
        for i, cell_value in enumerate(header_row):
            if cell_value and str(cell_value).strip().lower() == col_name.lower():
                return i
    return -1


def process_xlsx_file(filepath: Path, clients_map: Dict[str, int]) -> Tuple[Dict[int, List[Dict]], List[str]]:
    logger.info(f"\n{'='*60}")
    logger.info(f"PROCESSING EXCEL FILE: {filepath}")
    logger.info(f"{'='*60}")
    
    if not filepath.exists():
        raise MigrationError(f"Excel file not found: {filepath}")
    
    sheet_name = 'Data'
    header_row = None
    row_iter = None

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise MigrationError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        ws = wb[sheet_name]
        header_row = tuple(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        row_iter = ws.iter_rows(min_row=2, values_only=True)
    except MigrationError:
        raise
    except Exception as e:
        try:
            df = pd.read_excel(filepath, sheet_name=sheet_name)
            header_row = tuple(df.columns.astype(str))
            row_iter = (tuple(x) for x in df.to_numpy())
            logger.info("Loaded Excel via pandas (openpyxl fallback)")
        except Exception as e2:
            raise MigrationError(f"Failed to load Excel file: {e2}") from e
    
    logger.info(f"Header row has {len(header_row)} columns")
    
    col_service_location_name = find_excel_column_index(header_row, ['Service Location Name'])
    col_service_type_desc = find_excel_column_index(header_row, ['Planned Service Type Description'])
    col_service_req_type_desc = find_excel_column_index(header_row, ['Planned Service Requirement Type Description'])
    col_start_datetime = find_excel_column_index(header_row, ['Service Requirement Start Date And Time'])
    col_end_datetime = find_excel_column_index(header_row, ['Service Requirement End Date And Time'])
    
    logger.info(f"Column mappings:")
    logger.info(f"  - Service Location Name: column {col_service_location_name}")
    logger.info(f"  - Planned Service Type Description: column {col_service_type_desc}")
    logger.info(f"  - Planned Service Requirement Type Description: column {col_service_req_type_desc}")
    logger.info(f"  - Service Requirement Start Date And Time: column {col_start_datetime}")
    logger.info(f"  - Service Requirement End Date And Time: column {col_end_datetime}")
    
    required_cols = {
        'Service Location Name': col_service_location_name,
        'Planned Service Type Description': col_service_type_desc,
        'Planned Service Requirement Type Description': col_service_req_type_desc,
        'Service Requirement Start Date And Time': col_start_datetime,
        'Service Requirement End Date And Time': col_end_datetime,
    }
    
    missing_cols = [name for name, idx in required_cols.items() if idx == -1]
    if missing_cols:
        raise MigrationError(f"Missing required columns: {missing_cols}")
    
    client_records = defaultdict(list)
    unmatched_clients = set()
    skipped_non_personal_care = 0
    skipped_missing_data = 0
    total_rows = 0
    
    for row_num, row in enumerate(row_iter, start=2):
        total_rows += 1
        
        service_location_name = row[col_service_location_name] if col_service_location_name < len(row) else None
        service_type_desc = row[col_service_type_desc] if col_service_type_desc < len(row) else None
        service_req_type_desc = row[col_service_req_type_desc] if col_service_req_type_desc < len(row) else None
        start_datetime_val = row[col_start_datetime] if col_start_datetime < len(row) else None
        end_datetime_val = row[col_end_datetime] if col_end_datetime < len(row) else None
        
        # Check for empty/NaN values
        is_empty = False
        if hasattr(pd, 'isna'):
            is_empty = pd.isna(service_location_name)
        else:
            is_empty = not service_location_name
        
        if is_empty:
            logger.warning(
                "Row %d: SKIPPED - empty Service Location Name | Type=%r, ReqType=%r, Start=%r, End=%r",
                row_num, service_type_desc, service_req_type_desc, start_datetime_val, end_datetime_val
            )
            continue

        service_type_str = str(service_type_desc).strip() if service_type_desc else ''
        service_req_type_str = str(service_req_type_desc).strip() if service_req_type_desc else ''

        if service_type_str.lower() != 'personal care' or service_req_type_str.lower() != 'personal care':
            skipped_non_personal_care += 1
            logger.warning(
                "Row %d: SKIPPED - not Personal Care | Location=%r, Type=%r, ReqType=%r, Start=%r, End=%r",
                row_num, service_location_name, service_type_str, service_req_type_str, start_datetime_val, end_datetime_val
            )
            continue

        client_key = str(service_location_name).strip().lower()

        client_id = clients_map.get(client_key)
        if not client_id:
            unmatched_clients.add(str(service_location_name))
            logger.warning(
                "Row %d: SKIPPED - client not found | Location=%r, Type=%r, ReqType=%r, Start=%r, End=%r",
                row_num, service_location_name, service_type_str, service_req_type_str, start_datetime_val, end_datetime_val
            )
            continue

        start_datetime = parse_datetime_value(start_datetime_val)
        end_datetime = parse_datetime_value(end_datetime_val)

        # Extra safety check for None/NaT values
        if not start_datetime or not end_datetime:
            skipped_missing_data += 1
            logger.warning(
                "Row %d: SKIPPED - missing datetime | Location=%r, Start=%r, End=%r",
                row_num, service_location_name, start_datetime_val, end_datetime_val
            )
            continue

        # Ensure we have actual datetime objects
        try:
            start_date_val = start_datetime.date()
            start_time_val = start_datetime.time()
            end_time_val = end_datetime.time()
        except (ValueError, AttributeError) as e:
            skipped_missing_data += 1
            logger.warning(
                "Row %d: SKIPPED - invalid datetime | Location=%r, Start=%r, End=%r, error=%s",
                row_num, service_location_name, start_datetime_val, end_datetime_val, e
            )
            continue

        client_records[client_id].append({
            'client_id': client_id,
            'client_name': service_location_name,
            'start_datetime': start_datetime,
            'end_datetime': end_datetime,
            'start_date': start_date_val,
            'start_time': start_time_val,
            'end_time': end_time_val,
            'day_of_week': get_day_of_week(start_date_val),
            'source_row': row_num,
        })
        logger.info(
            "Row %d: ADDED | client=%r (id=%s), start=%s %s, end_time=%s, day=%s",
            row_num, service_location_name, client_id, start_date_val, start_time_val, end_time_val, get_day_of_week(start_date_val)
        )

    logger.info(f"\n{'='*60}")
    logger.info("EXCEL PROCESSING SUMMARY")
    logger.info(f"{'='*60}")
    logger.info(f"Total rows processed: {total_rows}")
    logger.info(f"Clients with records: {len(client_records)}")
    logger.info(f"Skipped (non-Personal Care): {skipped_non_personal_care}")
    logger.info(f"Skipped (missing data): {skipped_missing_data}")
    logger.info(f"Unmatched clients: {len(unmatched_clients)}")
    
    return client_records, list(unmatched_clients)


def analyze_client_schedule(records: List[Dict]) -> Dict[str, Any]:
    """
    Analyze a client's schedule to determine recurrence pattern.
    
    Record semantics:
    - Each distinct (day_of_week, start_time, end_time) → one schedule record.
      E.g. 08:00-08:30 and 08:00-09:00 produce two separate records.
    - When X source rows have the exact same (start_time, end_time), one record
      is produced with number_of_care_givers = X (max over dates for that slot).
    
    Recurrence logic:
    - Group records by (day_of_week, start_time, end_time)
    - Check if each slot appears in week 1, week 2, or both
    - If ALL slots appear in BOTH weeks → weekly (occurs_every=1)
    - If some slots are DIFFERENT between weeks → bi-weekly (occurs_every=2)
    """
    if not records:
        return {'schedules': [], 'occurs_every': WEEK_ROTATION, 'min_date': None}
    
    min_date = min(r['start_date'] for r in records)
    logger.info(f"  Analyzing schedule - min_date: {min_date}, total records: {len(records)}")
    
    # Group by (day_of_week, start_time, end_time)
    # Track which weeks each slot appears in
    # Count source rows (caregivers) PER DATE so we get "caregivers per occurrence", not total across all dates
    slot_weeks = defaultdict(set)
    slot_record_count_per_date = defaultdict(lambda: defaultdict(int))
    
    for record in records:
        # Normalize times so "same" slot (e.g. 20:00-20:30) groups together even if Excel had microsecond variance
        start_n = normalize_time_for_slot(record['start_time'])
        end_n = normalize_time_for_slot(record['end_time'])
        key = (record['day_of_week'], start_n, end_n)
        slot_record_count_per_date[key][record['start_date']] += 1
        week_num = get_week_number(record['start_date'], min_date)
        # Only consider weeks 1 and 2 for comparison
        if 1 <= week_num <= 2:
            slot_weeks[key].add(week_num)
    
    logger.info(f"  Found {len(slot_weeks)} unique slot types")
    
    # number_of_care_givers = max over all dates of (rows for that slot on that date) = caregivers per occurrence
    def _caregivers_for_slot(slot_key):
        counts = slot_record_count_per_date.get(slot_key, {})
        return max(1, max(counts.values())) if counts else 1
    
    # Classify slots; number_of_care_givers = count of source rows for that slot on a single date (same time = multiple caregivers)
    both_weeks = []  # Same slot in both week 1 and week 2
    week1_only = []  # Slot only in week 1
    week2_only = []  # Slot only in week 2
    
    for (day, start_time, end_time), weeks in slot_weeks.items():
        slot_key = (day, start_time, end_time)
        num_care_givers = _caregivers_for_slot(slot_key)
        has_week1 = 1 in weeks
        has_week2 = 2 in weeks
        
        if has_week1 and has_week2:
            both_weeks.append({
                'day': day,
                'start_time': start_time,
                'end_time': end_time,
                'number_of_care_givers': num_care_givers,
            })
            logger.debug(f"    Slot {day} {format_time_str(start_time)}-{format_time_str(end_time)}: BOTH weeks, caregivers={num_care_givers}")
        elif has_week1:
            week1_only.append({
                'day': day,
                'start_time': start_time,
                'end_time': end_time,
                'number_of_care_givers': num_care_givers,
            })
            logger.debug(f"    Slot {day} {format_time_str(start_time)}-{format_time_str(end_time)}: WEEK 1 only, caregivers={num_care_givers}")
        elif has_week2:
            week2_only.append({
                'day': day,
                'start_time': start_time,
                'end_time': end_time,
                'number_of_care_givers': num_care_givers,
            })
            logger.debug(f"    Slot {day} {format_time_str(start_time)}-{format_time_str(end_time)}: WEEK 2 only, caregivers={num_care_givers}")
    
    logger.info(f"  Slot analysis: both_weeks={len(both_weeks)}, week1_only={len(week1_only)}, week2_only={len(week2_only)}")
    
    # Determine occurs_every
    if AUTO_DETECT_WEEK_ROTATION:
        # If all slots appear in both weeks, it's weekly
        # If any slots are different between weeks, it's bi-weekly
        if not week1_only and not week2_only:
            occurs_every = 1
            logger.info(f"  ✓ Auto-detected: WEEKLY (occurs_every=1) - all slots in both weeks")
        else:
            occurs_every = 2
            logger.info(f"  ✓ Auto-detected: BI-WEEKLY (occurs_every=2) - different slots between weeks")
    else:
        occurs_every = WEEK_ROTATION
        logger.info(f"  Using configured WEEK_ROTATION: {WEEK_ROTATION}")
    
    # Build schedules list
    schedules = []
    
    if occurs_every == 1:
        # Weekly: Create one record per unique slot
        for item in both_weeks:
            schedules.append({
                'day': item['day'],
                'start_time': item['start_time'],
                'end_time': item['end_time'],
                'start_date': min_date,
                'occurs_every': 1,
                'number_of_care_givers': item.get('number_of_care_givers', 1),
            })
        # Also include week1_only and week2_only if any (they become weekly too)
        for item in week1_only + week2_only:
            schedules.append({
                'day': item['day'],
                'start_time': item['start_time'],
                'end_time': item['end_time'],
                'start_date': min_date,
                'occurs_every': 1,
                'number_of_care_givers': item.get('number_of_care_givers', 1),
            })
    else:
        # Bi-weekly: Create records based on which week they appear
        for item in both_weeks:
            # Slot in both weeks with bi-weekly = one record starting week 1
            schedules.append({
                'day': item['day'],
                'start_time': item['start_time'],
                'end_time': item['end_time'],
                'start_date': min_date,
                'occurs_every': 2,
                'number_of_care_givers': item.get('number_of_care_givers', 1),
            })
        
        for item in week1_only:
            schedules.append({
                'day': item['day'],
                'start_time': item['start_time'],
                'end_time': item['end_time'],
                'start_date': min_date,
                'occurs_every': 2,
                'number_of_care_givers': item.get('number_of_care_givers', 1),
            })
        
        for item in week2_only:
            week2_start = min_date + timedelta(days=7)
            schedules.append({
                'day': item['day'],
                'start_time': item['start_time'],
                'end_time': item['end_time'],
                'start_date': week2_start,
                'occurs_every': 2,
                'number_of_care_givers': item.get('number_of_care_givers', 1),
            })
    
    return {
        'schedules': schedules,
        'occurs_every': occurs_every,
        'min_date': min_date,
        'both_weeks': len(both_weeks),
        'week1_only': len(week1_only),
        'week2_only': len(week2_only)
    }


def generate_availability_records(
    client_records: Dict[int, List[Dict]], 
    type_id: int,
    is_unavailability: bool
) -> List[Dict]:
    logger.info(f"\n{'='*60}")
    logger.info("GENERATING AVAILABILITY RECORDS")
    logger.info(f"{'='*60}")
    
    availabilities = []
    
    for client_id, records in client_records.items():
        logger.info(f"\nProcessing client_id={client_id} ({len(records)} records)")
        
        analysis = analyze_client_schedule(records)
        schedules = analysis['schedules']
        min_date = analysis['min_date']
        
        if not min_date:
            logger.warning(f"  No valid schedules for client {client_id}")
            continue
        
        # Start date: 2 weeks before first service
        target_start_date = min_date - timedelta(days=14)
        logger.info(f"  First service: {min_date}, Start date (-14 days): {target_start_date}")
        
        for schedule in schedules:
            num_care_givers = schedule.get('number_of_care_givers', DEFAULT_NUMBER_OF_CARE_GIVERS)
            availabilities.append({
                'client_id': client_id,
                'days': [schedule['day']],
                'requested_start_time': format_time_str(schedule['start_time']),
                'requested_end_time': format_time_str(schedule['end_time']),
                'start_time': format_time_str(schedule['start_time']),
                'end_time': format_time_str(schedule['end_time']),
                'is_temp': False,
                'is_unavailability': is_unavailability,
                'type_id': type_id,
                'number_of_care_givers': num_care_givers,
                'flex_start': DEFAULT_FLEX_START,
                'flex_end': DEFAULT_FLEX_END,
                'fix_window': DEFAULT_FIX_WINDOW,
                'min_duration': DEFAULT_MIN_DURATION,
                'duration': DEFAULT_DURATION,
                'note': None,
                'start_date': format_date_str(schedule['start_date']),
                'end_date': None,
                'occurs_every': schedule['occurs_every'],
                'effective_date_from': None,
                'effective_date_to': None,
            })
            
            logger.info(f"  → {schedule['day']} {format_time_str(schedule['start_time'])}-{format_time_str(schedule['end_time'])}, "
                       f"start={format_date_str(schedule['start_date'])}, occurs_every={schedule['occurs_every']}, "
                       f"number_of_care_givers={num_care_givers}")
    
    logger.info(f"\nGenerated {len(availabilities)} total availability records")
    return availabilities


def deduplicate_availabilities(availabilities: List[Dict]) -> List[Dict]:
    unique_map = {}
    duplicates = 0
    
    for avail in availabilities:
        key = (
            avail['client_id'],
            tuple(avail['days']),
            avail['start_time'],
            avail['end_time'],
            avail['start_date'],
            avail['occurs_every'],
        )
        
        if key in unique_map:
            duplicates += 1
            # Keep the one with higher number_of_care_givers (same slot, multiple caregivers)
            if avail.get('number_of_care_givers', 1) > unique_map[key].get('number_of_care_givers', 1):
                unique_map[key] = avail
        else:
            unique_map[key] = avail
    
    logger.info(f"Deduplication: {len(availabilities)} → {len(unique_map)} records ({duplicates} duplicates removed)")
    return list(unique_map.values())


def clear_client_availabilities(connection) -> None:
    """Remove all existing rows from client_availabilities before seeding."""
    cursor = connection.cursor()
    try:
        cursor.execute("DELETE FROM client_availabilities")
        deleted = cursor.rowcount
        connection.commit()
        logger.info(f"Cleared client_availabilities: {deleted} existing row(s) removed")
    except Exception as e:
        connection.rollback()
        logger.error(f"Failed to clear client_availabilities: {e}")
        raise MigrationError(f"Failed to clear client_availabilities: {e}")
    finally:
        cursor.close()


def seed_availabilities(connection, availabilities: List[Dict]) -> int:
    logger.info(f"\n{'='*60}")
    logger.info("SEEDING DATABASE")
    logger.info(f"{'='*60}")
    
    clear_client_availabilities(connection)
    
    if not availabilities:
        logger.warning("No availabilities to insert")
        return 0
    
    cursor = connection.cursor()
    try:
        cursor.execute("SELECT typname FROM pg_type WHERE typname = 'client_availabilities_days_enum'")
        enum_exists = cursor.fetchone()
        
        array_cast = '::text[]::client_availabilities_days_enum[]' if enum_exists else '::text[]'
        
        # Note: minDuration and duration use camelCase in DB (no explicit name mapping in entity)
        insert_query = f"""
            INSERT INTO client_availabilities (
                client_id, days, requested_start_time, requested_end_time,
                start_time, end_time, is_temp, is_unavailability, type_id,
                number_of_care_givers, flex_start, flex_end, fix_window,
                "minDuration", duration, note,
                start_date, end_date, occurs_every,
                effective_date_from, effective_date_to,
                created_date, last_modified_date
            ) VALUES %s
            RETURNING id
        """
        
        availability_tuples = [
            (
                avail['client_id'],
                avail['days'],
                avail['requested_start_time'],
                avail['requested_end_time'],
                avail['start_time'],
                avail['end_time'],
                avail['is_temp'],
                avail['is_unavailability'],
                avail['type_id'],
                avail['number_of_care_givers'],
                avail['flex_start'],
                avail['flex_end'],
                avail['fix_window'],
                avail['min_duration'],
                avail['duration'],
                avail['note'],
                avail['start_date'],
                avail['end_date'],
                avail['occurs_every'],
                avail['effective_date_from'],
                avail['effective_date_to'],
            )
            for avail in availabilities
        ]
        
        logger.info(f"Inserting {len(availability_tuples)} records...")
        
        execute_values(
            cursor,
            insert_query,
            availability_tuples,
            template=f"(%s, %s{array_cast}, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW())"
        )
        
        inserted = cursor.fetchall()
        connection.commit()
        
        logger.info(f"✓ Successfully inserted {len(inserted)} client availability records")
        return len(inserted)
        
    except Exception as e:
        connection.rollback()
        logger.error(f"✗ Failed to insert: {e}")
        raise MigrationError(f"Database insertion failed: {e}")
    finally:
        cursor.close()


def generate_summary_report(
    client_records: Dict[int, List[Dict]],
    availabilities: List[Dict],
    unmatched_clients: List[str],
    inserted_count: int
) -> str:
    report = []
    report.append("\n" + "="*70)
    report.append("CLIENT AVAILABILITY MIGRATION SUMMARY REPORT")
    report.append("="*70)
    
    report.append(f"\n📊 CONFIGURATION:")
    report.append(f"   - Week Rotation: {WEEK_ROTATION}")
    report.append(f"   - Auto-detect: {AUTO_DETECT_WEEK_ROTATION}")
    report.append(f"   - Availability Type: {AVAILABILITY_TYPE_NAME}")
    
    report.append(f"\n📊 INPUT DATA:")
    report.append(f"   - Clients with records: {len(client_records)}")
    report.append(f"   - Total source records: {sum(len(r) for r in client_records.values())}")
    report.append(f"   - Unmatched clients: {len(unmatched_clients)}")
    
    if unmatched_clients:
        report.append(f"\n⚠️  UNMATCHED CLIENTS ({len(unmatched_clients)}):")
        for client in sorted(unmatched_clients)[:20]:
            report.append(f"   - {client}")
        if len(unmatched_clients) > 20:
            report.append(f"   ... and {len(unmatched_clients) - 20} more")
    
    report.append(f"\n📋 GENERATED RECORDS:")
    report.append(f"   - Total: {len(availabilities)}")
    
    by_client = defaultdict(int)
    for a in availabilities:
        by_client[a['client_id']] += 1
    
    report.append(f"   - Unique clients: {len(by_client)}")
    
    by_occurs = defaultdict(int)
    for a in availabilities:
        by_occurs[a['occurs_every']] += 1
    
    report.append(f"\n📅 RECURRENCE PATTERNS:")
    for occurs, count in sorted(by_occurs.items()):
        pattern = "Weekly" if occurs == 1 else f"Bi-weekly (every {occurs} weeks)"
        report.append(f"   - {pattern}: {count} records")
    
    report.append(f"\n✅ DATABASE:")
    report.append(f"   - Records inserted: {inserted_count}")
    
    report.append(f"\n📝 SAMPLE RECORDS (first 5):")
    for i, a in enumerate(availabilities[:5]):
        report.append(f"   {i+1}. Client ID: {a['client_id']}, Day: {a['days']}, "
                     f"Time: {a['requested_start_time']}-{a['requested_end_time']}, "
                     f"Start: {a['start_date']}, Occurs: {a['occurs_every']}w")
    
    report.append("\n" + "="*70)
    return "\n".join(report)


def run(xlsx_path: Optional[str] = None) -> bool:
    print(f"""
    ╔════════════════════════════════════════════════════════════════╗
    ║   CLIENT AVAILABILITY MIGRATION                                ║
    ║                                                                ║
    ║   - Filter: Personal Care only                                 ║
    ║   - Start Date: First service date - 14 days                   ║
    ║   - Auto-detect weekly vs bi-weekly                            ║
    ╚════════════════════════════════════════════════════════════════╝
    """)
    
    from migration_support import get_assets_dir
    filepath = Path(xlsx_path) if xlsx_path else get_assets_dir() / 'clientAvailability' / 'ClientHoursWithServiceType.xlsx'
    logger.info(f"Excel file path: {filepath}")
    
    connection = None
    
    try:
        logger.info(f"\n{'='*60}")
        logger.info("STEP 1: DATABASE CONNECTION")
        logger.info(f"{'='*60}")
        
        config = get_db_config()
        connection = connect_to_database(config)
        
        logger.info(f"\n{'='*60}")
        logger.info("STEP 2: LOAD REFERENCE DATA")
        logger.info(f"{'='*60}")
        
        clients_map = get_all_clients(connection)
        type_id, is_unavailability = get_availability_type(connection, AVAILABILITY_TYPE_NAME)
        
        logger.info(f"\n{'='*60}")
        logger.info("STEP 3: PROCESS EXCEL FILE")
        logger.info(f"{'='*60}")
        
        client_records, unmatched_clients = process_xlsx_file(filepath, clients_map)
        
        if not client_records:
            logger.error("No valid Personal Care records found")
            return False
        
        logger.info(f"\n{'='*60}")
        logger.info("STEP 4: GENERATE AVAILABILITY RECORDS")
        logger.info(f"{'='*60}")
        
        availabilities = generate_availability_records(client_records, type_id, is_unavailability)
        availabilities = deduplicate_availabilities(availabilities)
        
        logger.info(f"\n{'='*60}")
        logger.info("STEP 5: SEED DATABASE")
        logger.info(f"{'='*60}")
        
        inserted_count = seed_availabilities(connection, availabilities)
        
        report = generate_summary_report(client_records, availabilities, unmatched_clients, inserted_count)
        print(report)
        logger.info(report)
        
        print("\n" + "="*60)
        print("✓ MIGRATION COMPLETED SUCCESSFULLY")
        print("="*60)
        
        return True
        
    except MigrationError as e:
        logger.error(f"Migration error: {e}")
        print(f"\n✗ MIGRATION FAILED: {e}")
        return False
    except Exception as e:
        logger.exception(f"Unexpected error: {e}")
        print(f"\n✗ MIGRATION FAILED: {e}")
        return False
    finally:
        if connection:
            connection.close()
            logger.info("\nDatabase connection closed")


if __name__ == "__main__":
    xlsx_arg = sys.argv[1] if len(sys.argv) > 1 else None
    success = run(xlsx_arg)
    sys.exit(0 if success else 1)