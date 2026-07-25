"""
Update Today Visits Migration
=============================
Update roster visits for a selected calendar date using Client Hours with Service Type:

1. Rows with non-empty Cancellation Description → cancel matching ALLOCATED/UNALLOCATED
   visits (client + start/end minutes) with that cancellation type (insert type if missing).
2. All ALLOCATED/UNALLOCATED visits for terminated clients on that date → cancel with
   type "Terminated" (insert if missing).
3. Personal Care rows where Service Requirement start AND end are empty, but Actual
   start/end are present → create a one-day temporary client_schedule (is_temporary,
   effective for the target date only) and an UNALLOCATED roster_visit linked to it.

Missing visits for cancel path → skip and log. Missing roster → created when temps needed.
"""

from __future__ import annotations

import logging
import os
import sys
import uuid
from datetime import date, datetime, timedelta, time
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Set, Tuple

try:
    import psycopg2
    from psycopg2.extras import RealDictCursor
    import openpyxl
except ImportError as e:
    print(f"Missing required package: {e}")
    print("Please install: pip install psycopg2-binary openpyxl")
    sys.exit(1)

try:
    from connection_manager import ConnectionLostError
except ImportError:
    ConnectionLostError = None

from encoding_utils import fix_utf8_mojibake, normalize_name_for_match
from person_match_utils import (
    add_person_to_name_map,
    extract_eircode_from_address,
    resolve_person_id,
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - [%(filename)s:%(lineno)d] - %(message)s",
    handlers=[
        logging.FileHandler("migration_update_today_visits.log", mode="w"),
        logging.StreamHandler(sys.stdout),
    ],
)
logger = logging.getLogger(__name__)

TERMINATED_CANCELLATION_TYPE = "Terminated"
ACTIVE_VISIT_STATUSES = ("UNALLOCATED", "ALLOCATED")
PERSONAL_CARE = "Personal Care"
DAYS_OF_WEEK = [
    "Monday",
    "Tuesday",
    "Wednesday",
    "Thursday",
    "Friday",
    "Saturday",
    "Sunday",
]
TEMP_SCHEDULE_NOTE = "migration temp from Actual"


class MigrationError(Exception):
    pass


def get_db_config() -> Dict[str, Any]:
    config = {
        "host": os.getenv("DB_HOST", "localhost"),
        "port": int(os.getenv("DB_PORT", "5432")),
        "database": os.getenv("DB_NAME"),
        "user": os.getenv("DB_USER"),
        "password": os.getenv("DB_PASSWORD"),
    }
    missing = [k for k, v in config.items() if not v]
    if missing:
        raise MigrationError(f"Missing database configuration: {missing}")
    return config


def connect_to_database(config: Dict[str, Any]):
    connection = psycopg2.connect(
        host=config["host"],
        port=config["port"],
        database=config["database"],
        user=config["user"],
        password=config["password"],
        cursor_factory=RealDictCursor,
        connect_timeout=10,
        keepalives=1,
        keepalives_idle=30,
        keepalives_interval=10,
        keepalives_count=5,
    )
    connection.autocommit = False
    return connection


def parse_datetime_value(datetime_val) -> Optional[datetime]:
    if datetime_val is None:
        return None
    if isinstance(datetime_val, datetime):
        return datetime_val
    if isinstance(datetime_val, date) and not isinstance(datetime_val, datetime):
        return datetime.combine(datetime_val, time.min)
    if isinstance(datetime_val, str):
        datetime_str = datetime_val.strip()
        if not datetime_str:
            return None
        for fmt in (
            "%d-%m-%Y %H:%M:%S",
            "%Y-%m-%d %H:%M:%S",
            "%d/%m/%Y %H:%M:%S",
            "%d-%m-%Y %H:%M",
            "%Y-%m-%d %H:%M",
            "%d/%m/%Y %H:%M",
            "%Y-%m-%d",
            "%d-%m-%Y",
            "%d/%m/%Y",
        ):
            try:
                return datetime.strptime(datetime_str, fmt)
            except ValueError:
                continue
    if isinstance(datetime_val, (int, float)):
        try:
            base = datetime(1899, 12, 30)
            return base + timedelta(days=float(datetime_val))
        except (OverflowError, ValueError):
            pass
    return None


def datetime_to_minutes(dt: datetime) -> int:
    return dt.hour * 60 + dt.minute


def parse_target_date(value: Any) -> date:
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    raise MigrationError(f"Invalid target date: {value!r} (expected YYYY-MM-DD)")


def resolve_start_end(
    req_start_val,
    req_end_val,
    act_start_val,
    act_end_val,
) -> Tuple[Optional[datetime], Optional[datetime], str]:
    """
    Prefer Requirement start/end; fall back to Actual per side.
    Returns (start, end, source_label).
    """
    req_start = parse_datetime_value(req_start_val)
    req_end = parse_datetime_value(req_end_val)
    act_start = parse_datetime_value(act_start_val)
    act_end = parse_datetime_value(act_end_val)

    start = req_start or act_start
    end = req_end or act_end
    if req_start and req_end:
        source = "requirement"
    elif act_start and act_end and not (req_start or req_end):
        source = "actual"
    else:
        source = "mixed"
    return start, end, source


def _cell_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _is_empty_datetime_val(val) -> bool:
    """True when Excel cell has no usable datetime (None / blank / unparseable)."""
    if val is None:
        return True
    if isinstance(val, str) and not val.strip():
        return True
    return parse_datetime_value(val) is None


def get_day_of_week(date_obj: date) -> str:
    return DAYS_OF_WEEK[date_obj.weekday()]


def format_time_str(t: time) -> str:
    return t.strftime("%H:%M:%S")


def format_date_str(d: date) -> str:
    return d.strftime("%Y-%m-%d")


def requested_duration_minutes(
    start_dt: datetime,
    end_dt: datetime,
    offset_days: int = 0,
) -> int:
    start_m = datetime_to_minutes(start_dt)
    end_m = datetime_to_minutes(end_dt)
    offset = max(0, int(offset_days or 0))
    if offset == 0 and end_m <= start_m:
        offset = 1
    total = offset * 24 * 60 + (end_m - start_m)
    return max(1, total)


def offset_days_between(start_dt: datetime, end_dt: datetime) -> int:
    offset_days = (end_dt.date() - start_dt.date()).days
    if offset_days < 0:
        return 0
    if offset_days == 0 and end_dt.time() <= start_dt.time():
        return 1
    return offset_days


def get_all_clients(connection) -> Dict[str, List[Dict[str, Any]]]:
    """Name key -> list of candidate client dicts (Active/postcode ranking at resolve time)."""
    cursor = connection.cursor()
    try:
        cursor.execute(
            "SELECT id, name, lastname, status, postcode FROM client WHERE deleted_at IS NULL"
        )
        clients: Dict[str, List[Dict[str, Any]]] = {}
        for row in cursor.fetchall():
            name = (row["name"] or "").strip()
            lastname = (row["lastname"] or "").strip()
            person = {
                "id": row["id"],
                "name": name,
                "lastname": lastname,
                "status": row["status"],
                "postcode": row["postcode"],
            }
            key_comma = normalize_name_for_match(f"{lastname}, {name}")
            key_space = normalize_name_for_match(f"{name} {lastname}")
            add_person_to_name_map(clients, key_comma, person)
            add_person_to_name_map(clients, key_space, person)
        return clients
    finally:
        cursor.close()


def ensure_cancellation_types(connection, names: Sequence[str]) -> Dict[str, int]:
    """
    Ensure each name exists in cancellation_types (is_paid=false for inserts).
    Returns name -> id map for all requested names (plus any already present).
    """
    unique_names = sorted({(n or "").strip() for n in names if (n or "").strip()})
    if not unique_names:
        return {}

    cursor = connection.cursor()
    try:
        cursor.execute("SELECT id, name FROM cancellation_types WHERE name = ANY(%s)", (unique_names,))
        existing = {row["name"]: row["id"] for row in cursor.fetchall()}
        missing = [n for n in unique_names if n not in existing]
        for name in missing:
            cursor.execute(
                """
                INSERT INTO cancellation_types (name, is_paid, created_date, last_modified_date)
                VALUES (%s, false, NOW(), NOW())
                ON CONFLICT (name) DO NOTHING
                RETURNING id, name
                """,
                (name,),
            )
            row = cursor.fetchone()
            if row:
                existing[row["name"]] = row["id"]
                logger.info("Inserted cancellation type %r (id=%s)", name, row["id"])
            else:
                cursor.execute("SELECT id, name FROM cancellation_types WHERE name = %s", (name,))
                found = cursor.fetchone()
                if found:
                    existing[found["name"]] = found["id"]
        connection.commit()
        return existing
    except Exception:
        connection.rollback()
        raise
    finally:
        cursor.close()


def load_roster_visits_for_date(connection, target_date: date) -> Tuple[Optional[str], List[Dict[str, Any]]]:
    cursor = connection.cursor()
    try:
        cursor.execute("SELECT id FROM roster WHERE date = %s", (target_date,))
        roster = cursor.fetchone()
        if not roster:
            return None, []
        roster_id = roster["id"]
        cursor.execute(
            """
            SELECT id, receiver_client_id, start_minute, end_minute, status, cancellation_type_id
            FROM roster_visit
            WHERE roster_id = %s
              AND receiver_type = 'CLIENT'
              AND receiver_client_id IS NOT NULL
            """,
            (roster_id,),
        )
        return roster_id, list(cursor.fetchall())
    finally:
        cursor.close()


def get_terminated_client_ids(connection, target_date: date) -> Set[int]:
    cursor = connection.cursor()
    try:
        cursor.execute(
            """
            SELECT id
            FROM client
            WHERE deleted_at IS NULL
              AND (
                status = 'Deactive'
                OR (termination_date IS NOT NULL AND termination_date < %s)
              )
            """,
            (target_date,),
        )
        return {int(row["id"]) for row in cursor.fetchall()}
    finally:
        cursor.close()


def _col_idx(headers: Sequence[str], names: Sequence[str]) -> int:
    lower_headers = [str(h or "").strip().lower() for h in headers]
    for name in names:
        nl = name.lower()
        for i, h in enumerate(lower_headers):
            if h == nl:
                return i
    return -1


def extract_cancellation_rows(
    filepath: Path,
    target_date: date,
    clients_map: Dict[str, List[Dict[str, Any]]],
) -> Tuple[List[Dict[str, Any]], Dict[str, int]]:
    """
    Parse Client Hours XLSX Data sheet; return cancellation rows for target_date
    and stats counters.
    """
    stats = {
        "total_rows": 0,
        "with_cancellation": 0,
        "on_target_date": 0,
        "skipped_missing_datetime": 0,
        "skipped_unknown_client": 0,
        "skipped_wrong_date": 0,
    }
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        if "Data" not in wb.sheetnames:
            raise MigrationError("Sheet 'Data' not found in workbook")
        ws = wb["Data"]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            raise MigrationError("Workbook has no header row")

        col_loc = _col_idx(header, ["Service Location Name"])
        col_addr = _col_idx(header, ["Service Location Address"])
        col_req_start = _col_idx(header, ["Service Requirement Start Date And Time"])
        col_req_end = _col_idx(header, ["Service Requirement End Date And Time"])
        col_act_start = _col_idx(header, ["Actual Start Date And Time"])
        col_act_end = _col_idx(header, ["Actual End Date And Time"])
        col_cancel = _col_idx(header, ["Cancellation Description"])

        if col_loc == -1:
            raise MigrationError("Missing required column: Service Location Name")
        if col_cancel == -1:
            raise MigrationError("Missing required column: Cancellation Description")
        if col_req_start == -1 and col_act_start == -1:
            raise MigrationError(
                "Missing start datetime columns (need Requirement and/or Actual Start)"
            )
        if col_req_end == -1 and col_act_end == -1:
            raise MigrationError(
                "Missing end datetime columns (need Requirement and/or Actual End)"
            )

        results: List[Dict[str, Any]] = []
        for row_num, row in enumerate(rows_iter, start=2):
            stats["total_rows"] += 1
            if not row:
                continue
            cancel_val = row[col_cancel] if col_cancel < len(row) else None
            cancel_name = str(cancel_val).strip() if cancel_val is not None else ""
            if not cancel_name or cancel_name.lower() == "none":
                continue
            stats["with_cancellation"] += 1

            raw_loc = row[col_loc] if col_loc < len(row) else None
            loc = fix_utf8_mojibake(raw_loc) if raw_loc is not None else None
            loc_str = str(loc).strip() if loc is not None else ""
            if not loc_str:
                logger.warning("Row %d: SKIPPED - empty Service Location Name", row_num)
                continue

            raw_addr = row[col_addr] if col_addr != -1 and col_addr < len(row) else None
            addr = fix_utf8_mojibake(raw_addr) if raw_addr is not None else None
            source_eircode = extract_eircode_from_address(addr)

            client_key = normalize_name_for_match(loc_str)
            client_id = resolve_person_id(clients_map.get(client_key) or [], source_eircode)
            if not client_id:
                stats["skipped_unknown_client"] += 1
                logger.warning(
                    "Row %d: SKIPPED - client not found | Location=%r eircode=%r Cancellation=%r",
                    row_num,
                    loc_str,
                    source_eircode or None,
                    cancel_name,
                )
                continue

            req_start = row[col_req_start] if col_req_start != -1 and col_req_start < len(row) else None
            req_end = row[col_req_end] if col_req_end != -1 and col_req_end < len(row) else None
            act_start = row[col_act_start] if col_act_start != -1 and col_act_start < len(row) else None
            act_end = row[col_act_end] if col_act_end != -1 and col_act_end < len(row) else None

            start_dt, end_dt, source = resolve_start_end(req_start, req_end, act_start, act_end)
            if not start_dt or not end_dt:
                stats["skipped_missing_datetime"] += 1
                logger.warning(
                    "Row %d: SKIPPED - missing datetime (Requirement and Actual empty) | Location=%r",
                    row_num,
                    loc_str,
                )
                continue

            visit_date = start_dt.date()
            if visit_date != target_date:
                stats["skipped_wrong_date"] += 1
                continue

            stats["on_target_date"] += 1
            results.append(
                {
                    "row_num": row_num,
                    "client_id": client_id,
                    "client_name": loc_str,
                    "cancellation_name": cancel_name,
                    "start_minute": datetime_to_minutes(start_dt),
                    "end_minute": datetime_to_minutes(end_dt),
                    "source": source,
                }
            )
        return results, stats
    finally:
        wb.close()


def extract_temp_visit_rows(
    filepath: Path,
    target_date: date,
    clients_map: Dict[str, List[Dict[str, Any]]],
) -> Tuple[List[Dict[str, Any]], Dict[str, int]]:
    """
    Personal Care rows where Requirement start AND end are empty, Actual start/end
    are present, Actual start date == target_date, and no Cancellation Description.
    """
    stats = {
        "total_rows": 0,
        "personal_care": 0,
        "req_both_empty": 0,
        "candidates": 0,
        "skipped_not_personal_care": 0,
        "skipped_has_cancellation": 0,
        "skipped_req_not_both_empty": 0,
        "skipped_missing_actual": 0,
        "skipped_unknown_client": 0,
        "skipped_wrong_date": 0,
        "skipped_empty_location": 0,
        "skipped_bad_offset": 0,
    }
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        if "Data" not in wb.sheetnames:
            raise MigrationError("Sheet 'Data' not found in workbook")
        ws = wb["Data"]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            raise MigrationError("Workbook has no header row")

        col_loc = _col_idx(header, ["Service Location Name"])
        col_addr = _col_idx(header, ["Service Location Address"])
        col_planned = _col_idx(header, ["Planned Service Type Description"])
        col_planned_req = _col_idx(
            header, ["Planned Service Requirement Type Description"]
        )
        col_req_start = _col_idx(header, ["Service Requirement Start Date And Time"])
        col_req_end = _col_idx(header, ["Service Requirement End Date And Time"])
        col_act_start = _col_idx(header, ["Actual Start Date And Time"])
        col_act_end = _col_idx(header, ["Actual End Date And Time"])
        col_cancel = _col_idx(header, ["Cancellation Description"])

        if col_loc == -1:
            raise MigrationError("Missing required column: Service Location Name")
        if col_planned == -1 or col_planned_req == -1:
            raise MigrationError(
                "Missing Personal Care filter columns: Planned Service Type Description "
                "and Planned Service Requirement Type Description"
            )
        if col_act_start == -1 or col_act_end == -1:
            raise MigrationError(
                "Missing Actual Start/End Date And Time columns (required for temp visits)"
            )

        results: List[Dict[str, Any]] = []
        for row_num, row in enumerate(rows_iter, start=2):
            stats["total_rows"] += 1
            if not row:
                continue

            planned = _cell_str(row[col_planned] if col_planned < len(row) else None)
            planned_req = _cell_str(
                row[col_planned_req] if col_planned_req < len(row) else None
            )
            if planned != PERSONAL_CARE or planned_req != PERSONAL_CARE:
                stats["skipped_not_personal_care"] += 1
                continue
            stats["personal_care"] += 1

            cancel_val = (
                row[col_cancel] if col_cancel != -1 and col_cancel < len(row) else None
            )
            cancel_name = _cell_str(cancel_val)
            if cancel_name and cancel_name.lower() != "none":
                stats["skipped_has_cancellation"] += 1
                continue

            req_start = (
                row[col_req_start]
                if col_req_start != -1 and col_req_start < len(row)
                else None
            )
            req_end = (
                row[col_req_end]
                if col_req_end != -1 and col_req_end < len(row)
                else None
            )
            if not (_is_empty_datetime_val(req_start) and _is_empty_datetime_val(req_end)):
                stats["skipped_req_not_both_empty"] += 1
                continue
            stats["req_both_empty"] += 1

            act_start_val = row[col_act_start] if col_act_start < len(row) else None
            act_end_val = row[col_act_end] if col_act_end < len(row) else None
            act_start = parse_datetime_value(act_start_val)
            act_end = parse_datetime_value(act_end_val)
            if not act_start or not act_end:
                stats["skipped_missing_actual"] += 1
                logger.warning(
                    "Row %d: SKIPPED temp - missing Actual Start/End | ActStart=%r ActEnd=%r",
                    row_num,
                    act_start_val,
                    act_end_val,
                )
                continue

            if act_start.date() != target_date:
                stats["skipped_wrong_date"] += 1
                continue

            offset_days = offset_days_between(act_start, act_end)
            if (act_end.date() - act_start.date()).days < 0:
                stats["skipped_bad_offset"] += 1
                logger.warning(
                    "Row %d: SKIPPED temp - end before start | start=%s end=%s",
                    row_num,
                    act_start,
                    act_end,
                )
                continue

            raw_loc = row[col_loc] if col_loc < len(row) else None
            loc = fix_utf8_mojibake(raw_loc) if raw_loc is not None else None
            loc_str = str(loc).strip() if loc is not None else ""
            if not loc_str:
                stats["skipped_empty_location"] += 1
                logger.warning("Row %d: SKIPPED temp - empty Service Location Name", row_num)
                continue

            raw_addr = (
                row[col_addr] if col_addr != -1 and col_addr < len(row) else None
            )
            addr = fix_utf8_mojibake(raw_addr) if raw_addr is not None else None
            source_eircode = extract_eircode_from_address(addr)

            client_key = normalize_name_for_match(loc_str)
            client_id = resolve_person_id(
                clients_map.get(client_key) or [], source_eircode
            )
            if not client_id:
                stats["skipped_unknown_client"] += 1
                logger.warning(
                    "Row %d: SKIPPED temp - client not found | Location=%r eircode=%r",
                    row_num,
                    loc_str,
                    source_eircode or None,
                )
                continue

            start_minute = datetime_to_minutes(act_start)
            end_minute = datetime_to_minutes(act_end)
            duration = requested_duration_minutes(act_start, act_end, offset_days)
            stats["candidates"] += 1
            results.append(
                {
                    "row_num": row_num,
                    "client_id": int(client_id),
                    "client_name": loc_str,
                    "start_minute": start_minute,
                    "end_minute": end_minute,
                    "requested_start_time": format_time_str(act_start.time()),
                    "requested_end_time": format_time_str(act_end.time()),
                    "requested_duration": duration,
                    "end_time_date_offset_days": offset_days,
                    "day_of_week": get_day_of_week(target_date),
                }
            )
        return results, stats
    finally:
        wb.close()


def match_and_cancel_from_file(
    connection,
    visits: List[Dict[str, Any]],
    cancel_rows: List[Dict[str, Any]],
    type_ids: Dict[str, int],
) -> Tuple[int, int]:
    """
    Cancel visits matching file rows. Returns (cancelled_count, skipped_unmatched).
    """
    # Index active visits by (client_id, start_minute, end_minute)
    index: Dict[Tuple[int, int, int], List[Dict[str, Any]]] = {}
    for v in visits:
        if v["status"] not in ACTIVE_VISIT_STATUSES:
            continue
        key = (int(v["receiver_client_id"]), int(v["start_minute"]), int(v["end_minute"]))
        index.setdefault(key, []).append(v)

    cancelled = 0
    skipped = 0
    cursor = connection.cursor()
    try:
        for row in cancel_rows:
            type_id = type_ids.get(row["cancellation_name"])
            if type_id is None:
                logger.warning(
                    "Row %d: SKIPPED - cancellation type id missing for %r",
                    row["row_num"],
                    row["cancellation_name"],
                )
                skipped += 1
                continue
            key = (row["client_id"], row["start_minute"], row["end_minute"])
            matches = index.get(key) or []
            if not matches:
                skipped += 1
                logger.warning(
                    "Row %d: SKIPPED - no matching roster visit | client_id=%s start=%s end=%s cancel=%r",
                    row["row_num"],
                    row["client_id"],
                    row["start_minute"],
                    row["end_minute"],
                    row["cancellation_name"],
                )
                continue
            # Cancel all matching slots (e.g. multi-caregiver)
            still_active = [m for m in matches if m["status"] in ACTIVE_VISIT_STATUSES]
            if not still_active:
                logger.info(
                    "Row %d: visit(s) already cancelled | client_id=%s start=%s end=%s",
                    row["row_num"],
                    row["client_id"],
                    row["start_minute"],
                    row["end_minute"],
                )
                continue
            for visit in still_active:
                cursor.execute(
                    """
                    UPDATE roster_visit
                    SET status = 'CANCELLED',
                        cancellation_type_id = %s,
                        updated_at = NOW()
                    WHERE id = %s
                      AND status::text = ANY(%s)
                    """,
                    (type_id, visit["id"], list(ACTIVE_VISIT_STATUSES)),
                )
                if cursor.rowcount:
                    visit["status"] = "CANCELLED"
                    visit["cancellation_type_id"] = type_id
                    cancelled += 1
                    logger.info(
                        "Cancelled visit %s from file | client_id=%s type=%r",
                        visit["id"],
                        row["client_id"],
                        row["cancellation_name"],
                    )
        connection.commit()
    except Exception:
        connection.rollback()
        raise
    finally:
        cursor.close()
    return cancelled, skipped


def cancel_terminated_client_visits(
    connection,
    visits: List[Dict[str, Any]],
    terminated_ids: Set[int],
    terminated_type_id: int,
) -> int:
    cancelled = 0
    cursor = connection.cursor()
    try:
        for visit in visits:
            if visit["status"] not in ACTIVE_VISIT_STATUSES:
                continue
            client_id = int(visit["receiver_client_id"])
            if client_id not in terminated_ids:
                continue
            cursor.execute(
                """
                UPDATE roster_visit
                SET status = 'CANCELLED',
                    cancellation_type_id = %s,
                    updated_at = NOW()
                WHERE id = %s
                  AND status::text = ANY(%s)
                """,
                (terminated_type_id, visit["id"], list(ACTIVE_VISIT_STATUSES)),
            )
            if cursor.rowcount:
                visit["status"] = "CANCELLED"
                visit["cancellation_type_id"] = terminated_type_id
                cancelled += 1
                logger.info(
                    "Cancelled visit %s for terminated client_id=%s with type %r",
                    visit["id"],
                    client_id,
                    TERMINATED_CANCELLATION_TYPE,
                )
        connection.commit()
    except Exception:
        connection.rollback()
        raise
    finally:
        cursor.close()
    return cancelled


def find_personal_care_service_type_id(connection) -> Optional[int]:
    cursor = connection.cursor()
    try:
        cursor.execute(
            """
            SELECT id FROM service_type
            WHERE name = %s AND status = 'Active'
            ORDER BY id
            LIMIT 1
            """,
            (PERSONAL_CARE,),
        )
        row = cursor.fetchone()
        return int(row["id"]) if row else None
    finally:
        cursor.close()


def ensure_roster_for_date(connection, target_date: date) -> str:
    """Return roster id for date, inserting an empty roster if missing."""
    cursor = connection.cursor()
    try:
        cursor.execute("SELECT id FROM roster WHERE date = %s", (target_date,))
        row = cursor.fetchone()
        if row:
            return str(row["id"])
        roster_id = str(uuid.uuid4())
        cursor.execute(
            """
            INSERT INTO roster (id, date, seeded_client_schedule_ids, created_at, updated_at)
            VALUES (%s, %s, '{}', NOW(), NOW())
            ON CONFLICT (date) DO NOTHING
            RETURNING id
            """,
            (roster_id, target_date),
        )
        inserted = cursor.fetchone()
        if inserted:
            connection.commit()
            logger.info("Created roster %s for date %s", inserted["id"], target_date)
            return str(inserted["id"])
        cursor.execute("SELECT id FROM roster WHERE date = %s", (target_date,))
        existing = cursor.fetchone()
        if not existing:
            raise MigrationError(f"Failed to ensure roster for {target_date}")
        connection.commit()
        return str(existing["id"])
    except Exception:
        connection.rollback()
        raise
    finally:
        cursor.close()


def _find_existing_temp_schedule_id(
    cursor,
    client_id: int,
    target_date: date,
    requested_start: str,
    requested_end: str,
) -> Optional[int]:
    cursor.execute(
        """
        SELECT cs.id
        FROM client_schedules cs
        JOIN client_schedule_preferences csp ON csp.client_schedule_id = cs.id
        WHERE cs.client_id = %s
          AND cs.deleted_at IS NULL
          AND csp.is_temporary = true
          AND csp.effective_date_from = %s
          AND csp.effective_date_to = %s
          AND cs.requested_start_time = %s::time
          AND cs.requested_end_time = %s::time
        ORDER BY cs.id
        LIMIT 1
        """,
        (client_id, target_date, target_date, requested_start, requested_end),
    )
    row = cursor.fetchone()
    return int(row["id"]) if row else None


def _active_visit_exists(
    cursor,
    roster_id: str,
    client_id: int,
    start_minute: int,
    end_minute: int,
    client_schedule_id: Optional[int] = None,
) -> bool:
    if client_schedule_id is not None:
        cursor.execute(
            """
            SELECT id FROM roster_visit
            WHERE roster_id = %s
              AND client_schedule_id = %s
              AND COALESCE(slot_index, 0) = 0
              AND status::text = ANY(%s)
            LIMIT 1
            """,
            (roster_id, client_schedule_id, list(ACTIVE_VISIT_STATUSES)),
        )
        if cursor.fetchone():
            return True
    cursor.execute(
        """
        SELECT id FROM roster_visit
        WHERE roster_id = %s
          AND receiver_type = 'CLIENT'
          AND receiver_client_id = %s
          AND start_minute = %s
          AND end_minute = %s
          AND status::text = ANY(%s)
        LIMIT 1
        """,
        (
            roster_id,
            client_id,
            start_minute,
            end_minute,
            list(ACTIVE_VISIT_STATUSES),
        ),
    )
    return cursor.fetchone() is not None


def create_temp_schedules_and_visits(
    connection,
    roster_id: str,
    target_date: date,
    candidates: List[Dict[str, Any]],
    terminated_ids: Set[int],
    personal_care_service_type_id: Optional[int],
) -> Dict[str, int]:
    """
    Insert one-day temp client_schedules + UNALLOCATED roster_visits.
    Returns counts: created_schedules, created_visits, skipped_existing,
    skipped_terminated, skipped_errors.
    """
    counts = {
        "created_schedules": 0,
        "created_visits": 0,
        "skipped_existing": 0,
        "skipped_terminated": 0,
        "skipped_errors": 0,
    }
    if not candidates:
        return counts

    cursor = connection.cursor()
    try:
        cursor.execute(
            "SELECT typname FROM pg_type WHERE typname = 'client_schedules_days_enum'"
        )
        enum_exists = cursor.fetchone() is not None
        array_cast = (
            "::text[]::client_schedules_days_enum[]" if enum_exists else "::text[]"
        )
        date_str = format_date_str(target_date)

        for row in candidates:
            client_id = int(row["client_id"])
            if client_id in terminated_ids:
                counts["skipped_terminated"] += 1
                logger.info(
                    "Row %d: SKIPPED temp - client_id=%s is terminated",
                    row["row_num"],
                    client_id,
                )
                continue

            cursor.execute("SAVEPOINT temp_visit_row")
            try:
                existing_schedule_id = _find_existing_temp_schedule_id(
                    cursor,
                    client_id,
                    target_date,
                    row["requested_start_time"],
                    row["requested_end_time"],
                )
                schedule_id = existing_schedule_id
                if schedule_id is None:
                    cursor.execute(
                        f"""
                        INSERT INTO client_schedules (
                            client_id, days, requested_start_time, requested_end_time,
                            requested_duration, start_date, end_date, occurs_every,
                            number_of_care_givers, end_time_date_offset_days,
                            created_date, last_modified_date
                        ) VALUES (
                            %s, %s{array_cast}, %s, %s,
                            %s, %s, %s, %s,
                            %s, %s,
                            NOW(), NOW()
                        )
                        RETURNING id
                        """,
                        (
                            client_id,
                            [row["day_of_week"]],
                            row["requested_start_time"],
                            row["requested_end_time"],
                            row["requested_duration"],
                            date_str,
                            date_str,
                            1,
                            1,
                            row["end_time_date_offset_days"],
                        ),
                    )
                    schedule_id = int(cursor.fetchone()["id"])
                    cursor.execute(
                        """
                        INSERT INTO client_schedule_preferences (
                            client_schedule_id, window_start, window_end, min_duration,
                            suggested_duration,
                            is_temporary, effective_date_from, effective_date_to,
                            note, not_send_to_engine, is_unavailability, type_id,
                            created_date, last_modified_date
                        ) VALUES (
                            %s, %s, %s, %s,
                            %s,
                            true, %s, %s,
                            %s, true, false, NULL,
                            NOW(), NOW()
                        )
                        """,
                        (
                            schedule_id,
                            row["requested_start_time"],
                            row["requested_end_time"],
                            row["requested_duration"],
                            None,
                            date_str,
                            date_str,
                            TEMP_SCHEDULE_NOTE,
                        ),
                    )
                    counts["created_schedules"] += 1
                    logger.info(
                        "Row %d: created temp schedule id=%s client_id=%s %s-%s date=%s",
                        row["row_num"],
                        schedule_id,
                        client_id,
                        row["requested_start_time"],
                        row["requested_end_time"],
                        date_str,
                    )
                else:
                    logger.info(
                        "Row %d: reusing existing temp schedule id=%s client_id=%s",
                        row["row_num"],
                        schedule_id,
                        client_id,
                    )

                if _active_visit_exists(
                    cursor,
                    roster_id,
                    client_id,
                    row["start_minute"],
                    row["end_minute"],
                    client_schedule_id=schedule_id,
                ):
                    counts["skipped_existing"] += 1
                    logger.info(
                        "Row %d: SKIPPED visit - already exists | client_id=%s "
                        "start=%s end=%s schedule_id=%s",
                        row["row_num"],
                        client_id,
                        row["start_minute"],
                        row["end_minute"],
                        schedule_id,
                    )
                    cursor.execute("RELEASE SAVEPOINT temp_visit_row")
                    continue

                visit_id = str(uuid.uuid4())
                cursor.execute(
                    """
                    INSERT INTO roster_visit (
                        id, roster_id, receiver_type, receiver_client_id, receiver_user_id,
                        provider_user_id, client_schedule_id, slot_index,
                        start_minute, end_minute, status,
                        cancellation_type_id, cancellation_note, pinned,
                        created_at, updated_at
                    ) VALUES (
                        %s, %s, 'CLIENT', %s, NULL,
                        NULL, %s, 0,
                        %s, %s, 'UNALLOCATED',
                        NULL, NULL, false,
                        NOW(), NOW()
                    )
                    """,
                    (
                        visit_id,
                        roster_id,
                        client_id,
                        schedule_id,
                        row["start_minute"],
                        row["end_minute"],
                    ),
                )
                if personal_care_service_type_id is not None:
                    cursor.execute(
                        """
                        INSERT INTO roster_visit_service_types
                            (roster_visit_id, service_type_id)
                        VALUES (%s, %s)
                        ON CONFLICT DO NOTHING
                        """,
                        (visit_id, personal_care_service_type_id),
                    )
                counts["created_visits"] += 1
                logger.info(
                    "Row %d: created UNALLOCATED visit %s | client_id=%s "
                    "start=%s end=%s schedule_id=%s",
                    row["row_num"],
                    visit_id,
                    client_id,
                    row["start_minute"],
                    row["end_minute"],
                    schedule_id,
                )
                cursor.execute("RELEASE SAVEPOINT temp_visit_row")
            except Exception as e:
                cursor.execute("ROLLBACK TO SAVEPOINT temp_visit_row")
                counts["skipped_errors"] += 1
                logger.exception(
                    "Row %d: SKIPPED temp - error creating schedule/visit: %s",
                    row["row_num"],
                    e,
                )
                continue

        connection.commit()
    except Exception:
        connection.rollback()
        raise
    finally:
        cursor.close()
    return counts


def run(
    excel_path: Optional[str] = None,
    target_date: Optional[Any] = None,
    connection_manager=None,
    state=None,
) -> bool:
    print(
        """
    ╔════════════════════════════════════════════════════════════════╗
    ║   UPDATE TODAY VISITS                                          ║
    ║   Cancel visits + create temp schedules/visits from Actual     ║
    ╚════════════════════════════════════════════════════════════════╝
    """
    )
    if state and state.is_completed("update_today_visits"):
        logger.info("Update today visits already completed (resume).")
        return True

    from migration_support import get_assets_dir

    filepath = (
        Path(excel_path)
        if excel_path
        else get_assets_dir() / "updateTodayVisits" / "ClientHoursWithServiceType.xlsx"
    )
    if not filepath.exists():
        logger.error("Excel file not found: %s", filepath)
        return False

    if target_date is None:
        target_date = date.today()
    try:
        parsed_date = parse_target_date(target_date)
    except MigrationError as e:
        logger.error("%s", e)
        return False

    logger.info("Excel: %s | target date: %s", filepath, parsed_date)
    connection = None
    try:
        config = get_db_config()
        if connection_manager:
            connection = connection_manager.get_connection()
        else:
            connection = connect_to_database(config)

        clients_map = get_all_clients(connection)
        cancel_rows, extract_stats = extract_cancellation_rows(
            filepath, parsed_date, clients_map
        )
        logger.info(
            "Cancel extract: total=%s with_cancel=%s on_date=%s missing_dt=%s "
            "unknown_client=%s wrong_date=%s",
            extract_stats["total_rows"],
            extract_stats["with_cancellation"],
            extract_stats["on_target_date"],
            extract_stats["skipped_missing_datetime"],
            extract_stats["skipped_unknown_client"],
            extract_stats["skipped_wrong_date"],
        )

        temp_rows, temp_stats = extract_temp_visit_rows(
            filepath, parsed_date, clients_map
        )
        logger.info(
            "Temp extract: personal_care=%s req_both_empty=%s candidates=%s "
            "not_pc=%s has_cancel=%s req_filled=%s missing_actual=%s "
            "unknown_client=%s wrong_date=%s",
            temp_stats["personal_care"],
            temp_stats["req_both_empty"],
            temp_stats["candidates"],
            temp_stats["skipped_not_personal_care"],
            temp_stats["skipped_has_cancellation"],
            temp_stats["skipped_req_not_both_empty"],
            temp_stats["skipped_missing_actual"],
            temp_stats["skipped_unknown_client"],
            temp_stats["skipped_wrong_date"],
        )

        type_names = [r["cancellation_name"] for r in cancel_rows]
        type_names.append(TERMINATED_CANCELLATION_TYPE)
        type_ids = ensure_cancellation_types(connection, type_names)
        terminated_type_id = type_ids.get(TERMINATED_CANCELLATION_TYPE)
        if terminated_type_id is None:
            raise MigrationError("Failed to ensure Terminated cancellation type")

        terminated_ids = get_terminated_client_ids(connection, parsed_date)

        roster_id, visits = load_roster_visits_for_date(connection, parsed_date)
        file_cancelled = 0
        file_skipped = 0
        term_cancelled = 0

        if not roster_id and temp_rows:
            roster_id = ensure_roster_for_date(connection, parsed_date)
            visits = []
        elif not roster_id and cancel_rows:
            logger.warning(
                "No roster found for %s — nothing to cancel from file.",
                parsed_date,
            )
        elif not roster_id:
            logger.info("No roster found for %s and no temp candidates.", parsed_date)

        if roster_id and cancel_rows:
            logger.info("Roster %s has %d client visits", roster_id, len(visits))
            file_cancelled, file_skipped = match_and_cancel_from_file(
                connection, visits, cancel_rows, type_ids
            )

        if roster_id and visits:
            term_cancelled = cancel_terminated_client_visits(
                connection, visits, terminated_ids, terminated_type_id
            )

        temp_counts = {
            "created_schedules": 0,
            "created_visits": 0,
            "skipped_existing": 0,
            "skipped_terminated": 0,
            "skipped_errors": 0,
        }
        if temp_rows:
            if not roster_id:
                roster_id = ensure_roster_for_date(connection, parsed_date)
            pc_type_id = find_personal_care_service_type_id(connection)
            if pc_type_id is None:
                logger.warning(
                    "Active service_type %r not found — visits will have no service types",
                    PERSONAL_CARE,
                )
            temp_counts = create_temp_schedules_and_visits(
                connection,
                roster_id,
                parsed_date,
                temp_rows,
                terminated_ids,
                pc_type_id,
            )

        print("\n" + "=" * 60)
        print("✓ UPDATE TODAY VISITS COMPLETED")
        print("=" * 60)
        print(f"  Target date: {parsed_date}")
        print(f"  Cancellation rows on date: {len(cancel_rows)}")
        print(f"  Cancelled from file: {file_cancelled}")
        print(f"  Skipped unmatched file rows: {file_skipped}")
        print(f"  Terminated clients: {len(terminated_ids)}")
        print(f"  Cancelled for terminated: {term_cancelled}")
        print(f"  Temp candidates: {len(temp_rows)}")
        print(f"  Temp schedules created: {temp_counts['created_schedules']}")
        print(f"  Temp visits created: {temp_counts['created_visits']}")
        print(f"  Temp skipped existing: {temp_counts['skipped_existing']}")
        print(f"  Temp skipped terminated: {temp_counts['skipped_terminated']}")
        print(f"  Temp skipped errors: {temp_counts['skipped_errors']}")

        if state:
            state.clear_step("update_today_visits")
        return True
    except (psycopg2.OperationalError, psycopg2.InterfaceError) as e:
        if ConnectionLostError:
            raise ConnectionLostError("update_today_visits", {}) from e
        raise
    except MigrationError as e:
        logger.error("Migration error: %s", e)
        return False
    except Exception as e:
        logger.exception("Unexpected error: %s", e)
        return False
    finally:
        if connection and not connection_manager:
            connection.close()


if __name__ == "__main__":
    excel_arg = sys.argv[1] if len(sys.argv) > 1 else None
    date_arg = sys.argv[2] if len(sys.argv) > 2 else None
    success = run(excel_path=excel_arg, target_date=date_arg)
    sys.exit(0 if success else 1)
