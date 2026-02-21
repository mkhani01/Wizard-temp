"""
User Availability Migration Script
===================================
Migrates user availability data from Excel file to database.

Rules:
- "Core" records: Recurring availability (Start Date - 28 days, occurs_every=4).
- Other Types (found in seeded availability_types): Treated as availability or unavailability
  based on the type's classification in DB (search Type in availability_types).
  - is_unavailability = from type (availability vs unavailability)
  - is_temp = True
  - No recurrence: start_date/end_date/occurs_every = None
  - effective_date_from = effective_date_to = that specific date only
  - Handles multi-day and overnight shifts by splitting into daily records.
"""

import os
import sys
import logging
from pathlib import Path
from datetime import datetime, timedelta, date, time
from collections import defaultdict
from typing import Optional, Dict, List, Tuple, Any

try:
    import psycopg2
    from psycopg2.extras import RealDictCursor, execute_values
    import openpyxl
except ImportError as e:
    print(f"Missing required package: {e}")
    print("Please install: pip install psycopg2-binary openpyxl")
    sys.exit(1)


# Configure comprehensive logging
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - [%(filename)s:%(lineno)d] - %(message)s',
    handlers=[
        logging.FileHandler('migration_availability.log', mode='w'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


# Excel column indices (0-based)
COLUMN_CARE_ASSISTANT_NAME = 0
COLUMN_FRANCHISE = 1
COLUMN_TEAM = 2
COLUMN_CARE_ASSISTANT_TYPE = 3
COLUMN_GRADE_TYPE = 4
COLUMN_START_DATE = 5
COLUMN_START_TIME = 6
COLUMN_END_DATE = 7
COLUMN_END_TIME = 8
COLUMN_HOURS = 9
COLUMN_TYPE = 10
COLUMN_NOTES = 11

DAYS_OF_WEEK = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
TITLES = ['Mr', 'Mrs', 'Miss', 'Ms', 'Dr', 'Prof', 'Mr.', 'Mrs.', 'Miss.', 'Ms.', 'Dr.', 'Prof.']


class MigrationError(Exception):
    """Custom exception for migration errors"""
    pass


def get_db_config() -> Dict[str, Any]:
    """Get database configuration from environment variables"""
    config = {
        'host': os.getenv('DB_HOST', 'localhost'),
        'port': int(os.getenv('DB_PORT', '5432')),
        'database': os.getenv('DB_NAME'),
        'user': os.getenv('DB_USER'),
        'password': os.getenv('DB_PASSWORD')
    }
    
    missing = [k for k, v in config.items() if not v]
    if missing:
        raise MigrationError(f"Missing database configuration: {missing}")
    
    return config


def connect_to_database(config: Dict[str, Any]):
    """Connect to PostgreSQL database"""
    try:
        logger.info(f"Connecting to PostgreSQL at {config['host']}:{config['port']}/{config['database']}...")
        connection = psycopg2.connect(
            host=config['host'],
            port=config['port'],
            database=config['database'],
            user=config['user'],
            password=config['password'],
            cursor_factory=RealDictCursor,
            connect_timeout=10,
        )
        connection.autocommit = False
        logger.info("✓ Database connection established successfully")
        return connection
    except Exception as e:
        logger.error(f"✗ Failed to connect to database: {e}")
        raise MigrationError(f"Database connection failed: {e}")


def get_all_users(connection) -> Dict[str, int]:
    """Get all users from database mapped by 'name lastname' (lowercase)"""
    cursor = connection.cursor()
    try:
        cursor.execute('SELECT id, name, lastname FROM "user" WHERE deleted_at IS NULL')
        users = {}
        for row in cursor.fetchall():
            name = (row['name'] or '').strip()
            lastname = (row['lastname'] or '').strip()
            key = f"{name} {lastname}".strip().lower()
            if key:
                users[key] = row['id']
        
        logger.info(f"✓ Loaded {len(users)} users from database")
        return users
    finally:
        cursor.close()


def get_availability_types(connection) -> Dict[str, Dict]:
    """
    Get all availability types from database.
    Returns dict: { 'name_lower': { 'id': int, 'is_unavailability': bool } }
    """
    cursor = connection.cursor()
    try:
        cursor.execute("SELECT id, name, type FROM availability_types WHERE deleted_at IS NULL")
        types_map = {}
        for row in cursor.fetchall():
            name_lower = (row['name'] or '').strip().lower()
            is_unavailability = (row['type'] or '').strip().lower() == 'unavailability'
            types_map[name_lower] = {
                'id': row['id'],
                'is_unavailability': is_unavailability
            }
        
        if 'core' not in types_map:
            raise MigrationError("CRITICAL: 'Core' availability type not found in database.")
        
        logger.info(f"✓ Loaded {len(types_map)} availability types from database")
        return types_map
    finally:
        cursor.close()


def strip_title(full_name: str) -> str:
    """Strip title from name and return cleaned name"""
    if not full_name:
        return ''
    
    parts = full_name.strip().split()
    if not parts:
        return ''
    
    while parts and parts[0] in TITLES:
        parts = parts[1:]
    
    return ' '.join(parts)


def parse_time_value(time_val) -> Optional[time]:
    """Parse time from various formats"""
    if time_val is None:
        return None
    
    if isinstance(time_val, time):
        return time_val
    
    if isinstance(time_val, datetime):
        return time_val.time()
    
    if isinstance(time_val, str):
        try:
            parts = time_val.split(':')
            if len(parts) >= 2:
                hours = int(parts[0])
                minutes = int(parts[1])
                seconds = int(parts[2]) if len(parts) > 2 else 0
                return time(hours, minutes, seconds)
        except (ValueError, IndexError):
            pass
    
    return None


def parse_date_value(date_val) -> Optional[date]:
    """Parse date from various formats"""
    if date_val is None:
        return None
    
    if isinstance(date_val, datetime):
        return date_val.date()
    
    if isinstance(date_val, date):
        return date_val
    
    if isinstance(date_val, str):
        date_str = date_val.strip()
        for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y']:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
    
    return None


def get_day_of_week(date_obj: date) -> str:
    """Get day of week name from date"""
    return DAYS_OF_WEEK[date_obj.weekday()]


def format_time_str(t: time) -> str:
    """Format time as HH:MM:SS string"""
    return t.strftime('%H:%M:%S')


def format_date_str(d: date) -> str:
    """Format date as YYYY-MM-DD string"""
    return d.strftime('%Y-%m-%d')


def process_xlsx_file(filepath: Path, users_map: Dict[str, int], types_map: Dict[str, Dict]) -> Tuple[List[Dict], List[str]]:
    """
    Process the Excel file and extract availability records.
    """
    logger.info(f"\n{'='*60}")
    logger.info(f"PROCESSING EXCEL FILE: {filepath}")
    logger.info(f"{'='*60}")
    
    if not filepath.exists():
        raise MigrationError(f"Excel file not found: {filepath}")
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        raise MigrationError(f"Failed to load Excel file: {e}")
    
    sheet_name = 'Care Assistant Availability'
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
        logger.warning(f"Default sheet not found. Using: {sheet_name}")
    
    ws = wb[sheet_name]
    
    valid_records = []
    unmatched_users = set()
    skipped_types = defaultdict(int)
    total_rows = 0
    
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        total_rows += 1
        
        care_assistant_name = row[COLUMN_CARE_ASSISTANT_NAME] if len(row) > COLUMN_CARE_ASSISTANT_NAME else None
        start_date_val = row[COLUMN_START_DATE] if len(row) > COLUMN_START_DATE else None
        end_date_val = row[COLUMN_END_DATE] if len(row) > COLUMN_END_DATE else None
        start_time_val = row[COLUMN_START_TIME] if len(row) > COLUMN_START_TIME else None
        end_time_val = row[COLUMN_END_TIME] if len(row) > COLUMN_END_TIME else None
        type_val = row[COLUMN_TYPE] if len(row) > COLUMN_TYPE else None
        notes = row[COLUMN_NOTES] if len(row) > COLUMN_NOTES else None
        
        if not care_assistant_name:
            logger.warning(
                "Row %d: SKIPPED - empty Care Assistant name | Start=%r, End=%r, Type=%r, StartTime=%r, EndTime=%r",
                row_num, start_date_val, end_date_val, type_val, start_time_val, end_time_val
            )
            continue

        type_str = str(type_val).strip() if type_val else ''
        type_key = type_str.lower()

        # Check if type is known (not in seeded availability_types) — log row info, do not seed
        if type_key not in types_map:
            skipped_types[type_str] += 1
            logger.warning(
                f"Row {row_num}: Skipped (Type not in DB) — Care Assistant: {care_assistant_name!r}, "
                f"Type: {type_str!r}, Start: {start_date_val}, End: {end_date_val}, "
                f"StartTime: {start_time_val}, EndTime: {end_time_val}, Notes: {notes!r}"
            )
            continue
        
        # Match User
        name_without_title = strip_title(str(care_assistant_name))
        user_key = name_without_title.strip().lower()
        user_id = users_map.get(user_key)
        
        if not user_id:
            unmatched_users.add(str(care_assistant_name))
            logger.warning(
                "Row %d: SKIPPED - user not found | Care Assistant=%r, Type=%r, Start=%r, End=%r, StartTime=%r, EndTime=%r",
                row_num, care_assistant_name, type_str, start_date_val, end_date_val, start_time_val, end_time_val
            )
            continue

        # Parse dates/times
        start_date = parse_date_value(start_date_val)
        end_date = parse_date_value(end_date_val)
        start_time = parse_time_value(start_time_val)
        end_time = parse_time_value(end_time_val)

        if not start_date or not start_time or not end_time:
            logger.warning(
                "Row %d: SKIPPED - missing date/time | Care Assistant=%r, Type=%r, Start=%r, End=%r, StartTime=%r, EndTime=%r",
                row_num, care_assistant_name, type_str, start_date_val, end_date_val, start_time_val, end_time_val
            )
            continue
        
        # If end_date is missing, assume same day
        if not end_date:
            end_date = start_date

        type_info = types_map[type_key]
        is_core = (type_key == 'core')

        valid_records.append({
            'user_id': user_id,
            'user_name': care_assistant_name,
            'start_date': start_date,
            'end_date': end_date,
            'start_time': start_time,
            'end_time': end_time,
            'type_id': type_info['id'],
            'is_unavailability': type_info['is_unavailability'],
            'is_core': is_core,
            'notes': notes,
            'source_row': row_num,
        })
        logger.info(
            "Row %d: ADDED | user=%r (id=%s), type=%r, core=%s, start=%s end=%s, start_time=%s end_time=%s",
            row_num, care_assistant_name, user_id, type_str, is_core,
            start_date, end_date, start_time, end_time
        )

    logger.info(f"\n{'='*60}")
    logger.info("EXCEL PROCESSING SUMMARY")
    logger.info(f"{'='*60}")
    logger.info(f"Total rows: {total_rows}")
    logger.info(f"Valid records extracted: {len(valid_records)}")
    logger.info(f"Skipped types (not in DB): {dict(skipped_types)}")
    logger.info(f"Unmatched users: {len(unmatched_users)}")
    
    return valid_records, list(unmatched_users)


def generate_availability_records(records: List[Dict]) -> List[Dict]:
    """
    Generate database records.
    Splits shifts into daily chunks if they span multiple days or overnight.
    """
    logger.info(f"\n{'='*60}")
    logger.info("GENERATING AVAILABILITY RECORDS")
    logger.info(f"{'='*60}")
    
    availabilities = []
    
    for rec in records:
        user_id = rec['user_id']
        start_date = rec['start_date']
        end_date = rec['end_date']
        start_time = rec['start_time']
        end_time = rec['end_time']
        type_id = rec['type_id']
        is_unavailability = rec['is_unavailability']
        is_core = rec['is_core']
        
        # --- CORE LOGIC ---
        if is_core:
            # Rule: Start Date - 28 days = recurring start
            target_start_date = start_date - timedelta(days=28)
            day_of_week = get_day_of_week(target_start_date)
            
            # Handle overnight core shifts (rare, but possible)
            # Note: Core logic in prompt implies single day block, but we should handle splits safely
            
            if end_time <= start_time:
                # Overnight core shift (e.g., Mon 22:00 - Tue 06:00)
                # Part 1
                availabilities.append({
                    'user_id': user_id,
                    'days': [day_of_week],
                    'start_time': format_time_str(start_time),
                    'end_time': '23:59:59',
                    'is_temp': False,
                    'is_unavailability': is_unavailability,
                    'type_id': type_id,
                    'start_date': format_date_str(target_start_date),
                    'end_date': None,
                    'occurs_every': 4,
                    'effective_date_from': None,
                    'effective_date_to': None,
                    'note': rec['notes']
                })
                # Part 2 (Next day)
                next_day = target_start_date + timedelta(days=1)
                availabilities.append({
                    'user_id': user_id,
                    'days': [get_day_of_week(next_day)],
                    'start_time': '00:00:00',
                    'end_time': format_time_str(end_time),
                    'is_temp': False,
                    'is_unavailability': is_unavailability,
                    'type_id': type_id,
                    'start_date': format_date_str(next_day),
                    'end_date': None,
                    'occurs_every': 4,
                    'effective_date_from': None,
                    'effective_date_to': None,
                    'note': rec['notes']
                })
            else:
                # Standard core shift
                availabilities.append({
                    'user_id': user_id,
                    'days': [day_of_week],
                    'start_time': format_time_str(start_time),
                    'end_time': format_time_str(end_time),
                    'is_temp': False,
                    'is_unavailability': is_unavailability,
                    'type_id': type_id,
                    'start_date': format_date_str(target_start_date),
                    'end_date': None,
                    'occurs_every': 4,
                    'effective_date_from': None,
                    'effective_date_to': None,
                    'note': rec['notes']
                })

        # --- OTHER TYPES (availability or unavailability from seeded type) ---
        else:
            # Look up Type in availability_types: use its classification (availability vs unavailability).
            # is_temp = True, no recurrence; effective_date_from/to = that specific date only.
            # Split multi-day into individual days to satisfy 'days' enum and validation.
            current_date = start_date

            while current_date <= end_date:
                is_first_day = (current_date == start_date)
                is_last_day = (current_date == end_date)

                # Calculate chunk times
                chunk_start = start_time if is_first_day else time(0, 0, 0)
                chunk_end = end_time if is_last_day else time(23, 59, 59)

                # Adjust for exact overnight edge case (e.g. ends at 00:00 next day)
                if is_last_day and end_time == time(0, 0, 0):
                    # 00:00 means end of previous day, so this day is empty or full?
                    # If it's the ONLY day (start==end), 00:00 to 00:00 is invalid, treat as full day?
                    # Or ignore. Let's assume full day for 00:00-00:00 single day.
                    if start_date == end_date:
                        chunk_start = time(0,0,0)
                        chunk_end = time(23,59,59)
                    else:
                        # It's a subsequent day that ends at 00:00, meaning it ended at midnight.
                        # No record needed for this day.
                        current_date += timedelta(days=1)
                        continue

                # Skip if start >= end (invalid chunk, e.g. generated from logic above)
                if chunk_start >= chunk_end:
                    current_date += timedelta(days=1)
                    continue

                availabilities.append({
                    'user_id': user_id,
                    'days': [get_day_of_week(current_date)],
                    'start_time': format_time_str(chunk_start),
                    'end_time': format_time_str(chunk_end),
                    'is_temp': True,
                    'is_unavailability': is_unavailability,  # from seeded type (availability vs unavailability)
                    'type_id': type_id,
                    'start_date': None,
                    'end_date': None,
                    'occurs_every': None,
                    'effective_date_from': format_date_str(current_date),
                    'effective_date_to': format_date_str(current_date),
                    'note': rec['notes']
                })

                current_date += timedelta(days=1)
    
    logger.info(f"Generated {len(availabilities)} availability slots")
    return availabilities


def deduplicate_availabilities(availabilities: List[Dict]) -> List[Dict]:
    """Remove duplicates based on unique constraints"""
    unique_map = {}
    duplicates = 0
    
    for avail in availabilities:
        # Key based on unique constraints
        if avail['is_temp']:
            key = (
                avail['user_id'],
                avail['type_id'],
                avail['effective_date_from'],
                avail['start_time'],
                avail['end_time']
            )
        else:
            key = (
                avail['user_id'],
                tuple(avail['days']),
                avail['start_time'],
                avail['end_time'],
                avail['start_date'],
            )
        
        if key in unique_map:
            duplicates += 1
        else:
            unique_map[key] = avail
    
    logger.info(f"Deduplication: {len(availabilities)} -> {len(unique_map)} records ({duplicates} duplicates removed)")
    return list(unique_map.values())


def clear_user_availabilities(connection) -> None:
    """Remove all existing rows from user_availabilities before seeding."""
    cursor = connection.cursor()
    try:
        cursor.execute("DELETE FROM user_availabilities")
        deleted = cursor.rowcount
        connection.commit()
        logger.info(f"Cleared user_availabilities: {deleted} existing row(s) removed")
    except Exception as e:
        connection.rollback()
        logger.error(f"Failed to clear user_availabilities: {e}")
        raise MigrationError(f"Failed to clear user_availabilities: {e}")
    finally:
        cursor.close()


def seed_availabilities(connection, availabilities: List[Dict]) -> int:
    """Insert user availabilities into database"""
    logger.info(f"\n{'='*60}")
    logger.info("SEEDING DATABASE")
    logger.info(f"{'='*60}")
    
    clear_user_availabilities(connection)
    
    if not availabilities:
        logger.warning("No availabilities to insert")
        return 0
    
    cursor = connection.cursor()
    try:
        # Check enum
        cursor.execute("SELECT typname FROM pg_type WHERE typname = 'user_availabilities_days_enum'")
        enum_exists = cursor.fetchone()
        array_cast = '::text[]::user_availabilities_days_enum[]' if enum_exists else '::text[]'
        
        insert_query = f"""
            INSERT INTO user_availabilities (
                user_id, days, start_time, end_time, 
                is_temp, is_unavailability, type_id,
                start_date, end_date, occurs_every, 
                effective_date_from, effective_date_to,
                note, created_date, last_modified_date
            ) VALUES %s
            RETURNING id
        """
        
        availability_tuples = [
            (
                avail['user_id'],
                avail['days'],
                avail['start_time'],
                avail['end_time'],
                avail['is_temp'],
                avail['is_unavailability'],
                avail['type_id'],
                avail['start_date'],
                avail['end_date'],
                avail['occurs_every'],
                avail['effective_date_from'],
                avail['effective_date_to'],
                avail['note'],
            )
            for avail in availabilities
        ]
        
        logger.info(f"Inserting {len(availability_tuples)} records...")
        
        execute_values(
            cursor,
            insert_query,
            availability_tuples,
            template=f"(%s, %s{array_cast}, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW())"
        )
        
        inserted = cursor.fetchall()
        connection.commit()
        logger.info("SEEDED %d availability records (see Excel row ADDED/SKIPPED logs above for which rows produced them)", len(inserted))
        return len(inserted)
        
    except Exception as e:
        connection.rollback()
        logger.error(f"✗ Failed to insert availabilities: {e}")
        raise MigrationError(f"Database insertion failed: {e}")
    finally:
        cursor.close()


def generate_summary_report(
    records: List[Dict],
    availabilities: List[Dict],
    unmatched_users: List[str],
    inserted_count: int
) -> str:
    """Generate a summary report of the migration"""
    report = []
    report.append("\n" + "="*70)
    report.append("MIGRATION SUMMARY REPORT")
    report.append("="*70)
    
    report.append(f"\n📊 INPUT DATA:")
    report.append(f"   - Valid records extracted: {len(records)}")
    
    core_count = sum(1 for r in records if r['is_core'])
    other_count = len(records) - core_count
    report.append(f"   - Core Availabilities: {core_count}")
    report.append(f"   - Other types (temp, by seeded type): {other_count}")
    report.append(f"   - Unmatched users: {len(unmatched_users)}")
    
    if unmatched_users:
        report.append(f"\n⚠️  UNMATCHED USERS ({len(unmatched_users)}):")
        for user in sorted(unmatched_users)[:10]: # Show first 10
            report.append(f"   - {user}")
        if len(unmatched_users) > 10:
            report.append(f"   ... and {len(unmatched_users) - 10} more")
    
    report.append(f"\n📋 GENERATED RECORDS:")
    report.append(f"   - Total slots to insert: {len(availabilities)}")
    temp_avail = sum(1 for a in availabilities if a.get('is_temp') and not a.get('is_unavailability'))
    temp_unavail = sum(1 for a in availabilities if a.get('is_temp') and a.get('is_unavailability'))
    report.append(f"   - Temp availability: {temp_avail}")
    report.append(f"   - Temp unavailability: {temp_unavail}")

    report.append(f"\n✅ DATABASE:")
    report.append(f"   - Records inserted: {inserted_count}")
    
    report.append("\n" + "="*70)
    return "\n".join(report)


def run(xlsx_path: Optional[str] = None) -> bool:
    """Main execution function"""
    print("""
    ╔══════════════════════════════════════════════════════════════╗
    ║   USER AVAILABILITY MIGRATION                                ║
    ║   - Core: Recurring (Start Date - 28d)                       ║
    ║   - Others: Temp availability/unavailability by type (Exact Date) ║
    ╚══════════════════════════════════════════════════════════════╝
    """)
    
    if xlsx_path:
        filepath = Path(xlsx_path)
    else:
        # Fallback logic if needed
        filepath = Path('userAvailabilities.xlsx') 
    
    logger.info(f"Excel file path: {filepath}")
    
    connection = None
    
    try:
        # Step 1: Database connection
        config = get_db_config()
        connection = connect_to_database(config)
        
        # Step 2: Load reference data
        users_map = get_all_users(connection)
        types_map = get_availability_types(connection)
        
        # Step 3: Process Excel file
        records, unmatched_users = process_xlsx_file(filepath, users_map, types_map)
        
        if not records:
            logger.error("No valid records found in Excel file")
            return False
        
        # Step 4: Generate availability records
        availabilities = generate_availability_records(records)
        availabilities = deduplicate_availabilities(availabilities)
        
        # Step 5: Seed database
        inserted_count = seed_availabilities(connection, availabilities)
        
        # Generate summary report
        report = generate_summary_report(records, availabilities, unmatched_users, inserted_count)
        print(report)
        
        print("\n" + "="*60)
        print("✓ MIGRATION COMPLETED SUCCESSFULLY")
        print("="*60)
        
        return True
        
    except MigrationError as e:
        logger.error(f"Migration error: {e}")
        print(f"\n✗ MIGRATION FAILED: {e}")
        return False
        
    except Exception as e:
        logger.exception(f"Unexpected error: {e}")
        print(f"\n✗ MIGRATION FAILED: {e}")
        return False
        
    finally:
        if connection:
            connection.close()
            logger.info("\nDatabase connection closed")


if __name__ == "__main__":
    xlsx_arg = sys.argv[1] if len(sys.argv) > 1 else None
    success = run(xlsx_arg)
    sys.exit(0 if success else 1)