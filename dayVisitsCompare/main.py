"""
Compare DB roster_visit CSV export with Client Hours with Service Type XLSX.

Match key (same as cancel migration): (client_id, start_minute, end_minute)

Usage:
  python3 -m dayVisitsCompare.main <db_visits.csv> <client_hours.xlsx> [YYYY-MM-DD]
  python3 dayVisitsCompare/main.py <db_visits.csv> <client_hours.xlsx> [YYYY-MM-DD]
  python3 -m dayVisitsCompare.main ... --clients-csv clients.csv

Resolves Service Location Name → client_id via DB (.env) or optional --clients-csv
(columns: id, name, lastname).

Excel filter: only rows where both
  "Planned Service Type Description" and
  "Planned Service Requirement Type Description"
equal "Personal Care".
"""

from __future__ import annotations

import argparse
import csv
import os
import sys
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# Allow running as script: python3 dayVisitsCompare/main.py ...
_PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

from dotenv import load_dotenv

load_dotenv(_PROJECT_ROOT / ".env")

from encoding_utils import fix_utf8_mojibake, normalize_name_for_match
from person_match_utils import flatten_candidate_ids
from updateTodayVisitsMigration.main import (
    MigrationError,
    _col_idx,
    connect_to_database,
    datetime_to_minutes,
    get_all_clients,
    get_db_config,
    parse_target_date,
    resolve_start_end,
)

MatchKey = Tuple[int, int, int]


PERSONAL_CARE = "Personal Care"


def minutes_to_hhmm(m: int) -> str:
    return f"{m // 60:02d}:{m % 60:02d}"


def load_db_visits_csv(path: Path, target_date: Optional[date]) -> Tuple[date, List[Dict[str, Any]]]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            raise MigrationError(f"Empty or invalid CSV: {path}")
        required = {"receiver_client_id", "start_minute", "end_minute"}
        missing = required - {h.strip() for h in reader.fieldnames}
        if missing:
            raise MigrationError(f"CSV missing columns: {sorted(missing)}")

        rows: List[Dict[str, Any]] = []
        dates_seen: Counter = Counter()
        for i, raw in enumerate(reader, start=2):
            date_raw = (raw.get("date") or "").strip()
            row_date = None
            if date_raw:
                try:
                    row_date = parse_target_date(date_raw[:10])
                    dates_seen[row_date] += 1
                except MigrationError:
                    pass

            if target_date and row_date and row_date != target_date:
                continue

            try:
                client_id = int(str(raw["receiver_client_id"]).strip())
                start_minute = int(str(raw["start_minute"]).strip())
                end_minute = int(str(raw["end_minute"]).strip())
            except (TypeError, ValueError) as e:
                raise MigrationError(f"CSV row {i}: invalid client/start/end: {e}") from e

            rows.append(
                {
                    "source": "db",
                    "row_num": i,
                    "id": (raw.get("id") or "").strip(),
                    "client_id": client_id,
                    "start_minute": start_minute,
                    "end_minute": end_minute,
                    "status": (raw.get("status") or "").strip(),
                    "slot_index": (raw.get("slot_index") or "").strip(),
                    "date": row_date,
                }
            )

    if not rows:
        raise MigrationError("No DB visit rows to compare (after date filter)")

    if target_date is None:
        if len(dates_seen) == 1:
            target_date = next(iter(dates_seen))
        elif dates_seen:
            raise MigrationError(
                f"CSV has multiple dates {sorted(str(d) for d in dates_seen)}; pass YYYY-MM-DD"
            )
        else:
            raise MigrationError("CSV has no date column values; pass YYYY-MM-DD")

    return target_date, rows


def load_excel_visits(
    path: Path,
    target_date: date,
    clients_map: Dict[str, int],
) -> Tuple[List[Dict[str, Any]], Dict[str, int]]:
    import openpyxl

    stats = {
        "total_rows": 0,
        "on_target_date": 0,
        "personal_care": 0,
        "skipped_not_personal_care": 0,
        "skipped_missing_datetime": 0,
        "skipped_unknown_client": 0,
        "skipped_wrong_date": 0,
        "skipped_empty_location": 0,
    }
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if "Data" not in wb.sheetnames:
            raise MigrationError("Sheet 'Data' not found in workbook")
        ws = wb["Data"]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            raise MigrationError("Workbook has no header row")

        col_loc = _col_idx(header, ["Service Location Name"])
        col_req_start = _col_idx(header, ["Service Requirement Start Date And Time"])
        col_req_end = _col_idx(header, ["Service Requirement End Date And Time"])
        col_act_start = _col_idx(header, ["Actual Start Date And Time"])
        col_act_end = _col_idx(header, ["Actual End Date And Time"])
        col_cancel = _col_idx(header, ["Cancellation Description"])
        col_count = _col_idx(header, ["Count"])
        col_svc_type = _col_idx(header, ["Planned Service Type Description"])
        col_svc_req = _col_idx(header, ["Planned Service Requirement Type Description"])

        if col_loc == -1:
            raise MigrationError("Missing required column: Service Location Name")
        if col_svc_type == -1:
            raise MigrationError("Missing required column: Planned Service Type Description")
        if col_svc_req == -1:
            raise MigrationError(
                "Missing required column: Planned Service Requirement Type Description"
            )
        if col_req_start == -1 and col_act_start == -1:
            raise MigrationError("Missing start datetime columns")
        if col_req_end == -1 and col_act_end == -1:
            raise MigrationError("Missing end datetime columns")

        results: List[Dict[str, Any]] = []
        for row_num, row in enumerate(rows_iter, start=2):
            stats["total_rows"] += 1
            if not row:
                continue

            svc_type = _cell_str(row, col_svc_type)
            svc_req = _cell_str(row, col_svc_req)
            if svc_type != PERSONAL_CARE or svc_req != PERSONAL_CARE:
                stats["skipped_not_personal_care"] += 1
                continue

            raw_loc = row[col_loc] if col_loc < len(row) else None
            loc = fix_utf8_mojibake(raw_loc) if raw_loc is not None else None
            loc_str = str(loc).strip() if loc is not None else ""
            if not loc_str:
                stats["skipped_empty_location"] += 1
                continue

            req_start = row[col_req_start] if col_req_start != -1 and col_req_start < len(row) else None
            req_end = row[col_req_end] if col_req_end != -1 and col_req_end < len(row) else None
            act_start = row[col_act_start] if col_act_start != -1 and col_act_start < len(row) else None
            act_end = row[col_act_end] if col_act_end != -1 and col_act_end < len(row) else None
            start_dt, end_dt, source = resolve_start_end(req_start, req_end, act_start, act_end)
            if not start_dt or not end_dt:
                stats["skipped_missing_datetime"] += 1
                continue

            if start_dt.date() != target_date:
                stats["skipped_wrong_date"] += 1
                continue
            stats["on_target_date"] += 1
            stats["personal_care"] += 1

            client_key = normalize_name_for_match(loc_str)
            client_id = clients_map.get(client_key)
            if not client_id:
                stats["skipped_unknown_client"] += 1
                results.append(
                    {
                        "source": "excel",
                        "row_num": row_num,
                        "client_id": None,
                        "client_name": loc_str,
                        "start_minute": datetime_to_minutes(start_dt),
                        "end_minute": datetime_to_minutes(end_dt),
                        "time_source": source,
                        "cancellation": _cell_str(row, col_cancel),
                        "service_type": svc_type,
                        "service_requirement_type": svc_req,
                        "count": _cell_int(row, col_count, default=1),
                        "unresolved": True,
                    }
                )
                continue

            cancel = _cell_str(row, col_cancel)
            results.append(
                {
                    "source": "excel",
                    "row_num": row_num,
                    "client_id": client_id,
                    "client_name": loc_str,
                    "start_minute": datetime_to_minutes(start_dt),
                    "end_minute": datetime_to_minutes(end_dt),
                    "time_source": source,
                    "cancellation": cancel,
                    "service_type": svc_type,
                    "service_requirement_type": svc_req,
                    "count": _cell_int(row, col_count, default=1),
                    "unresolved": False,
                }
            )
        return results, stats
    finally:
        wb.close()


def _cell_str(row, col: int) -> str:
    if col == -1 or col >= len(row) or row[col] is None:
        return ""
    return str(row[col]).strip()


def _cell_int(row, col: int, default: int = 1) -> int:
    if col == -1 or col >= len(row) or row[col] is None:
        return default
    try:
        return max(1, int(row[col]))
    except (TypeError, ValueError):
        return default


def expand_excel_for_match(excel_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Expand Count>1 into multiple match slots (multi-caregiver)."""
    expanded: List[Dict[str, Any]] = []
    for row in excel_rows:
        if row.get("unresolved"):
            expanded.append(row)
            continue
        for slot in range(int(row.get("count") or 1)):
            item = dict(row)
            item["slot"] = slot
            expanded.append(item)
    return expanded


def compare(
    db_rows: List[Dict[str, Any]],
    excel_rows: List[Dict[str, Any]],
) -> Dict[str, Any]:
    db_by_key: Dict[MatchKey, List[Dict[str, Any]]] = defaultdict(list)
    for r in db_rows:
        key = (int(r["client_id"]), int(r["start_minute"]), int(r["end_minute"]))
        db_by_key[key].append(r)

    excel_resolved = [r for r in excel_rows if not r.get("unresolved")]
    excel_unresolved = [r for r in excel_rows if r.get("unresolved")]

    excel_by_key: Dict[MatchKey, List[Dict[str, Any]]] = defaultdict(list)
    for r in excel_resolved:
        key = (int(r["client_id"]), int(r["start_minute"]), int(r["end_minute"]))
        excel_by_key[key].append(r)

    only_db: List[Dict[str, Any]] = []
    only_excel: List[Dict[str, Any]] = []
    matched = 0

    all_keys = set(db_by_key) | set(excel_by_key)
    for key in sorted(all_keys):
        db_list = db_by_key.get(key, [])
        ex_list = excel_by_key.get(key, [])
        n_match = min(len(db_list), len(ex_list))
        matched += n_match
        only_db.extend(db_list[n_match:])
        only_excel.extend(ex_list[n_match:])

    return {
        "matched": matched,
        "only_db": only_db,
        "only_excel": only_excel,
        "unresolved_excel": excel_unresolved,
    }


def _print_row_db(r: Dict[str, Any], id_to_name: Dict[int, str]) -> None:
    cid = int(r["client_id"])
    name = id_to_name.get(cid, "")
    print(
        f"  DB  id={r.get('id','')} client_id={cid} name={name!r} "
        f"{minutes_to_hhmm(r['start_minute'])}-{minutes_to_hhmm(r['end_minute'])} "
        f"status={r.get('status','')} slot={r.get('slot_index','')}"
    )


def _print_row_excel(r: Dict[str, Any]) -> None:
    cid = r.get("client_id")
    print(
        f"  XLSX row={r['row_num']} client_id={cid} name={r.get('client_name')!r} "
        f"{minutes_to_hhmm(r['start_minute'])}-{minutes_to_hhmm(r['end_minute'])} "
        f"cancel={(r.get('cancellation') or '-')!r} svc={(r.get('service_type') or '-')!r}"
    )


def build_id_to_name(clients_map: Dict[str, int], excel_rows: List[Dict[str, Any]]) -> Dict[int, str]:
    id_to_name: Dict[int, str] = {}
    for r in excel_rows:
        cid = r.get("client_id")
        name = r.get("client_name")
        if cid and name and cid not in id_to_name:
            id_to_name[int(cid)] = name
    for key, cid in clients_map.items():
        if cid not in id_to_name and "," in key:
            id_to_name[cid] = key
    return id_to_name


def load_clients_map_from_csv(path: Path) -> Dict[str, int]:
    """Build name→id map from a clients export (id, name, lastname)."""
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            raise MigrationError(f"Empty clients CSV: {path}")
        field_map = {h.strip().lower(): h for h in reader.fieldnames}
        id_col = field_map.get("id")
        name_col = field_map.get("name") or field_map.get("first_name") or field_map.get("firstname")
        last_col = field_map.get("lastname") or field_map.get("last_name")
        if not id_col or not name_col or not last_col:
            raise MigrationError(
                "Clients CSV needs columns: id, name, lastname "
                f"(got: {reader.fieldnames})"
            )
        clients: Dict[str, int] = {}
        for raw in reader:
            try:
                client_id = int(str(raw[id_col]).strip())
            except (TypeError, ValueError):
                continue
            name = (raw.get(name_col) or "").strip()
            lastname = (raw.get(last_col) or "").strip()
            key_comma = normalize_name_for_match(f"{lastname}, {name}")
            key_space = normalize_name_for_match(f"{name} {lastname}")
            if key_comma:
                if key_comma not in clients or client_id > clients[key_comma]:
                    clients[key_comma] = client_id
            if key_space:
                if key_space not in clients or client_id > clients[key_space]:
                    clients[key_space] = client_id
        if not clients:
            raise MigrationError(f"No clients loaded from {path}")
        return clients


def load_clients_map(clients_csv: Optional[Path]) -> Dict[str, int]:
    if clients_csv is not None:
        if not clients_csv.exists():
            raise MigrationError(f"Clients CSV not found: {clients_csv}")
        return load_clients_map_from_csv(clients_csv)

    # Same defaults as wizard when .env has no DB_* keys
    os.environ.setdefault("DB_HOST", "localhost")
    os.environ.setdefault("DB_PORT", "5432")
    os.environ.setdefault("DB_NAME", "appDB")
    os.environ.setdefault("DB_USER", "root")
    os.environ.setdefault("DB_PASSWORD", "root")

    try:
        connection = connect_to_database(get_db_config())
    except Exception as e:
        raise MigrationError(
            f"Could not connect to DB for client name→id mapping: {e}\n"
            "Set DB_* in .env, or pass --clients-csv (id,name,lastname)."
        ) from e
    try:
        return flatten_candidate_ids(get_all_clients(connection))
    finally:
        connection.close()


def run(
    db_csv: Path,
    excel_path: Path,
    target_date: Optional[str] = None,
    clients_csv: Optional[Path] = None,
) -> int:
    parsed_date = parse_target_date(target_date) if target_date else None
    parsed_date, db_rows = load_db_visits_csv(db_csv, parsed_date)
    clients_map = load_clients_map(clients_csv)

    print(f"Comparing visits for {parsed_date}")
    print(f"  DB CSV : {db_csv} ({len(db_rows)} row(s))")
    print(f"  Excel  : {excel_path}")
    print(
        f"  Clients: {len(clients_map)} name key(s) "
        f"({'CSV ' + str(clients_csv) if clients_csv else 'from DB'})"
    )

    excel_raw, excel_stats = load_excel_visits(excel_path, parsed_date, clients_map)
    excel_rows = expand_excel_for_match(excel_raw)

    print(
        f"  Excel Personal Care on date: {excel_stats['personal_care']} "
        f"(skipped non-PC: {excel_stats['skipped_not_personal_care']}, "
        f"unresolved clients: {excel_stats['skipped_unknown_client']}, "
        f"expanded slots: {len(excel_rows)})"
    )

    result = compare(db_rows, excel_rows)
    id_to_name = build_id_to_name(clients_map, excel_raw)

    print()
    print("=== Summary ===")
    print(f"  Matched (client_id + start + end): {result['matched']}")
    print(f"  Only in DB CSV                   : {len(result['only_db'])}")
    print(f"  Only in Excel                    : {len(result['only_excel'])}")
    print(f"  Excel unresolved (name→id fail)  : {len(result['unresolved_excel'])}")

    if result["only_db"]:
        print()
        print(f"=== Only in DB CSV ({len(result['only_db'])}) ===")
        for r in result["only_db"]:
            _print_row_db(r, id_to_name)

    if result["only_excel"]:
        print()
        print(f"=== Only in Excel ({len(result['only_excel'])}) ===")
        for r in result["only_excel"]:
            _print_row_excel(r)

    if result["unresolved_excel"]:
        print()
        print(f"=== Excel rows with unknown client ({len(result['unresolved_excel'])}) ===")
        for r in result["unresolved_excel"]:
            _print_row_excel(r)

    return 0


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        prog="dayVisitsCompare",
        description="Compare DB roster visits CSV vs Client Hours XLSX for a date",
    )
    parser.add_argument("db_csv", type=Path, help="DB roster_visit export CSV")
    parser.add_argument("excel", type=Path, help="Client Hours with Service Type XLSX")
    parser.add_argument(
        "date",
        nargs="?",
        default=None,
        help="Target date YYYY-MM-DD (optional if CSV has a single date)",
    )
    parser.add_argument(
        "--clients-csv",
        type=Path,
        default=None,
        help="Optional clients export (id,name,lastname) instead of reading DB",
    )
    args = parser.parse_args(argv)

    if not args.db_csv.exists():
        print(f"DB CSV not found: {args.db_csv}", file=sys.stderr)
        return 1
    if not args.excel.exists():
        print(f"Excel not found: {args.excel}", file=sys.stderr)
        return 1

    try:
        return run(args.db_csv, args.excel, args.date, args.clients_csv)
    except MigrationError as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
